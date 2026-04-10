# -*- coding: utf-8 -*-
"""
Enhanced version V2 of process_I15_report() with:
- Historical tracking
- Protection against multiple runs per day
- Force update option
"""

import pandas as pd
import os
import sqlite3
import re
from datetime import datetime, timedelta


def normalize_nvkt(x):
    """Chuẩn hóa tên NVKT_DB - giữ phần sau dấu '-'"""
    if not isinstance(x, str):
        return x
    if '-' in x:
        x = x.split('-')[1].strip()
    x = re.sub(r'\([^)]*\)', '', x).strip()
    return x


def add_tt_column(df):
    """Thêm cột TT (Thứ tự) vào đầu DataFrame, đánh số từ 1"""
    if len(df) == 0:
        return df
    df_copy = df.copy()
    # Xóa cột TT cũ nếu đã tồn tại
    if 'TT' in df_copy.columns:
        df_copy = df_copy.drop(columns=['TT'])
    df_copy.insert(0, 'TT', range(1, len(df_copy) + 1))
    return df_copy


def format_excel_detail(file_path, df, sheet_name=None):
    """
    Định dạng file Excel chi tiết NVKT:
    - Kẻ bảng với borders
    - Điều chỉnh độ rộng cột tự động
    - Wrap text
    - Tô màu cột SA cho các giá trị trùng lặp
    
    Args:
        sheet_name: Tên sheet cần format. Nếu None thì format sheet active.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Định nghĩa border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header style
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Các màu để tô SA trùng lặp
        sa_colors = [
            'FFFF99',  # Vàng nhạt
            'FFCCCC',  # Hồng nhạt
            'CCFFCC',  # Xanh lá nhạt
            'CCCCFF',  # Tím nhạt
            'FFCC99',  # Cam nhạt
            'CCFFFF',  # Cyan nhạt
            'FF99CC',  # Hồng đậm hơn
            '99CCFF',  # Xanh dương nhạt
        ]
        
        # Tìm cột SA
        sa_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'SA':
                sa_col_idx = idx
                break
        
        # Đếm các giá trị SA và tìm các giá trị trùng lặp
        sa_duplicates = {}
        if sa_col_idx and len(df) > 0 and 'SA' in df.columns:
            sa_counts = df['SA'].value_counts()
            duplicate_sas = sa_counts[sa_counts > 1].index.tolist()
            for i, sa_val in enumerate(duplicate_sas):
                sa_duplicates[sa_val] = sa_colors[i % len(sa_colors)]
        
        # Áp dụng định dạng cho từng ô
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                                     min_col=1, max_col=ws.max_column), 1):
            for col_idx, cell in enumerate(row, 1):
                # Border cho tất cả các ô
                cell.border = thin_border
                # Alignment wrap text và center vertical
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Header row
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                else:
                    # Tô màu cột SA nếu là giá trị trùng lặp
                    if col_idx == sa_col_idx and cell.value in sa_duplicates:
                        color = sa_duplicates[cell.value]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Điều chỉnh độ rộng cột tự động
        column_widths = {
            'TT': 5,
            'MA_TB': 18,
            'TEN_TB_ONE': 25,
            'DIACHI_ONE': 35,
            'DT_ONE': 12,
            'NGAY_SUYHAO': 12,
            'THIETBI': 25,
            'SA': 28,
            'KETCUOI': 18,
        }
        
        for col_idx, cell in enumerate(ws[1], 1):
            col_letter = get_column_letter(col_idx)
            col_name = cell.value
            if col_name in column_widths:
                ws.column_dimensions[col_letter].width = column_widths[col_name]
            else:
                ws.column_dimensions[col_letter].width = 15
        
        # Đặt chiều cao hàng header
        ws.row_dimensions[1].height = 25
        
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"⚠️ Lỗi format Excel: {e}")
        return False


def create_k2_threshold_report(df, output_dir):
    """
    Tạo file I1.5_k2_theo_nguong_report.xlsx với 3 sheet lọc theo ngưỡng
    Chỉ số OLT RX và Chỉ số ONU RX.
    """
    try:
        print("\n" + "="*80)
        print("TẠO BÁO CÁO K2 THEO NGƯỠNG")
        print("="*80)

        olt_col = 'Chỉ số OLT RX'
        onu_col = 'Chỉ số ONU RX'

        if olt_col not in df.columns or onu_col not in df.columns:
            print(f"⚠️ Không tìm thấy cột '{olt_col}' hoặc '{onu_col}', bỏ qua tạo báo cáo theo ngưỡng")
            return False

        # Chuyển sang kiểu số
        df_work = df.copy()
        df_work[olt_col] = pd.to_numeric(df_work[olt_col], errors='coerce')
        df_work[onu_col] = pd.to_numeric(df_work[onu_col], errors='coerce')

        # Các cột cần xuất (DOI_ONE, NVKT_DB_NORMALIZED đứng đầu, sau STT)
        output_columns = ['DOI_ONE', 'NVKT_DB_NORMALIZED',
                          'ACCOUNT_CTS', 'TEN_TB_ONE', 'DIACHI_ONE', 'DT_ONE',
                          'NGAY_SUYHAO', 'THIETBI', 'SA', 'KETCUOI',
                          olt_col, onu_col]
        output_columns = [c for c in output_columns if c in df_work.columns]

        # Định nghĩa 3 ngưỡng
        thresholds = [
            {
                'sheet': 'k2_25_26',
                'low': -26,
                'high': -25,
                'low_inclusive': True,
                'high_inclusive': True,
            },
            {
                'sheet': 'k2_26_26.5',
                'low': -26.5,
                'high': -26,
                'low_inclusive': True,
                'high_inclusive': False,
            },
            {
                'sheet': 'k2_26.5_27',
                'low': -27,
                'high': -26.5,
                'low_inclusive': True,
                'high_inclusive': False,
            },
        ]

        output_file = os.path.join(output_dir, 'I1.5_k2_theo_nguong_report.xlsx')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for t in thresholds:
                low, high = t['low'], t['high']

                # Xây dựng điều kiện cho từng cột
                if t['low_inclusive'] and t['high_inclusive']:
                    olt_mask = (df_work[olt_col] >= low) & (df_work[olt_col] <= high)
                    onu_mask = (df_work[onu_col] >= low) & (df_work[onu_col] <= high)
                elif t['low_inclusive'] and not t['high_inclusive']:
                    olt_mask = (df_work[olt_col] >= low) & (df_work[olt_col] < high)
                    onu_mask = (df_work[onu_col] >= low) & (df_work[onu_col] < high)
                else:
                    olt_mask = (df_work[olt_col] > low) & (df_work[olt_col] <= high)
                    onu_mask = (df_work[onu_col] > low) & (df_work[onu_col] <= high)

                # OR: 1 trong 2 chỉ số nằm trong khoảng
                df_filtered = df_work[olt_mask | onu_mask][output_columns].copy()
                df_filtered = df_filtered.sort_values(by=['DOI_ONE', 'NVKT_DB_NORMALIZED']).reset_index(drop=True)

                # Thêm cột STT
                df_filtered.insert(0, 'STT', range(1, len(df_filtered) + 1))

                df_filtered.to_excel(writer, sheet_name=t['sheet'], index=False)
                print(f"  ✅ Sheet '{t['sheet']}': {len(df_filtered)} bản ghi")

        print(f"\n✅ Đã tạo file: {output_file}")
        return True

    except Exception as e:
        print(f"❌ Lỗi khi tạo báo cáo K2 theo ngưỡng: {e}")
        import traceback
        traceback.print_exc()
        return False


def process_I15_report_with_tracking(force_update=False, report_date=None):
    """Wrapper for K1 report"""
    input_file = os.path.join("downloads", "baocao_hanoi", "I1.5 report.xlsx")
    return _process_I15_generic_with_tracking(input_file, k_suffix="K1", history_db="suy_hao_history.db", force_update=force_update, report_date=report_date)


def process_I15_k2_report_with_tracking(force_update=False, report_date=None):
    """Wrapper for K2 report"""
    input_file = os.path.join("downloads", "baocao_hanoi", "I1.5_k2 report.xlsx")
    return _process_I15_generic_with_tracking(input_file, k_suffix="K2", history_db="suy_hao_history_k2.db", force_update=force_update, report_date=report_date)


def _process_I15_generic_with_tracking(input_file, k_suffix="K1", history_db="suy_hao_history.db", force_update=False, report_date=None):
    """
    Xử lý báo cáo I1.5 (K1 hoặc K2) với tracking lịch sử:
    1. Đọc file đầu vào
    2. Tra cứu thông tin từ danhba.db
    3. Chuẩn hóa cột NVKT_DB
    4. Kiểm tra đã xử lý ngày này chưa
    5. So sánh với dữ liệu ngày hôm qua
    6. Tạo các sheet: TH_SHC_I15, Tang_moi, Giam_het, Van_con, Bien_dong_tong_hop
    7. Lưu vào database để tracking lịch sử (Database riêng cho K1/K2)

    Args:
        input_file (str): Đường dẫn file Excel đầu vào
        k_suffix (str): Hậu tố K1 hoặc K2 để đặt tên
        history_db (str): Tên file database history tương ứng
        force_update (bool): Nếu True, cho phép ghi đè dữ liệu đã tồn tại trong ngày
        report_date (str): Ngày báo cáo thủ công (YYYY-MM-DD), nếu None sẽ đọc từ file
    """
    try:
        print("\n" + "="*80)
        print(f"BẮT ĐẦU XỬ LÝ BÁO CÁO I1.5 {k_suffix} (VỚI TRACKING LỊCH SỬ V2)")
        print(f"Database: {history_db}")
        print("="*80)

        # Đường dẫn file
        db_file = "danhba.db"
        history_db = history_db # Dùng giá trị truyền vào

        if not os.path.exists(input_file):
            print(f"❌ Không tìm thấy file: {input_file}")
            return False

        print(f"\n✓ Đang đọc file: {input_file}")

        # Đọc file Excel
        df = pd.read_excel(input_file)
        print(f"✅ Đã đọc file, tổng số dòng: {len(df)}, tổng số cột: {df.shape[1]}")

        # Lấy ngày báo cáo - ưu tiên: ngày thủ công > ngày hiện tại > ngày trong file
        if report_date is not None:
            try:
                # Kiểm tra định dạng ngày thủ công
                pd.to_datetime(report_date, format='%Y-%m-%d')
                print(f"✓ Sử dụng ngày báo cáo thủ công: {report_date}")
            except Exception as e:
                print(f"❌ Định dạng ngày thủ công không hợp lệ (cần YYYY-MM-DD): {report_date}")
                return False
        else:
            report_date = datetime.now().strftime('%Y-%m-%d')
            print(f"✓ Sử dụng ngày hiện tại: {report_date}")

        # Cập nhật lại cột NGAY_SUYHAO trong DataFrame nếu dùng ngày thủ công
        df['NGAY_SUYHAO'] = pd.to_datetime(report_date).strftime('%d/%m/%Y')

        # Tra cứu thông tin từ danhba.db
        print("\n✓ Đang tra cứu thông tin từ danhba.db...")
        if os.path.exists(db_file):
            try:
                conn = sqlite3.connect(db_file)
                # Thêm DOI_VT và NVKT vào query để cập nhật DOI_ONE và NVKT_DB
                query = "SELECT MA_TB, THIETBI, SA, KETCUOI, DOI_VT, NVKT FROM danhba"
                df_danhba = pd.read_sql_query(query, conn)

                print(f"✅ Đã đọc {len(df_danhba)} bản ghi từ danhba.db")

                if 'ACCOUNT_CTS' in df.columns:
                    # Xóa các cột cũ nếu tồn tại
                    cols_to_remove = ['MA_TB', 'THIETBI', 'SA', 'KETCUOI', 'DOI_VT_DB', 'NVKT_DB_NEW']
                    for col in cols_to_remove:
                        if col in df.columns:
                            df = df.drop(columns=[col])

                    # Merge để lấy thông tin từ danhba.db
                    df = df.merge(df_danhba, left_on='ACCOUNT_CTS', right_on='MA_TB', how='left')
                    if 'MA_TB' in df.columns:
                        df = df.drop(columns=['MA_TB'])
                    print(f"✅ Đã tra cứu và thêm các cột: THIETBI, SA, KETCUOI")
                    
                    # Cập nhật DOI_ONE và NVKT_DB từ danhba.db (nếu có dữ liệu)
                    updated_doi = 0
                    updated_nvkt = 0
                    
                    if 'DOI_VT' in df.columns and 'DOI_ONE' in df.columns:
                        # Đếm số bản ghi sẽ được cập nhật
                        mask_doi = df['DOI_VT'].notna() & (df['DOI_VT'] != df['DOI_ONE'])
                        updated_doi = mask_doi.sum()
                        # Cập nhật DOI_ONE từ DOI_VT (danhba.db)
                        df.loc[df['DOI_VT'].notna(), 'DOI_ONE'] = df.loc[df['DOI_VT'].notna(), 'DOI_VT']
                        df = df.drop(columns=['DOI_VT'])
                    
                    if 'NVKT' in df.columns and 'NVKT_DB' in df.columns:
                        # Đếm số bản ghi sẽ được cập nhật
                        mask_nvkt = df['NVKT'].notna() & (df['NVKT'] != df['NVKT_DB'])
                        updated_nvkt = mask_nvkt.sum()
                        # Cập nhật NVKT_DB từ NVKT (danhba.db)
                        df.loc[df['NVKT'].notna(), 'NVKT_DB'] = df.loc[df['NVKT'].notna(), 'NVKT']
                        df = df.drop(columns=['NVKT'])
                    
                    if updated_doi > 0 or updated_nvkt > 0:
                        print(f"✅ Đã cập nhật từ danhba.db: {updated_doi} DOI_ONE, {updated_nvkt} NVKT_DB")
                
                # Đọc bảng thong_ke (cho NVKT)
                df_thong_ke = pd.read_sql_query(
                    "SELECT DOI_VT, NVKT, so_thue_bao_pon_qly FROM thong_ke", conn)
                print(f"✅ Đã đọc {len(df_thong_ke)} bản ghi từ bảng thong_ke")
                
                # Đọc bảng thong_ke_theo_don_vi (cho đơn vị)
                df_thong_ke_dv = pd.read_sql_query(
                    "SELECT don_vi, so_thue_bao_pon_qly FROM thong_ke_theo_don_vi", conn)
                print(f"✅ Đã đọc {len(df_thong_ke_dv)} bản ghi từ bảng thong_ke_theo_don_vi")
                
                conn.close()
            except Exception as e:
                print(f"⚠️ Lỗi khi tra cứu danhba.db: {e}")
                df_thong_ke = pd.DataFrame()
                df_thong_ke_dv = pd.DataFrame()
        else:
            df_thong_ke = pd.DataFrame()
            df_thong_ke_dv = pd.DataFrame()
            print(f"⚠️ Không tìm thấy file {db_file}")

        # Chuẩn hóa cột NVKT_DB
        print("\n✓ Đang chuẩn hóa cột NVKT_DB...")
        if 'NVKT_DB' in df.columns:
            df['NVKT_DB_NORMALIZED'] = df['NVKT_DB'].apply(normalize_nvkt)
            print("✅ Đã chuẩn hóa cột NVKT_DB")
        else:
            print("⚠️ Không tìm thấy cột NVKT_DB")
            df['NVKT_DB_NORMALIZED'] = None

        # ==================================================================
        # SO SÁNH VỚI NGÀY HÔM QUA VÀ LƯU VÀO DATABASE
        # ==================================================================
        print("\n" + "="*80)
        print("TRACKING LỊCH SỬ VÀ SO SÁNH VỚI NGÀY HÔM QUA")
        print("="*80)

        # Khởi tạo biến
        df_tang_moi = pd.DataFrame()
        df_giam_het = pd.DataFrame()
        df_van_con = pd.DataFrame()
        df_bien_dong = pd.DataFrame()
        should_save_to_db = True

        if not os.path.exists(history_db):
            print(f"⚠️ Không tìm thấy {history_db}, bỏ qua tracking lịch sử")
            should_save_to_db = False
        else:
            hist_conn = sqlite3.connect(history_db)
            cursor = hist_conn.cursor()

            # KIỂM TRA ĐÃ XỬ LÝ NGÀY NÀY CHƯA
            cursor.execute("SELECT COUNT(*) FROM suy_hao_snapshots WHERE ngay_bao_cao = ?", (report_date,))
            existing_count = cursor.fetchone()[0]

            if existing_count > 0 and not force_update:
                print(f"\n⚠️  ĐÃ CÓ DỮ LIỆU NGÀY {report_date} TRONG DATABASE ({existing_count} bản ghi)")
                print(f"⚠️  BỎ QUA lưu database để tránh trùng lặp và sai số liệu")
                print(f"✓  Chỉ xử lý và tạo file Excel output")
                print(f"\nℹ️  Gợi ý:")
                print(f"   - Nếu muốn tải lại: Xóa dữ liệu ngày {report_date} trong DB trước")
                print(f"   - Hoặc chạy với tham số: process_I15_report_with_tracking(force_update=True)")

                # Đọc dữ liệu từ database thay vì tính lại
                print(f"\n✓ Đang đọc dữ liệu biến động từ database...")
                df_tang_moi = pd.read_sql_query(f"""
                    SELECT * FROM suy_hao_daily_changes
                    WHERE ngay_bao_cao = '{report_date}' AND loai_bien_dong = 'TANG_MOI'
                """, hist_conn)

                df_giam_het = pd.read_sql_query(f"""
                    SELECT * FROM suy_hao_daily_changes
                    WHERE ngay_bao_cao = '{report_date}' AND loai_bien_dong = 'GIAM_HET'
                """, hist_conn)

                df_van_con = pd.read_sql_query(f"""
                    SELECT * FROM suy_hao_daily_changes
                    WHERE ngay_bao_cao = '{report_date}' AND loai_bien_dong = 'VAN_CON'
                """, hist_conn)

                df_bien_dong = pd.read_sql_query(f"""
                    SELECT
                        doi_one as "Đơn vị",
                        nvkt_db_normalized as "NVKT_DB",
                        tong_so_hien_tai as "Tổng số hiện tại",
                        so_tang_moi as "Tăng mới",
                        so_giam_het as "Giảm/Hết",
                        so_van_con as "Vẫn còn"
                    FROM suy_hao_daily_summary
                    WHERE ngay_bao_cao = '{report_date}'
                    ORDER BY doi_one, nvkt_db_normalized
                """, hist_conn)

                print(f"✅ Đã đọc dữ liệu biến động từ DB:")
                print(f"   - TĂNG MỚI: {len(df_tang_moi)} thuê bao")
                print(f"   - GIẢM/HẾT: {len(df_giam_het)} thuê bao")
                print(f"   - VẪN CÒN: {len(df_van_con)} thuê bao")

                # Chuẩn hóa tên cột để khớp với logic tạo Excel
                if len(df_tang_moi) > 0:
                    df_tang_moi = df_tang_moi.rename(columns={
                        'account_cts': 'ACCOUNT_CTS',
                        'ten_tb_one': 'TEN_TB_ONE',
                        'dt_onediachi_one': 'DT_ONEDIACHI_ONE',
                        'doi_one': 'DOI_ONE',
                        'nvkt_db_normalized': 'NVKT_DB_NORMALIZED',
                        'sa': 'SA',
                        'olt_cts': 'OLT_CTS',
                        'port_cts': 'PORT_CTS',
                        'thietbi': 'THIETBI',
                        'ketcuoi': 'KETCUOI'
                    })

                if len(df_van_con) > 0:
                    df_van_con = df_van_con.rename(columns={
                        'account_cts': 'ACCOUNT_CTS',
                        'ten_tb_one': 'TEN_TB_ONE',
                        'dt_onediachi_one': 'DT_ONEDIACHI_ONE',
                        'doi_one': 'DOI_ONE',
                        'nvkt_db_normalized': 'NVKT_DB_NORMALIZED',
                        'sa': 'SA',
                        'olt_cts': 'OLT_CTS',
                        'port_cts': 'PORT_CTS',
                        'thietbi': 'THIETBI',
                        'ketcuoi': 'KETCUOI'
                    })

                hist_conn.close()
                should_save_to_db = False

            elif existing_count > 0 and force_update:
                print(f"\n⚠️  ĐÃ CÓ DỮ LIỆU NGÀY {report_date} ({existing_count} bản ghi)")
                print(f"✓  FORCE_UPDATE=True → Sẽ ghi đè dữ liệu cũ")
                should_save_to_db = True
                # Tiếp tục xử lý bình thường

        # Nếu cần lưu vào DB (lần đầu hoặc force_update)
        if should_save_to_db and os.path.exists(history_db):
            hist_conn = sqlite3.connect(history_db)
            cursor = hist_conn.cursor()

            # Tính ngày hôm qua
            yesterday = (datetime.strptime(report_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')

            # Đọc dữ liệu ngày hôm qua
            print(f"\n✓ Đang đọc dữ liệu ngày {yesterday}...")
            df_yesterday = pd.read_sql_query(f"""
                SELECT account_cts FROM suy_hao_snapshots
                WHERE ngay_bao_cao = '{yesterday}'
            """, hist_conn)

            print(f"  Ngày {yesterday}: {len(df_yesterday)} thuê bao")
            print(f"  Ngày {report_date}: {len(df)} thuê bao")

            # Phân loại (loại bỏ NaN/None)
            # Hỗ trợ cả ACCOUNT_CTS (từ file gốc) và MA_TB (từ i15_cts_converter)
            account_col = 'ACCOUNT_CTS' if 'ACCOUNT_CTS' in df.columns else 'MA_TB'
            if account_col in df.columns:
                accounts_today = set([x for x in df[account_col].tolist() if pd.notna(x) and str(x).strip() != ''])
            else:
                accounts_today = set()

            if len(df_yesterday) > 0:
                accounts_yesterday = set([x for x in df_yesterday['account_cts'].tolist() if pd.notna(x) and str(x).strip() != ''])
            else:
                accounts_yesterday = set()

            tang_moi_set = accounts_today - accounts_yesterday
            giam_het_set = accounts_yesterday - accounts_today
            van_con_set = accounts_today & accounts_yesterday

            print(f"\n✓ Phân tích biến động:")
            print(f"  🆕 TĂNG MỚI: {len(tang_moi_set)} thuê bao")
            print(f"  ⬇️  GIẢM/HẾT: {len(giam_het_set)} thuê bao")
            print(f"  ↔️  VẪN CÒN: {len(van_con_set)} thuê bao")

            # Tạo DataFrame cho từng loại
            df_tang_moi = df[df[account_col].isin(tang_moi_set)].copy() if len(tang_moi_set) > 0 else pd.DataFrame()
            df_van_con = df[df[account_col].isin(van_con_set)].copy() if len(van_con_set) > 0 else pd.DataFrame()

            # Lấy thông tin GIẢM/HẾT từ database
            if len(giam_het_set) > 0:
                accounts_str = ','.join([f"'{x}'" for x in list(giam_het_set)[:1000]])
                df_giam_het = pd.read_sql_query(f"""
                    SELECT s.*, t.so_ngay_lien_tuc
                    FROM suy_hao_snapshots s
                    LEFT JOIN suy_hao_tracking t ON s.account_cts = t.account_cts
                    WHERE s.ngay_bao_cao = '{yesterday}'
                      AND s.account_cts IN ({accounts_str})
                """, hist_conn)
            else:
                df_giam_het = pd.DataFrame()

            # Thêm số ngày liên tục cho VẪN CÒN
            if len(van_con_set) > 0 and len(df_van_con) > 0:
                print("\n✓ Đang lấy số ngày liên tục cho thuê bao VẪN CÒN...")
                tracking_data = pd.read_sql_query(f"""
                    SELECT account_cts, so_ngay_lien_tuc
                    FROM suy_hao_tracking
                    WHERE account_cts IN ({','.join([f"'{x}'" for x in list(van_con_set)[:1000]])})
                """, hist_conn)

                df_van_con = df_van_con.merge(
                    tracking_data,
                    left_on=account_col,
                    right_on='account_cts',
                    how='left'
                )
                if 'account_cts' in df_van_con.columns:
                    df_van_con = df_van_con.drop(columns=['account_cts'])

                df_van_con['so_ngay_lien_tuc'] = df_van_con['so_ngay_lien_tuc'].fillna(1) + 1
            else:
                if len(df_van_con) > 0:
                    df_van_con['so_ngay_lien_tuc'] = 2

            # Lưu snapshot hôm nay vào database
            print(f"\n✓ Đang lưu snapshot ngày {report_date} vào database...")

            # Xóa dữ liệu cũ nếu có
            cursor.execute("DELETE FROM suy_hao_snapshots WHERE ngay_bao_cao = ?", (report_date,))

            inserted = 0
            skipped = 0
            for idx, row in df.iterrows():
                account = row.get(account_col)
                if pd.isna(account) or account is None or str(account).strip() == '':
                    skipped += 1
                    continue

                try:
                    cursor.execute("""
                        INSERT INTO suy_hao_snapshots (
                            ngay_bao_cao, account_cts, ten_tb_one, dt_onediachi_one,
                            doi_one, nvkt_db, nvkt_db_normalized, sa,
                            olt_cts, port_cts, thietbi, ketcuoi, trangthai_tb
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        report_date, account,
                        row.get('TEN_TB_ONE'), row.get('DT_ONEDIACHI_ONE'),
                        row.get('DOI_ONE'), row.get('NVKT_DB'), row.get('NVKT_DB_NORMALIZED'),
                        row.get('SA'), row.get('OLT_CTS'), row.get('PORT_CTS'),
                        row.get('THIETBI'), row.get('KETCUOI'), row.get('TRANGTHAI_TB')
                    ))
                    inserted += 1
                except Exception as e:
                    print(f"  ⚠️  Lỗi insert account '{account}': {e}")
                    skipped += 1

            if skipped > 0:
                print(f"  ✅ Đã lưu {inserted} bản ghi vào snapshots (bỏ qua {skipped} dòng)")
            else:
                print(f"  ✅ Đã lưu {inserted} bản ghi vào snapshots")

            # Cập nhật tracking table
            print(f"\n✓ Đang cập nhật bảng tracking...")

            for account in tang_moi_set:
                df_filtered = df[df[account_col] == account]
                if len(df_filtered) > 0:
                    row_data = df_filtered.iloc[0]
                    cursor.execute("""
                        INSERT OR REPLACE INTO suy_hao_tracking (
                            account_cts, ngay_xuat_hien_dau_tien, ngay_thay_cuoi_cung,
                            so_ngay_lien_tuc, doi_one, nvkt_db, sa, trang_thai
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        account, report_date, report_date, 1,
                        row_data.get('DOI_ONE'), row_data.get('NVKT_DB_NORMALIZED'),
                        row_data.get('SA'), 'DANG_SUY_HAO'
                    ))

            for account in van_con_set:
                cursor.execute("""
                    UPDATE suy_hao_tracking
                    SET ngay_thay_cuoi_cung = ?,
                        so_ngay_lien_tuc = so_ngay_lien_tuc + 1,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE account_cts = ?
                """, (report_date, account))

            for account in giam_het_set:
                cursor.execute("""
                    UPDATE suy_hao_tracking
                    SET trang_thai = 'DA_HET_SUY_HAO',
                        updated_at = CURRENT_TIMESTAMP
                    WHERE account_cts = ?
                """, (account,))

            # Lưu daily changes
            print(f"\n✓ Đang lưu daily changes...")
            cursor.execute("DELETE FROM suy_hao_daily_changes WHERE ngay_bao_cao = ?", (report_date,))

            def save_changes(df_changes, loai):
                for _, row in df_changes.iterrows():
                    so_ngay = row.get('so_ngay_lien_tuc', 1) if loai != 'TANG_MOI' else 1
                    # Hỗ trợ cả ACCOUNT_CTS và MA_TB
                    account_val = row.get('ACCOUNT_CTS') or row.get('account_cts') or row.get('MA_TB') or row.get('ma_tb')
                    cursor.execute("""
                        INSERT INTO suy_hao_daily_changes (
                            ngay_bao_cao, account_cts, loai_bien_dong,
                            doi_one, nvkt_db, nvkt_db_normalized, sa, so_ngay_lien_tuc,
                            ten_tb_one, dt_onediachi_one, olt_cts, port_cts, thietbi, ketcuoi
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        report_date, account_val, loai,
                        row.get('DOI_ONE') or row.get('doi_one'),
                        row.get('NVKT_DB') or row.get('nvkt_db'),
                        row.get('NVKT_DB_NORMALIZED') or row.get('nvkt_db_normalized'),
                        row.get('SA') or row.get('sa'), so_ngay,
                        row.get('TEN_TB_ONE') or row.get('ten_tb_one'),
                        row.get('DT_ONEDIACHI_ONE') or row.get('dt_onediachi_one') or row.get('DT_ONE'),
                        row.get('OLT_CTS') or row.get('olt_cts'),
                        row.get('PORT_CTS') or row.get('port_cts'),
                        row.get('THIETBI') or row.get('thietbi'),
                        row.get('KETCUOI') or row.get('ketcuoi')
                    ))

            if len(df_tang_moi) > 0:
                save_changes(df_tang_moi, 'TANG_MOI')
            if len(df_giam_het) > 0:
                save_changes(df_giam_het, 'GIAM_HET')
            if len(df_van_con) > 0:
                save_changes(df_van_con, 'VAN_CON')

            # Tạo daily summary
            print(f"\n✓ Đang tạo daily summary...")
            cursor.execute("DELETE FROM suy_hao_daily_summary WHERE ngay_bao_cao = ?", (report_date,))

            cursor.execute(f"""
                INSERT INTO suy_hao_daily_summary (
                    ngay_bao_cao, doi_one, nvkt_db_normalized,
                    tong_so_hien_tai, so_tang_moi, so_giam_het, so_van_con
                )
                SELECT
                    '{report_date}', doi_one, nvkt_db_normalized,
                    SUM(CASE WHEN loai_bien_dong IN ('TANG_MOI', 'VAN_CON') THEN 1 ELSE 0 END),
                    SUM(CASE WHEN loai_bien_dong = 'TANG_MOI' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN loai_bien_dong = 'GIAM_HET' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN loai_bien_dong = 'VAN_CON' THEN 1 ELSE 0 END)
                FROM suy_hao_daily_changes
                WHERE ngay_bao_cao = '{report_date}'
                GROUP BY doi_one, nvkt_db_normalized
            """)

            # Cập nhật so_tb_quan_ly và ty_le_shc từ bảng thong_ke
            if len(df_thong_ke) > 0:
                print(f"✓ Đang cập nhật tỉ lệ SHC vào daily summary...")
                for _, row in df_thong_ke.iterrows():
                    doi_vt = row['DOI_VT']
                    nvkt = row['NVKT']
                    so_tb_ql = row['so_thue_bao_pon_qly']
                    
                    # Cập nhật so_tb_quan_ly
                    cursor.execute("""
                        UPDATE suy_hao_daily_summary 
                        SET so_tb_quan_ly = ?,
                            ty_le_shc = ROUND(CAST(tong_so_hien_tai AS REAL) / ? * 100, 2)
                        WHERE ngay_bao_cao = ? AND doi_one = ? AND nvkt_db_normalized = ?
                    """, (so_tb_ql, so_tb_ql if so_tb_ql > 0 else 1, report_date, doi_vt, nvkt))
                
                print(f"✅ Đã cập nhật tỉ lệ SHC cho {len(df_thong_ke)} NVKT")

            hist_conn.commit()

            df_bien_dong = pd.read_sql_query(f"""
                SELECT
                    doi_one as "Đơn vị",
                    nvkt_db_normalized as "NVKT_DB",
                    tong_so_hien_tai as "Tổng số hiện tại",
                    so_tang_moi as "Tăng mới",
                    so_giam_het as "Giảm/Hết",
                    so_van_con as "Vẫn còn",
                    so_tb_quan_ly as "Số TB quản lý",
                    ty_le_shc as "Tỉ lệ SHC (%)"
                FROM suy_hao_daily_summary
                WHERE ngay_bao_cao = '{report_date}'
                ORDER BY doi_one, nvkt_db_normalized
            """, hist_conn)

            hist_conn.close()
            print(f"  ✅ Đã lưu toàn bộ vào database lịch sử")

        # ==================================================================
        # TẠO CÁC SHEET THỐNG KÊ (LOGIC CŨ)
        # ==================================================================
        print("\n" + "="*80)
        print("TẠO CÁC SHEET THỐNG KÊ")
        print("="*80)

        # Sheet tổng hợp theo NVKT_DB và DOI_ONE
        print("\n✓ Đang đếm theo NVKT_DB và DOI_ONE...")
        if 'NVKT_DB_NORMALIZED' in df.columns and 'DOI_ONE' in df.columns:
            df_result = df.groupby(['NVKT_DB_NORMALIZED', 'DOI_ONE']).size().reset_index(name='Count')
            df_result = df_result[['DOI_ONE', 'NVKT_DB_NORMALIZED', 'Count']]
            df_result.columns = ['Đơn vị', 'NVKT_DB', f'Số TB Suy hao cao {k_suffix}']
            df_result = df_result.sort_values(by='Đơn vị').reset_index(drop=True)
            
            # Merge với thong_ke để lấy số thuê bao quản lý và tính tỉ lệ
            if len(df_thong_ke) > 0:
                df_result = df_result.merge(
                    df_thong_ke.rename(columns={'DOI_VT': 'Đơn vị', 'NVKT': 'NVKT_DB', 'so_thue_bao_pon_qly': 'Số TB quản lý'}),
                    on=['Đơn vị', 'NVKT_DB'],
                    how='left'
                )
                # Tính tỉ lệ suy hao cao (%)
                df_result['Tỉ lệ SHC (%)'] = (df_result[f'Số TB Suy hao cao {k_suffix}'] / df_result['Số TB quản lý'] * 100).round(2)
                df_result['Tỉ lệ SHC (%)'] = df_result['Tỉ lệ SHC (%)'].fillna(0)
                print(f"✅ Đã thêm cột Số TB quản lý và Tỉ lệ SHC (%)")
            
            print(f"✅ Đã đếm xong, tổng số nhóm: {len(df_result)}")
        else:
            print("❌ Không tìm thấy cột NVKT_DB_NORMALIZED hoặc DOI_ONE")
            return False

        # Tổng hợp theo tổ
        print("\n✓ Đang tạo tổng hợp theo tổ...")
        df_by_to = df_result.groupby('Đơn vị')[f'Số TB Suy hao cao {k_suffix}'].sum().reset_index()
        df_by_to = df_by_to.sort_values(by='Đơn vị').reset_index(drop=True)
        
        # Merge với thong_ke_theo_don_vi để lấy số thuê bao quản lý
        if len(df_thong_ke_dv) > 0:
            df_by_to = df_by_to.merge(
                df_thong_ke_dv.rename(columns={'don_vi': 'Đơn vị', 'so_thue_bao_pon_qly': 'Số TB quản lý'}),
                on='Đơn vị',
                how='left'
            )
            # Tính tỉ lệ suy hao cao (%)
            df_by_to['Tỉ lệ SHC (%)'] = (df_by_to[f'Số TB Suy hao cao {k_suffix}'] / df_by_to['Số TB quản lý'] * 100).round(2)
            df_by_to['Tỉ lệ SHC (%)'] = df_by_to['Tỉ lệ SHC (%)'].fillna(0)
            print(f"✅ Đã thêm cột Số TB quản lý và Tỉ lệ SHC (%) cho sheet theo tổ")
        
        # Tạo dòng tổng
        total_shc = df_by_to[f'Số TB Suy hao cao {k_suffix}'].sum()
        total_ql = df_by_to['Số TB quản lý'].sum() if 'Số TB quản lý' in df_by_to.columns else 0
        total_rate = round(total_shc / total_ql * 100, 2) if total_ql > 0 else 0
        
        if 'Số TB quản lý' in df_by_to.columns:
            total_row = pd.DataFrame({
                'Đơn vị': ['Tổng'],
                f'Số TB Suy hao cao {k_suffix}': [total_shc],
                'Số TB quản lý': [total_ql],
                'Tỉ lệ SHC (%)': [total_rate]
            })
        else:
            total_row = pd.DataFrame({'Đơn vị': ['Tổng'], f'Số TB Suy hao cao {k_suffix}': [total_shc]})
        df_by_to = pd.concat([df_by_to, total_row], ignore_index=True)

        # Thống kê theo SA
        print("\n✓ Đang tạo thống kê theo SA...")
        if 'SA' in df.columns:
            df_by_sa = df.groupby('SA').size().reset_index(name='Số lượng')
            df_by_sa = df_by_sa.sort_values(by='Số lượng', ascending=False).reset_index(drop=True)
            total_sa_row = pd.DataFrame({'SA': ['Tổng'], 'Số lượng': [df_by_sa['Số lượng'].sum()]})
            df_by_sa = pd.concat([df_by_sa, total_sa_row], ignore_index=True)
        else:
            df_by_sa = None

        # Danh sách chi tiết cho từng NVKT_DB
        print("\n✓ Đang tạo danh sách chi tiết cho từng NVKT_DB...")
        columns_to_keep = ['MA_TB', 'ACCOUNT_CTS', 'TEN_TB_ONE', 'DIACHI_ONE', 'DT_ONE', 'DT_ONEDIACHI_ONE', 'NGAY_SUYHAO',
                          'THIETBI', 'SA', 'KETCUOI', 'NVKT_DB_NORMALIZED']
        missing_cols = [col for col in columns_to_keep if col not in df.columns]
        if missing_cols:
            print(f"⚠️ Không tìm thấy các cột: {', '.join(missing_cols)}")
            columns_to_keep = [col for col in columns_to_keep if col in df.columns]

        df_detail = df[columns_to_keep].copy()
        nvkt_list = df_detail['NVKT_DB_NORMALIZED'].unique()
        print(f"✅ Tìm thấy {len(nvkt_list)} NVKT_DB cần tạo sheet chi tiết")

        # ==================================================================
        # GHI VÀO FILE EXCEL
        # ==================================================================
        print("\n✓ Đang ghi vào các sheet...")

        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            add_tt_column(df).to_excel(writer, sheet_name='Sheet1', index=False)
            print(f"  ✅ Sheet1 (dữ liệu gốc): {len(df)} dòng")

            add_tt_column(df_result).to_excel(writer, sheet_name='TH_SHC_I15', index=False)
            add_tt_column(df_by_to).to_excel(writer, sheet_name='TH_SHC_theo_to', index=False)
            if df_by_sa is not None:
                add_tt_column(df_by_sa).to_excel(writer, sheet_name='shc_theo_SA', index=False)

            if len(df_bien_dong) > 0:
                add_tt_column(df_bien_dong).to_excel(writer, sheet_name='Bien_dong_tong_hop', index=False)
                print(f"  ✅ Bien_dong_tong_hop: {len(df_bien_dong)} dòng")

            if len(df_tang_moi) > 0:
                cols_tang = ['ACCOUNT_CTS', 'TEN_TB_ONE', 'DT_ONEDIACHI_ONE', 'DOI_ONE',
                            'NVKT_DB_NORMALIZED', 'SA', 'OLT_CTS', 'PORT_CTS', 'THIETBI', 'KETCUOI']
                cols_tang = [c for c in cols_tang if c in df_tang_moi.columns]
                add_tt_column(df_tang_moi[cols_tang]).to_excel(writer, sheet_name='Tang_moi', index=False)
                print(f"  ✅ Tang_moi: {len(df_tang_moi)} dòng")

            if len(df_giam_het) > 0:
                cols_giam = ['account_cts', 'ten_tb_one', 'dt_onediachi_one', 'doi_one',
                            'nvkt_db_normalized', 'sa', 'so_ngay_lien_tuc', 'olt_cts', 'port_cts', 'thietbi', 'ketcuoi']
                cols_giam = [c for c in cols_giam if c in df_giam_het.columns]
                df_giam_out = df_giam_het[cols_giam].copy()
                df_giam_out.columns = [c.upper() if c != 'so_ngay_lien_tuc' else 'Số ngày suy hao' for c in df_giam_out.columns]
                add_tt_column(df_giam_out).to_excel(writer, sheet_name='Giam_het', index=False)
                print(f"  ✅ Giam_het: {len(df_giam_het)} dòng")

            if len(df_van_con) > 0:
                cols_van = ['ACCOUNT_CTS', 'TEN_TB_ONE', 'DT_ONEDIACHI_ONE', 'DOI_ONE',
                           'NVKT_DB_NORMALIZED', 'SA', 'so_ngay_lien_tuc', 'OLT_CTS', 'PORT_CTS', 'THIETBI', 'KETCUOI']
                cols_van = [c for c in cols_van if c in df_van_con.columns]
                df_van_out = df_van_con[cols_van].copy()
                if 'so_ngay_lien_tuc' in df_van_out.columns:
                    df_van_out = df_van_out.rename(columns={'so_ngay_lien_tuc': 'Số ngày liên tục'})
                add_tt_column(df_van_out).to_excel(writer, sheet_name='Van_con', index=False)
                print(f"  ✅ Van_con: {len(df_van_con)} dòng")

            for nvkt in nvkt_list:
                df_nvkt = df_detail[df_detail['NVKT_DB_NORMALIZED'] == nvkt].copy()
                if 'SA' in df_nvkt.columns:
                    df_nvkt = df_nvkt.sort_values(by='SA').reset_index(drop=True)
                df_nvkt = df_nvkt.drop(columns=['NVKT_DB_NORMALIZED'])
                sheet_name = str(nvkt)[:31]
                add_tt_column(df_nvkt).to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"  ✅ Đã tạo {len(nvkt_list)} sheet chi tiết NVKT_DB")

        # ==================================================================
        # TẠO FILE EXCEL CHI TIẾT CHO TỪNG NVKT THEO THƯ MỤC TỔ
        # ==================================================================
        print("\n" + "="*80)
        print("TẠO FILE EXCEL CHI TIẾT CHO TỪNG NVKT")
        print("="*80)

        # Tạo thư mục gốc (tách riêng cho K1 và K2)
        base_dir = os.path.dirname(input_file)
        detail_dir = os.path.join(base_dir, f"shc_NVKT_danh_sach_chi_tiet_{k_suffix}")
        if not os.path.exists(detail_dir):
            os.makedirs(detail_dir)
            print(f"✓ Đã tạo thư mục: {detail_dir}")
        
        detail_dir_26_27 = None
        if k_suffix == "K2":
            detail_dir_26_27 = os.path.join(base_dir, "shc_NVKT_danh_sach_chi_tiet_K2_26_27")
            if not os.path.exists(detail_dir_26_27):
                os.makedirs(detail_dir_26_27)
                print(f"✓ Đã tạo thư mục: {detail_dir_26_27}")

        # Lấy danh sách các cột cần thiết cho file chi tiết
        detail_columns = ['MA_TB', 'ACCOUNT_CTS', 'TEN_TB_ONE', 'DIACHI_ONE', 'DT_ONE', 'DT_ONEDIACHI_ONE', 'NGAY_SUYHAO',
                         'THIETBI', 'SA', 'KETCUOI', 'Chỉ số OLT RX', 'Chỉ số ONU RX']
        detail_columns = [c for c in detail_columns if c in df.columns]

        # Đếm số file đã tạo
        file_count = 0
        file_count_26_27 = 0

        # Lấy danh sách các tổ có trong dữ liệu
        if 'DOI_ONE' in df.columns:
            doi_list = df['DOI_ONE'].dropna().unique()
            print(f"✓ Tìm thấy {len(doi_list)} tổ")

            for doi in doi_list:
                # Tạo thư mục cho tổ (thay thế ký tự không hợp lệ)
                doi_safe = str(doi).replace('/', '_').replace('\\', '_').replace(':', '_').strip()
                doi_dir = os.path.join(detail_dir, doi_safe)
                if not os.path.exists(doi_dir):
                    os.makedirs(doi_dir)
                
                doi_dir_26_27 = None
                if detail_dir_26_27:
                    doi_dir_26_27 = os.path.join(detail_dir_26_27, doi_safe)
                    if not os.path.exists(doi_dir_26_27):
                        os.makedirs(doi_dir_26_27)

                # Lấy danh sách NVKT trong tổ
                df_doi = df[df['DOI_ONE'] == doi]
                nvkt_in_doi = df_doi['NVKT_DB_NORMALIZED'].dropna().unique()

                for nvkt in nvkt_in_doi:
                    # Lấy dữ liệu của NVKT
                    df_nvkt_detail = df_doi[df_doi['NVKT_DB_NORMALIZED'] == nvkt][detail_columns].copy()
                    
                    if len(df_nvkt_detail) == 0:
                        continue

                    # Sắp xếp theo SA
                    if 'SA' in df_nvkt_detail.columns:
                        df_nvkt_detail = df_nvkt_detail.sort_values(by='SA').reset_index(drop=True)

                    # Tạo tên file (thay thế ký tự không hợp lệ)
                    nvkt_safe = str(nvkt).replace('/', '_').replace('\\', '_').replace(':', '_').strip()
                    file_name = f"{nvkt_safe}.xlsx"
                    file_path = os.path.join(doi_dir, file_name)

                    # Ghi file Excel với cột TT
                    df_formatted = add_tt_column(df_nvkt_detail)
                    df_formatted.to_excel(file_path, index=False, sheet_name='Chi tiết SHC')
                    # Định dạng file Excel
                    format_excel_detail(file_path, df_formatted)

                    # Thêm các sheet theo ngưỡng cho K2
                    if k_suffix == "K2":
                        olt_col = 'Chỉ số OLT RX'
                        onu_col = 'Chỉ số ONU RX'
                        # Lấy dữ liệu NVKT với đầy đủ cột (bao gồm OLT/ONU RX)
                        df_nvkt_full = df_doi[df_doi['NVKT_DB_NORMALIZED'] == nvkt].copy()
                        if olt_col in df_nvkt_full.columns and onu_col in df_nvkt_full.columns:
                            df_nvkt_full[olt_col] = pd.to_numeric(df_nvkt_full[olt_col], errors='coerce')
                            df_nvkt_full[onu_col] = pd.to_numeric(df_nvkt_full[onu_col], errors='coerce')
                            
                            nguong_cols = list(dict.fromkeys([c for c in detail_columns if c in df_nvkt_full.columns] + [olt_col, onu_col]))
                            nguong_thresholds = [
                                ('k2_25_26', -26, -25, True, True),
                                ('k2_26_26.5', -26.5, -26, True, False),
                                ('k2_26.5_27', -27, -26.5, True, False),
                            ]
                            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                for sheet_name, low, high, low_inc, high_inc in nguong_thresholds:
                                    if low_inc and high_inc:
                                        olt_m = (df_nvkt_full[olt_col] >= low) & (df_nvkt_full[olt_col] <= high)
                                        onu_m = (df_nvkt_full[onu_col] >= low) & (df_nvkt_full[onu_col] <= high)
                                    else:
                                        olt_m = (df_nvkt_full[olt_col] >= low) & (df_nvkt_full[olt_col] < high)
                                        onu_m = (df_nvkt_full[onu_col] >= low) & (df_nvkt_full[onu_col] < high)
                                    df_ng = df_nvkt_full[olt_m | onu_m][nguong_cols].copy()
                                    if 'SA' in df_ng.columns:
                                        df_ng = df_ng.sort_values(by='SA').reset_index(drop=True)
                                    df_ng.insert(0, 'STT', range(1, len(df_ng) + 1))
                                    df_ng.to_excel(writer, sheet_name=sheet_name, index=False)
                            # Format các sheet ngưỡng (tô màu SA trùng, kẻ bảng)
                            for sheet_name, _, _, _, _ in nguong_thresholds:
                                format_excel_detail(file_path, df_nvkt_full, sheet_name=sheet_name)
                            
                            # Tạo thêm file chỉ chứa SHC có OLT RX hoặc ONU RX <= -26 cho từng NVKT
                            if doi_dir_26_27:
                                range_cols = [c for c in detail_columns if c in df_nvkt_full.columns]
                                olt_m_26_27 = df_nvkt_full[olt_col] <= -26
                                onu_m_26_27 = df_nvkt_full[onu_col] <= -26
                                df_26_27 = df_nvkt_full[olt_m_26_27 | onu_m_26_27][range_cols].copy()

                                if len(df_26_27) > 0:
                                    if 'SA' in df_26_27.columns:
                                        df_26_27 = df_26_27.sort_values(by='SA').reset_index(drop=True)
                                    file_name_26_27 = f"{nvkt_safe}_K2_26-27.xlsx"
                                    file_path_26_27 = os.path.join(doi_dir_26_27, file_name_26_27)
                                    df_26_27_formatted = add_tt_column(df_26_27)
                                    df_26_27_formatted.to_excel(file_path_26_27, index=False, sheet_name='Chi tiết SHC')
                                    format_excel_detail(file_path_26_27, df_26_27_formatted)
                                    file_count_26_27 += 1

                    file_count += 1

            print(f"✅ Đã tạo {file_count} file Excel chi tiết trong thư mục {detail_dir}")
            if detail_dir_26_27:
                print(f"✅ Đã tạo {file_count_26_27} file Excel chi tiết ngưỡng <= -26 trong thư mục {detail_dir_26_27}")
        else:
            print("⚠️ Không tìm thấy cột DOI_ONE, bỏ qua tạo file chi tiết theo tổ")

        # Tạo báo cáo K2 theo ngưỡng (chỉ khi xử lý K2)
        if k_suffix == "K2":
            create_k2_threshold_report(df, os.path.dirname(input_file))

        print("\n" + "="*80)
        print("✅ HOÀN THÀNH XỬ LÝ BÁO CÁO I1.5")
        print("="*80)

        return True

    except Exception as e:
        print(f"\n❌ Lỗi khi xử lý báo cáo I1.5: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import argparse
    from suy_hao_reports import generate_daily_comparison_report, generate_daily_comparison_report_k2
    
    parser = argparse.ArgumentParser(description="Xử lý báo cáo I1.5 Suy hao cao")
    parser.add_argument("--k1", action="store_true", help="Chỉ xử lý báo cáo K1")
    parser.add_argument("--k2", action="store_true", help="Chỉ xử lý báo cáo K2")
    parser.add_argument("--force", action="store_true", help="Ghi đè dữ liệu đã tồn tại")
    parser.add_argument("--date", type=str, help="Chỉ định ngày báo cáo thủ công (định dạng YYYY-MM-DD)")
    
    args = parser.parse_args()
    
    # Nếu không chọn gì thì chạy cả hai
    run_k1 = args.k1 or (not args.k1 and not args.k2)
    run_k2 = args.k2 or (not args.k1 and not args.k2)
    
    if run_k1:
        # print("\n" + "="*80)
        # print("XỬ LÝ BÁO CÁO K1")
        # print("="*80)
        # process_I15_report_with_tracking(force_update=args.force, report_date=args.date)
        
        # # Tạo báo cáo so sánh SHC ngày (T so với T-1)
        # print("\n" + "="*80)
        # print("TẠO BÁO CÁO SO SÁNH SHC K1 (T so với T-1)")
        # print("="*80)
        # generate_daily_comparison_report()
        pass    
    
    if run_k2:
        print("\n" + "="*80)
        print("XỬ LÝ BÁO CÁO K2")
        print("="*80)
        process_I15_k2_report_with_tracking(force_update=args.force, report_date=args.date)
        
        # Tạo báo cáo so sánh SHC K2 ngày (T so với T-1)
        print("\n" + "="*80)
        print("TẠO BÁO CÁO SO SÁNH SHC K2 (T so với T-1)")
        print("="*80)
        generate_daily_comparison_report_k2()

