# -*- coding: utf-8 -*-
"""
Module tạo báo cáo xu hướng suy hao theo tuần/tháng/ngày
"""

import pandas as pd
import sqlite3
import os
from datetime import datetime, timedelta


def _build_daily_comparison(conn, today_date, yesterday_date):
    """
    Tính toán so sánh biến động SHC giữa 2 ngày từ snapshots (tính động).
    Trả về: df_compare, df_by_unit, total_row, yesterday_str, today_str,
            col_sl_qua, col_sl_nay, col_tl_qua, col_tl_nay
    Hoặc None nếu không có dữ liệu ngày T.
    """
    # Lấy dữ liệu summary ngày T (số lượng, TB quản lý, tỉ lệ)
    df_today = pd.read_sql_query(f"""
        SELECT
            doi_one,
            nvkt_db_normalized,
            tong_so_hien_tai as so_luong_hom_nay,
            so_tb_quan_ly,
            ty_le_shc as ty_le_hom_nay
        FROM suy_hao_daily_summary
        WHERE ngay_bao_cao = '{today_date.strftime('%Y-%m-%d')}'
    """, conn)

    # Lấy dữ liệu summary ngày T-1
    df_yesterday = pd.read_sql_query(f"""
        SELECT
            doi_one,
            nvkt_db_normalized,
            tong_so_hien_tai as so_luong_hom_qua,
            ty_le_shc as ty_le_hom_qua
        FROM suy_hao_daily_summary
        WHERE ngay_bao_cao = '{yesterday_date.strftime('%Y-%m-%d')}'
    """, conn)

    print(f"\n✓ Dữ liệu T ({today_date.strftime('%d/%m')}): {len(df_today)} NVKT")
    print(f"✓ Dữ liệu T-1 ({yesterday_date.strftime('%d/%m')}): {len(df_yesterday)} NVKT")

    if len(df_today) == 0:
        print(f"❌ Không có dữ liệu ngày {today_date.strftime('%d/%m/%Y')}")
        return None

    # Tính Phát sinh / Đã giảm / Vẫn còn ĐỘNG từ snapshots
    df_snap_today = pd.read_sql_query(f"""
        SELECT account_cts, doi_one, nvkt_db_normalized
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao = '{today_date.strftime('%Y-%m-%d')}'
    """, conn)
    df_snap_yesterday = pd.read_sql_query(f"""
        SELECT account_cts, doi_one, nvkt_db_normalized
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao = '{yesterday_date.strftime('%Y-%m-%d')}'
    """, conn)

    accounts_today = set(df_snap_today['account_cts'])
    accounts_yesterday = set(df_snap_yesterday['account_cts'])
    tang_moi_set = accounts_today - accounts_yesterday
    giam_het_set = accounts_yesterday - accounts_today
    van_con_set = accounts_today & accounts_yesterday

    print(f"✓ So sánh snapshots: +{len(tang_moi_set)} tăng mới, -{len(giam_het_set)} giảm/hết, {len(van_con_set)} vẫn còn")

    # Đếm theo NVKT cho từng loại biến động
    df_tang_count = df_snap_today[df_snap_today['account_cts'].isin(tang_moi_set)] \
        .groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_phat_sinh')
    df_giam_count = df_snap_yesterday[df_snap_yesterday['account_cts'].isin(giam_het_set)] \
        .groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_da_giam')
    df_van_count = df_snap_today[df_snap_today['account_cts'].isin(van_con_set)] \
        .groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_van_con')

    # Merge dữ liệu
    df_compare = df_today.merge(
        df_yesterday[['doi_one', 'nvkt_db_normalized', 'so_luong_hom_qua', 'ty_le_hom_qua']],
        on=['doi_one', 'nvkt_db_normalized'],
        how='outer'
    ).fillna(0)

    # Merge biến động động
    for df_bd in [df_tang_count, df_giam_count, df_van_count]:
        df_compare = df_compare.merge(df_bd, on=['doi_one', 'nvkt_db_normalized'], how='left')
    df_compare[['so_phat_sinh', 'so_da_giam', 'so_van_con']] = \
        df_compare[['so_phat_sinh', 'so_da_giam', 'so_van_con']].fillna(0).astype(int)

    # Tính toán chênh lệch
    df_compare['chenh_lech'] = df_compare['so_luong_hom_nay'] - df_compare['so_luong_hom_qua']
    df_compare['chenh_lech_ty_le'] = (df_compare['ty_le_hom_nay'] - df_compare['ty_le_hom_qua']).round(2)

    # Tạo tên cột với ngày cụ thể
    yesterday_str = yesterday_date.strftime('%d/%m')
    today_str = today_date.strftime('%d/%m')

    col_sl_qua = f'SL {yesterday_str}'
    col_sl_nay = f'SL {today_str}'
    col_tl_qua = f'TL% {yesterday_str}'
    col_tl_nay = f'TL% {today_str}'

    # Đổi tên cột
    df_compare = df_compare.rename(columns={
        'doi_one': 'Đơn vị',
        'nvkt_db_normalized': 'NVKT',
        'so_luong_hom_qua': col_sl_qua,
        'so_luong_hom_nay': col_sl_nay,
        'so_phat_sinh': 'Phát sinh',
        'so_da_giam': 'Đã giảm',
        'so_van_con': 'Vẫn còn',
        'chenh_lech': '+/- SL',
        'so_tb_quan_ly': 'TB Quản lý',
        'ty_le_hom_qua': col_tl_qua,
        'ty_le_hom_nay': col_tl_nay,
        'chenh_lech_ty_le': '+/- TL%'
    })

    # Sắp xếp cột
    columns_order = ['Đơn vị', 'NVKT', col_sl_qua, col_sl_nay, '+/- SL',
                     'Phát sinh', 'Đã giảm', 'Vẫn còn', 'TB Quản lý',
                     col_tl_qua, col_tl_nay, '+/- TL%']
    df_compare = df_compare[[c for c in columns_order if c in df_compare.columns]]
    df_compare = df_compare.sort_values(by=['Đơn vị', 'NVKT'])

    # Tổng hợp theo đơn vị
    df_by_unit = df_compare.groupby('Đơn vị').agg({
        col_sl_qua: 'sum',
        col_sl_nay: 'sum',
        'Phát sinh': 'sum',
        'Đã giảm': 'sum',
        'Vẫn còn': 'sum',
        'TB Quản lý': 'sum'
    }).reset_index()

    df_by_unit['+/- SL'] = df_by_unit[col_sl_nay] - df_by_unit[col_sl_qua]
    df_by_unit[col_tl_qua] = (df_by_unit[col_sl_qua] / df_by_unit['TB Quản lý'] * 100).round(2)
    df_by_unit[col_tl_nay] = (df_by_unit[col_sl_nay] / df_by_unit['TB Quản lý'] * 100).round(2)
    df_by_unit['+/- TL%'] = (df_by_unit[col_tl_nay] - df_by_unit[col_tl_qua]).round(2)

    # Xử lý inf/nan
    df_by_unit = df_by_unit.replace([float('inf'), -float('inf')], 0).fillna(0)

    # Thêm dòng tổng
    total_row = pd.DataFrame({
        'Đơn vị': ['TỔNG CỘNG'],
        col_sl_qua: [df_by_unit[col_sl_qua].sum()],
        col_sl_nay: [df_by_unit[col_sl_nay].sum()],
        '+/- SL': [df_by_unit['+/- SL'].sum()],
        'Phát sinh': [df_by_unit['Phát sinh'].sum()],
        'Đã giảm': [df_by_unit['Đã giảm'].sum()],
        'Vẫn còn': [df_by_unit['Vẫn còn'].sum()],
        'TB Quản lý': [df_by_unit['TB Quản lý'].sum()],
        col_tl_qua: [round(df_by_unit[col_sl_qua].sum() / df_by_unit['TB Quản lý'].sum() * 100, 2) if df_by_unit['TB Quản lý'].sum() > 0 else 0],
        col_tl_nay: [round(df_by_unit[col_sl_nay].sum() / df_by_unit['TB Quản lý'].sum() * 100, 2) if df_by_unit['TB Quản lý'].sum() > 0 else 0],
        '+/- TL%': [0]
    })
    total_row['+/- TL%'] = total_row[col_tl_nay] - total_row[col_tl_qua]
    df_by_unit = pd.concat([df_by_unit, total_row], ignore_index=True)

    # Sắp xếp cột cho df_by_unit
    unit_columns = ['Đơn vị', col_sl_qua, col_sl_nay, '+/- SL',
                    'Phát sinh', 'Đã giảm', 'Vẫn còn', 'TB Quản lý',
                    col_tl_qua, col_tl_nay, '+/- TL%']
    df_by_unit = df_by_unit[[c for c in unit_columns if c in df_by_unit.columns]]

    return {
        'df_compare': df_compare,
        'df_by_unit': df_by_unit,
        'total_row': total_row,
        'yesterday_str': yesterday_str,
        'today_str': today_str,
        'col_sl_qua': col_sl_qua,
        'col_sl_nay': col_sl_nay,
        'col_tl_qua': col_tl_qua,
        'col_tl_nay': col_tl_nay,
    }


def _find_previous_date(conn, today_date):
    """Tìm ngày gần nhất trước today_date có dữ liệu trong DB."""
    df_prev = pd.read_sql_query(f"""
        SELECT MAX(ngay_bao_cao) as prev_date FROM suy_hao_snapshots
        WHERE ngay_bao_cao < '{today_date.strftime('%Y-%m-%d')}'
    """, conn)

    if df_prev.empty or df_prev['prev_date'][0] is None:
        yesterday_date = today_date - timedelta(days=1)
        print(f"⚠️ Không tìm thấy dữ liệu ngày trước T, dùng T-1: {yesterday_date.strftime('%d/%m/%Y')}")
    else:
        yesterday_date = datetime.strptime(df_prev['prev_date'][0], '%Y-%m-%d')
    return yesterday_date


def generate_daily_comparison_report(today_date=None, output_file=None):
    """
    Tạo báo cáo so sánh suy hao cao ngày hôm nay với ngày hôm qua
    
    Args:
        today_date: Ngày cần so sánh (format: 'YYYY-MM-DD' hoặc datetime). 
                   Mặc định lấy ngày mới nhất trong database.
        output_file: Đường dẫn file Excel output (tùy chọn)
    
    Returns:
        Đường dẫn file Excel đã tạo
    """
    print(f"\n{'='*80}")
    print(f"TẠO BÁO CÁO SO SÁNH SHC NGÀY")
    print(f"{'='*80}\n")
    
    db_path = "suy_hao_history.db"
    if not os.path.exists(db_path):
        print(f"❌ Không tìm thấy database: {db_path}")
        return None

    conn = sqlite3.connect(db_path)

    # Xác định ngày T
    if today_date is None:
        df_latest = pd.read_sql_query("SELECT MAX(ngay_bao_cao) as latest FROM suy_hao_snapshots", conn)
        if df_latest.empty or df_latest['latest'][0] is None:
            print("❌ Không có dữ liệu trong database")
            conn.close()
            return None
        today_date = datetime.strptime(df_latest['latest'][0], '%Y-%m-%d')
    elif isinstance(today_date, str):
        today_date = datetime.strptime(today_date, '%Y-%m-%d')

    yesterday_date = _find_previous_date(conn, today_date)
    print(f"✓ Ngày T: {today_date.strftime('%d/%m/%Y')}")
    print(f"✓ Ngày T-1: {yesterday_date.strftime('%d/%m/%Y')}")

    result = _build_daily_comparison(conn, today_date, yesterday_date)
    conn.close()

    if result is None:
        return None

    df_compare = result['df_compare']
    df_by_unit = result['df_by_unit']
    total_row = result['total_row']
    yesterday_str = result['yesterday_str']
    today_str = result['today_str']
    col_sl_qua = result['col_sl_qua']
    col_sl_nay = result['col_sl_nay']

    # In thống kê
    print(f"\n{'─'*60}")
    print(f"📊 THỐNG KÊ TỔNG HỢP:")
    print(f"   SL {yesterday_str}: {int(total_row[col_sl_qua].iloc[0])}")
    print(f"   SL {today_str}: {int(total_row[col_sl_nay].iloc[0])}")
    print(f"   Phát sinh:  +{int(total_row['Phát sinh'].iloc[0])}")
    print(f"   Đã giảm:    -{int(total_row['Đã giảm'].iloc[0])}")
    print(f"   Chênh lệch: {int(total_row['+/- SL'].iloc[0]):+d}")
    print(f"{'─'*60}")

    # Tạo file Excel
    if output_file is None:
        output_file = "downloads/baocao_hanoi/So_sanh_SHC_theo_ngay_T-1.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"\n✓ Đang ghi file Excel: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_by_unit.to_excel(writer, sheet_name='Theo_don_vi', index=False)
        df_compare.to_excel(writer, sheet_name='Chi_tiet_NVKT', index=False)

    print(f"✅ Đã tạo báo cáo so sánh: {output_file}")
    print(f"\n{'='*80}")
    print(f"✅ HOÀN THÀNH BÁO CÁO SO SÁNH SHC NGÀY")
    print(f"{'='*80}\n")

    return output_file


def generate_daily_comparison_report_k2(today_date=None, output_file=None):
    """
    Tạo báo cáo so sánh suy hao cao K2 ngày hôm nay với ngày hôm qua
    Sử dụng database suy_hao_history_k2.db
    
    Args:
        today_date: Ngày cần so sánh (format: 'YYYY-MM-DD' hoặc datetime). 
                   Mặc định lấy ngày mới nhất trong database.
        output_file: Đường dẫn file Excel output (tùy chọn)
    
    Returns:
        Đường dẫn file Excel đã tạo
    """
    print(f"\n{'='*80}")
    print(f"TẠO BÁO CÁO SO SÁNH SHC K2 NGÀY")
    print(f"{'='*80}\n")
    
    db_path = "suy_hao_history_k2.db"
    if not os.path.exists(db_path):
        print(f"❌ Không tìm thấy database: {db_path}")
        return None

    conn = sqlite3.connect(db_path)

    # Xác định ngày T
    if today_date is None:
        df_latest = pd.read_sql_query("SELECT MAX(ngay_bao_cao) as latest FROM suy_hao_snapshots", conn)
        if df_latest.empty or df_latest['latest'][0] is None:
            print("❌ Không có dữ liệu trong database K2")
            conn.close()
            return None
        today_date = datetime.strptime(df_latest['latest'][0], '%Y-%m-%d')
    elif isinstance(today_date, str):
        today_date = datetime.strptime(today_date, '%Y-%m-%d')

    yesterday_date = _find_previous_date(conn, today_date)
    print(f"✓ Ngày T: {today_date.strftime('%d/%m/%Y')}")
    print(f"✓ Ngày T-1: {yesterday_date.strftime('%d/%m/%Y')}")

    result = _build_daily_comparison(conn, today_date, yesterday_date)
    conn.close()

    if result is None:
        return None

    df_compare = result['df_compare']
    df_by_unit = result['df_by_unit']
    total_row = result['total_row']
    yesterday_str = result['yesterday_str']
    today_str = result['today_str']
    col_sl_qua = result['col_sl_qua']
    col_sl_nay = result['col_sl_nay']

    # In thống kê
    print(f"\n{'─'*60}")
    print(f"📊 THỐNG KÊ TỔNG HỢP K2:")
    print(f"   SL {yesterday_str}: {int(total_row[col_sl_qua].iloc[0])}")
    print(f"   SL {today_str}: {int(total_row[col_sl_nay].iloc[0])}")
    print(f"   Phát sinh:  +{int(total_row['Phát sinh'].iloc[0])}")
    print(f"   Đã giảm:    -{int(total_row['Đã giảm'].iloc[0])}")
    print(f"   Chênh lệch: {int(total_row['+/- SL'].iloc[0]):+d}")
    print(f"{'─'*60}")

    # Tạo file Excel
    if output_file is None:
        output_file = "downloads/baocao_hanoi/So_sanh_SHC_k2_theo_ngay_T-1.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"\n✓ Đang ghi file Excel: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_by_unit.to_excel(writer, sheet_name='Theo_don_vi', index=False)
        df_compare.to_excel(writer, sheet_name='Chi_tiet_NVKT', index=False)

    print(f"✅ Đã tạo báo cáo so sánh K2: {output_file}")
    print(f"\n{'='*80}")
    print(f"✅ HOÀN THÀNH BÁO CÁO SO SÁNH SHC K2 NGÀY")
    print(f"{'='*80}\n")
    
    return output_file

def generate_weekly_report(year, week_number, output_file=None):
    """
    Tạo báo cáo so sánh theo tuần

    Args:
        year: Năm (VD: 2025)
        week_number: Số tuần trong năm (1-52)
        output_file: Đường dẫn file Excel output (tùy chọn)

    Returns:
        Đường dẫn file Excel đã tạo
    """
    print(f"\n{'='*80}")
    print(f"TẠO BÁO CÁO TUẦN {week_number}/{year}")
    print(f"{'='*80}\n")

    db_path = "suy_hao_history.db"
    if not os.path.exists(db_path):
        print(f"❌ Không tìm thấy database: {db_path}")
        return None

    conn = sqlite3.connect(db_path)

    # Tính ngày bắt đầu và kết thúc tuần
    week_start = datetime.strptime(f'{year}-W{week_number:02d}-1', '%Y-W%W-%w')
    week_end = week_start + timedelta(days=6)

    # Tuần trước
    prev_week_start = week_start - timedelta(days=7)
    prev_week_end = week_end - timedelta(days=7)

    print(f"✓ Tuần hiện tại: {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}")
    print(f"✓ Tuần trước: {prev_week_start.strftime('%d/%m/%Y')} - {prev_week_end.strftime('%d/%m/%Y')}")

    # Lấy danh sách thuê bao tuần hiện tại
    df_current = pd.read_sql_query(f"""
        SELECT DISTINCT account_cts, doi_one, nvkt_db_normalized, sa, ten_tb_one
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{week_start.strftime('%Y-%m-%d')}'
          AND '{week_end.strftime('%Y-%m-%d')}'
    """, conn)

    # Lấy danh sách thuê bao tuần trước
    df_previous = pd.read_sql_query(f"""
        SELECT DISTINCT account_cts, doi_one, nvkt_db_normalized, sa, ten_tb_one
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{prev_week_start.strftime('%Y-%m-%d')}'
          AND '{prev_week_end.strftime('%Y-%m-%d')}'
    """, conn)

    print(f"\n✓ Tuần hiện tại: {len(df_current)} thuê bao")
    print(f"✓ Tuần trước: {len(df_previous)} thuê bao")

    # Phân loại biến động
    accounts_current = set(df_current['account_cts'].tolist())
    accounts_previous = set(df_previous['account_cts'].tolist())

    tang_moi = accounts_current - accounts_previous
    giam_het = accounts_previous - accounts_current
    van_con = accounts_current & accounts_previous

    print(f"\n✓ Phân tích biến động:")
    print(f"  🆕 TĂNG MỚI: {len(tang_moi)} thuê bao")
    print(f"  ⬇️  GIẢM/HẾT: {len(giam_het)} thuê bao")
    print(f"  ↔️  VẪN CÒN: {len(van_con)} thuê bao")

    # Tạo DataFrames
    df_tang = df_current[df_current['account_cts'].isin(tang_moi)].copy()
    df_giam = df_previous[df_previous['account_cts'].isin(giam_het)].copy()
    df_van = df_current[df_current['account_cts'].isin(van_con)].copy()

    # Tổng hợp theo NVKT_DB
    summary_current = df_current.groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_luong_tuan_nay')
    summary_previous = df_previous.groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_luong_tuan_truoc')

    df_summary = summary_current.merge(
        summary_previous,
        on=['doi_one', 'nvkt_db_normalized'],
        how='outer'
    ).fillna(0)

    df_summary['tang_giam'] = df_summary['so_luong_tuan_nay'] - df_summary['so_luong_tuan_truoc']
    df_summary['ty_le_thay_doi'] = (df_summary['tang_giam'] / df_summary['so_luong_tuan_truoc'] * 100).round(1)
    df_summary['ty_le_thay_doi'] = df_summary['ty_le_thay_doi'].replace([float('inf'), -float('inf')], 0)

    # Lấy thông tin tỉ lệ SHC từ suy_hao_daily_summary (ngày cuối tuần)
    df_ratio = pd.read_sql_query(f"""
        SELECT doi_one, nvkt_db_normalized, so_tb_quan_ly, ty_le_shc
        FROM suy_hao_daily_summary
        WHERE ngay_bao_cao = (SELECT MAX(ngay_bao_cao) FROM suy_hao_daily_summary 
                              WHERE ngay_bao_cao <= '{week_end.strftime('%Y-%m-%d')}')
    """, conn)
    
    if len(df_ratio) > 0:
        df_summary = df_summary.merge(
            df_ratio,
            on=['doi_one', 'nvkt_db_normalized'],
            how='left'
        )
        df_summary['so_tb_quan_ly'] = df_summary['so_tb_quan_ly'].fillna(0).astype(int)
        df_summary['ty_le_shc'] = df_summary['ty_le_shc'].fillna(0)
        df_summary.columns = ['Đơn vị', 'NVKT_DB', 'Tuần này', 'Tuần trước', 'Tăng/Giảm', '% Thay đổi', 'Số TB quản lý', 'Tỉ lệ SHC (%)']
    else:
        df_summary.columns = ['Đơn vị', 'NVKT_DB', 'Tuần này', 'Tuần trước', 'Tăng/Giảm', '% Thay đổi']
    
    df_summary = df_summary.sort_values(by=['Đơn vị', 'NVKT_DB'])

    conn.close()

    # Tạo file Excel
    if output_file is None:
        output_file = f"downloads/baocao_hanoi/Bao_cao_tuan_{week_number}_{year}.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"\n✓ Đang ghi file Excel: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet tổng hợp
        df_summary.to_excel(writer, sheet_name='Tong_hop', index=False)

        # Sheet chi tiết TĂNG MỚI
        if len(df_tang) > 0:
            df_tang[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Tang_moi', index=False
            )

        # Sheet chi tiết GIẢM/HẾT
        if len(df_giam) > 0:
            df_giam[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Giam_het', index=False
            )

        # Sheet VẪN CÒN
        if len(df_van) > 0:
            df_van[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Van_con', index=False
            )

    print(f"✅ Đã tạo báo cáo tuần: {output_file}")
    print(f"\n{'='*80}")
    print(f"✅ HOÀN THÀNH BÁO CÁO TUẦN {week_number}/{year}")
    print(f"{'='*80}\n")

    return output_file


def generate_monthly_report(year, month, output_file=None):
    """
    Tạo báo cáo so sánh theo tháng

    Args:
        year: Năm (VD: 2025)
        month: Tháng (1-12)
        output_file: Đường dẫn file Excel output (tùy chọn)

    Returns:
        Đường dẫn file Excel đã tạo
    """
    print(f"\n{'='*80}")
    print(f"TẠO BÁO CÁO THÁNG {month}/{year}")
    print(f"{'='*80}\n")

    db_path = "suy_hao_history.db"
    if not os.path.exists(db_path):
        print(f"❌ Không tìm thấy database: {db_path}")
        return None

    conn = sqlite3.connect(db_path)

    # Tháng hiện tại
    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = datetime(year, month + 1, 1) - timedelta(days=1)

    # Tháng trước
    if month == 1:
        prev_month_start = datetime(year - 1, 12, 1)
        prev_month_end = datetime(year, 1, 1) - timedelta(days=1)
    else:
        prev_month_start = datetime(year, month - 1, 1)
        prev_month_end = datetime(year, month, 1) - timedelta(days=1)

    print(f"✓ Tháng hiện tại: {month_start.strftime('%d/%m/%Y')} - {month_end.strftime('%d/%m/%Y')}")
    print(f"✓ Tháng trước: {prev_month_start.strftime('%d/%m/%Y')} - {prev_month_end.strftime('%d/%m/%Y')}")

    # Lấy danh sách thuê bao tháng hiện tại
    df_current = pd.read_sql_query(f"""
        SELECT DISTINCT account_cts, doi_one, nvkt_db_normalized, sa, ten_tb_one
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{month_start.strftime('%Y-%m-%d')}'
          AND '{month_end.strftime('%Y-%m-%d')}'
    """, conn)

    # Lấy danh sách thuê bao tháng trước
    df_previous = pd.read_sql_query(f"""
        SELECT DISTINCT account_cts, doi_one, nvkt_db_normalized, sa, ten_tb_one
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{prev_month_start.strftime('%Y-%m-%d')}'
          AND '{prev_month_end.strftime('%Y-%m-%d')}'
    """, conn)

    print(f"\n✓ Tháng hiện tại: {len(df_current)} thuê bao")
    print(f"✓ Tháng trước: {len(df_previous)} thuê bao")

    # Phân loại biến động
    accounts_current = set(df_current['account_cts'].tolist())
    accounts_previous = set(df_previous['account_cts'].tolist())

    tang_moi = accounts_current - accounts_previous
    giam_het = accounts_previous - accounts_current
    van_con = accounts_current & accounts_previous

    print(f"\n✓ Phân tích biến động:")
    print(f"  🆕 TĂNG MỚI: {len(tang_moi)} thuê bao")
    print(f"  ⬇️  GIẢM/HẾT: {len(giam_het)} thuê bao")
    print(f"  ↔️  VẪN CÒN: {len(van_con)} thuê bao")

    # Tạo DataFrames
    df_tang = df_current[df_current['account_cts'].isin(tang_moi)].copy()
    df_giam = df_previous[df_previous['account_cts'].isin(giam_het)].copy()
    df_van = df_current[df_current['account_cts'].isin(van_con)].copy()

    # Tổng hợp theo NVKT_DB
    summary_current = df_current.groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_luong_thang_nay')
    summary_previous = df_previous.groupby(['doi_one', 'nvkt_db_normalized']).size().reset_index(name='so_luong_thang_truoc')

    df_summary = summary_current.merge(
        summary_previous,
        on=['doi_one', 'nvkt_db_normalized'],
        how='outer'
    ).fillna(0)

    df_summary['tang_giam'] = df_summary['so_luong_thang_nay'] - df_summary['so_luong_thang_truoc']
    df_summary['ty_le_thay_doi'] = (df_summary['tang_giam'] / df_summary['so_luong_thang_truoc'] * 100).round(1)
    df_summary['ty_le_thay_doi'] = df_summary['ty_le_thay_doi'].replace([float('inf'), -float('inf')], 0)

    # Lấy thông tin tỉ lệ SHC từ suy_hao_daily_summary (ngày cuối tháng)
    df_ratio = pd.read_sql_query(f"""
        SELECT doi_one, nvkt_db_normalized, so_tb_quan_ly, ty_le_shc
        FROM suy_hao_daily_summary
        WHERE ngay_bao_cao = (SELECT MAX(ngay_bao_cao) FROM suy_hao_daily_summary 
                              WHERE ngay_bao_cao <= '{month_end.strftime('%Y-%m-%d')}')
    """, conn)
    
    if len(df_ratio) > 0:
        df_summary = df_summary.merge(
            df_ratio,
            on=['doi_one', 'nvkt_db_normalized'],
            how='left'
        )
        df_summary['so_tb_quan_ly'] = df_summary['so_tb_quan_ly'].fillna(0).astype(int)
        df_summary['ty_le_shc'] = df_summary['ty_le_shc'].fillna(0)
        df_summary.columns = ['Đơn vị', 'NVKT_DB', 'Tháng này', 'Tháng trước', 'Tăng/Giảm', '% Thay đổi', 'Số TB quản lý', 'Tỉ lệ SHC (%)']
    else:
        df_summary.columns = ['Đơn vị', 'NVKT_DB', 'Tháng này', 'Tháng trước', 'Tăng/Giảm', '% Thay đổi']
    
    df_summary = df_summary.sort_values(by=['Đơn vị', 'NVKT_DB'])

    # Thống kê theo ngày trong tháng
    print(f"\n✓ Tạo xu hướng theo ngày...")
    df_daily = pd.read_sql_query(f"""
        SELECT
            ngay_bao_cao,
            COUNT(DISTINCT account_cts) as so_luong
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{month_start.strftime('%Y-%m-%d')}'
          AND '{month_end.strftime('%Y-%m-%d')}'
        GROUP BY ngay_bao_cao
        ORDER BY ngay_bao_cao
    """, conn)

    df_daily['ngay_bao_cao'] = pd.to_datetime(df_daily['ngay_bao_cao']).dt.strftime('%d/%m')
    df_daily.columns = ['Ngày', 'Số lượng TB suy hao']

    conn.close()

    # Tạo file Excel
    if output_file is None:
        output_file = f"downloads/baocao_hanoi/Bao_cao_thang_{month:02d}_{year}.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"\n✓ Đang ghi file Excel: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet tổng hợp
        df_summary.to_excel(writer, sheet_name='Tong_hop', index=False)

        # Sheet xu hướng theo ngày
        df_daily.to_excel(writer, sheet_name='Xu_huong_theo_ngay', index=False)

        # Sheet chi tiết TĂNG MỚI
        if len(df_tang) > 0:
            df_tang[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Tang_moi', index=False
            )

        # Sheet chi tiết GIẢM/HẾT
        if len(df_giam) > 0:
            df_giam[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Giam_het', index=False
            )

        # Sheet VẪN CÒN
        if len(df_van) > 0:
            df_van[['account_cts', 'ten_tb_one', 'doi_one', 'nvkt_db_normalized', 'sa']].to_excel(
                writer, sheet_name='Van_con', index=False
            )

    print(f"✅ Đã tạo báo cáo tháng: {output_file}")
    print(f"\n{'='*80}")
    print(f"✅ HOÀN THÀNH BÁO CÁO THÁNG {month}/{year}")
    print(f"{'='*80}\n")

    return output_file


def generate_trend_report(start_date, end_date, output_file=None):
    """
    Tạo báo cáo xu hướng tùy chỉnh theo khoảng thời gian

    Args:
        start_date: Ngày bắt đầu (format: 'YYYY-MM-DD' hoặc datetime)
        end_date: Ngày kết thúc (format: 'YYYY-MM-DD' hoặc datetime)
        output_file: Đường dẫn file Excel output (tùy chọn)

    Returns:
        Đường dẫn file Excel đã tạo
    """
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

    print(f"\n{'='*80}")
    print(f"TẠO BÁO CÁO XU HƯỚNG: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
    print(f"{'='*80}\n")

    db_path = "suy_hao_history.db"
    if not os.path.exists(db_path):
        print(f"❌ Không tìm thấy database: {db_path}")
        return None

    conn = sqlite3.connect(db_path)

    # Xu hướng theo ngày
    df_daily = pd.read_sql_query(f"""
        SELECT
            ngay_bao_cao,
            doi_one,
            COUNT(DISTINCT account_cts) as so_luong
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{start_date.strftime('%Y-%m-%d')}'
          AND '{end_date.strftime('%Y-%m-%d')}'
        GROUP BY ngay_bao_cao, doi_one
        ORDER BY ngay_bao_cao, doi_one
    """, conn)

    df_daily['ngay_bao_cao'] = pd.to_datetime(df_daily['ngay_bao_cao']).dt.strftime('%d/%m/%Y')
    df_daily.columns = ['Ngày', 'Đơn vị', 'Số lượng']

    # Pivot để hiển thị theo đơn vị (đơn vị làm cột index, ngày làm cột dữ liệu)
    df_pivot = df_daily.pivot(index='Đơn vị', columns='Ngày', values='Số lượng').fillna(0).astype(int)
    df_pivot = df_pivot.reset_index()

    # Biến động theo NVKT_DB
    df_nvkt_trend = pd.read_sql_query(f"""
        SELECT
            ngay_bao_cao,
            doi_one,
            nvkt_db_normalized,
            COUNT(DISTINCT account_cts) as so_luong
        FROM suy_hao_snapshots
        WHERE ngay_bao_cao BETWEEN '{start_date.strftime('%Y-%m-%d')}'
          AND '{end_date.strftime('%Y-%m-%d')}'
        GROUP BY ngay_bao_cao, doi_one, nvkt_db_normalized
        ORDER BY doi_one, nvkt_db_normalized, ngay_bao_cao
    """, conn)

    conn.close()

    # Pivot để hiển thị theo NVKT_DB với cột Đơn vị
    df_nvkt_trend['ngay_bao_cao'] = pd.to_datetime(df_nvkt_trend['ngay_bao_cao']).dt.strftime('%d/%m')
    df_nvkt_pivot = df_nvkt_trend.pivot_table(
        index=['doi_one', 'nvkt_db_normalized'], 
        columns='ngay_bao_cao', 
        values='so_luong', 
        aggfunc='sum',
        fill_value=0
    ).astype(int)
    df_nvkt_pivot = df_nvkt_pivot.reset_index()
    df_nvkt_pivot = df_nvkt_pivot.rename(columns={'doi_one': 'Đơn vị', 'nvkt_db_normalized': 'NVKT'})
    df_nvkt_pivot = df_nvkt_pivot.sort_values(by=['Đơn vị', 'NVKT'])

    # Tạo file Excel
    if output_file is None:
        output_file = f"downloads/baocao_hanoi/Bao_cao_xu_huong_SHC_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    print(f"\n✓ Đang ghi file Excel: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet xu hướng theo đơn vị
        df_pivot.to_excel(writer, sheet_name='Xu_huong_theo_don_vi', index=False)

        # Sheet xu hướng theo NVKT_DB (pivot table)
        df_nvkt_pivot.to_excel(writer, sheet_name='Xu_huong_theo_NVKT', index=False)
        
        # Lấy workbook để thêm biểu đồ
        workbook = writer.book
        
        # Import thêm các module cần thiết cho chart
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Tạo biểu đồ cho từng đơn vị
        units = df_nvkt_pivot['Đơn vị'].unique()
        date_cols = [col for col in df_nvkt_pivot.columns if col not in ['Đơn vị', 'NVKT']]
        
        for unit_name in units:
            # Lọc dữ liệu theo đơn vị
            df_unit = df_nvkt_pivot[df_nvkt_pivot['Đơn vị'] == unit_name].copy()
            
            if df_unit.empty:
                continue
            
            # Tạo sheet mới cho biểu đồ
            short_name = unit_name.replace('Tổ Kỹ thuật Địa bàn ', '')
            sheet_name = f'Bieu_do_{short_name[:10]}'  # Giới hạn 31 ký tự
            
            # Tạo sheet mới với dữ liệu
            ws = workbook.create_sheet(title=sheet_name)
            
            # Ghi dữ liệu vào sheet
            # Header: NVKT + các ngày
            headers = ['NVKT'] + list(date_cols)
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=str(header))
            
            # Data rows
            for row_idx, (_, row) in enumerate(df_unit.iterrows(), 2):
                ws.cell(row=row_idx, column=1, value=row['NVKT'])
                for col_idx, date_col in enumerate(date_cols, 2):
                    ws.cell(row=row_idx, column=col_idx, value=row[date_col])
            
            # Tạo biểu đồ bar
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = f"KẾT QUẢ XỬ LÝ SUY HAO CAO - {short_name}"
            chart.y_axis.title = "Số TB suy hao cao"
            chart.x_axis.title = None  # Không cần tiêu đề vì đã có tên NVKT
            chart.x_axis.tickLblPos = "low"  # Hiển thị nhãn ở dưới
            
            # Số hàng dữ liệu
            num_rows = len(df_unit) + 1
            num_cols = len(date_cols) + 1
            
            # Data reference (các cột ngày)
            data = Reference(ws, min_col=2, min_row=1, max_col=num_cols, max_row=num_rows)
            
            # Category reference (tên NVKT) - dùng cách thông thường
            cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            chart.shape = 4
            chart.width = 25  # Rộng hơn để tên NVKT không bị chồng
            chart.height = 12
            
            # Thêm data labels - chỉ hiện số lượng
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showSerName = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.showLegendKey = False
            
            # Đặt biểu đồ vào vị trí phù hợp (sau bảng dữ liệu)
            ws.add_chart(chart, f"A{num_rows + 3}")
        
        print(f"   ✅ Đã tạo biểu đồ cho {len(units)} đơn vị")

    print(f"✅ Đã tạo báo cáo xu hướng: {output_file}")
    
    # =========================================================================
    # TẠO BIỂU ĐỒ DẠNG ẢNH
    # =========================================================================
    import matplotlib.pyplot as plt
    import numpy as np
    
    chart_folder = "downloads/baocao_hanoi/shc_NVKT_chart"
    os.makedirs(chart_folder, exist_ok=True)
    
    print(f"\n📊 Tạo biểu đồ dạng ảnh...")
    
    for unit_name in df_nvkt_pivot['Đơn vị'].unique():
        df_unit = df_nvkt_pivot[df_nvkt_pivot['Đơn vị'] == unit_name].copy()
        
        if df_unit.empty:
            continue
        
        short_name = unit_name.replace('Tổ Kỹ thuật Địa bàn ', '')
        nvkt_list = df_unit['NVKT'].values
        date_cols = [col for col in df_unit.columns if col not in ['Đơn vị', 'NVKT']]
        
        # Tạo figure
        fig, ax = plt.subplots(figsize=(16, 8))
        
        x = np.arange(len(nvkt_list))
        n_dates = len(date_cols)
        width = 0.8 / n_dates
        
        # Màu sắc - tạo danh sách màu đảm bảo 2 bar cạnh nhau không trùng màu
        # Sử dụng 2 bảng màu xen kẽ để tạo sự khác biệt rõ ràng
        base_colors = [
            '#1f77b4',  # blue
            '#ff7f0e',  # orange
            '#2ca02c',  # green
            '#d62728',  # red
            '#9467bd',  # purple
            '#8c564b',  # brown
            '#e377c2',  # pink
            '#7f7f7f',  # gray
            '#bcbd22',  # olive
            '#17becf',  # cyan
            '#393b79',  # dark blue
            '#637939',  # dark green
            '#8c6d31',  # dark orange
            '#843c39',  # dark red
            '#7b4173',  # dark purple
            '#5254a3',  # indigo
            '#6b6ecf',  # light purple
            '#9c9ede',  # lavender
            '#bd9e39',  # gold
            '#ad494a',  # salmon
            '#8ca252',  # light green
            '#ce6dbd',  # magenta
            '#de9ed6',  # light pink
            '#3182bd',  # steel blue
            '#e6550d',  # dark orange
            '#31a354',  # emerald
        ]
        # Lặp lại nếu cần nhiều hơn số màu có sẵn
        colors = [base_colors[i % len(base_colors)] for i in range(n_dates)]
        
        # Vẽ từng ngày
        for i, date_col in enumerate(date_cols):
            values = df_unit[date_col].values
            offset = (i - n_dates/2 + 0.5) * width
            bars = ax.bar(x + offset, values, width, label=str(date_col), color=colors[i])
            
            # Thêm giá trị lên cột
            for bar, val in zip(bars, values):
                if val > 0:
                    ax.annotate(f'{int(val)}',
                               xy=(bar.get_x() + bar.get_width() / 2, bar.get_height()),
                               xytext=(0, 1),
                               textcoords="offset points",
                               ha='center', va='bottom',
                               fontsize=7, fontweight='bold')
        
        ax.set_xlabel('NVKT', fontsize=12)
        ax.set_ylabel('Số TB suy hao cao', fontsize=12)
        ax.set_title(f'KẾT QUẢ XỬ LÝ SUY HAO CAO - {short_name}\n({start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")})', 
                     fontsize=14, fontweight='bold', pad=15)
        ax.set_xticks(x)
        ax.set_xticklabels(nvkt_list, rotation=45, ha='right', fontsize=10)
        ax.legend(title='Ngày', loc='upper right', fontsize=9, ncol=2)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)
        
        plt.tight_layout()
        
        # Lưu file
        safe_name = short_name.replace(' ', '_')
        chart_path = os.path.join(chart_folder, f"SHC_{safe_name}.png")
        plt.savefig(chart_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        print(f"   ✅ {chart_path}")
    
    print(f"\n{'='*80}")
    print(f"✅ HOÀN THÀNH BÁO CÁO XU HƯỚNG")
    print(f"{'='*80}\n")

    return output_file


if __name__ == "__main__":
    # Test các hàm
    print("Test module suy_hao_reports.py")

    # Báo cáo so sánh SHC ngày (T-1)
    generate_daily_comparison_report()

    # Lấy ngày dữ liệu mới nhất từ database
    db_path = "suy_hao_history.db"
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        latest_date_query = """
            SELECT MAX(ngay_bao_cao) as latest_date
            FROM suy_hao_snapshots
        """
        df_latest = pd.read_sql_query(latest_date_query, conn)
        conn.close()

        if not df_latest.empty and df_latest['latest_date'][0]:
            latest_date = datetime.strptime(df_latest['latest_date'][0], '%Y-%m-%d')
            month_start = datetime(latest_date.year, latest_date.month, 1)
        else:
            # Nếu database trống, sử dụng tháng hiện tại
            today = datetime.now()
            month_start = datetime(today.year, today.month, 1)
            latest_date = today
    else:
        # Nếu không có database, sử dụng ngày hôm nay
        today = datetime.now()
        month_start = datetime(today.year, today.month, 1)
        latest_date = today

    # Test báo cáo tuần (tuần hiện tại)
    current_week = latest_date.isocalendar()[1]
    current_year = latest_date.year
    generate_weekly_report(current_year, current_week)

    # Test báo cáo tháng (tháng hiện tại)
    generate_monthly_report(latest_date.year, latest_date.month)

    # Test báo cáo xu hướng (từ đầu tháng đến ngày dữ liệu mới nhất)
    generate_trend_report(month_start.strftime('%Y-%m-%d'), latest_date.strftime('%Y-%m-%d'))
