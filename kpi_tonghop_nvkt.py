"""
Module tổng hợp KPI cho NVKT (Nhân viên Kỹ thuật) từ nhiều nguồn dữ liệu.

Tổng hợp 9 nguồn dữ liệu KPI thành 1 file Excel duy nhất với multi-level header
(2 dòng: dòng 1 là nhóm chỉ tiêu, dòng 2 là chỉ tiêu con).

Nguồn dữ liệu:
    1. C1.1 - Chất lượng sửa chữa dịch vụ BRCĐ
       File: KPI-DOWNLOAD/c11-nvktdb report.xlsx | Sheet: c11 kpi nvkt
       Cột: SM1, SM2, Tỷ lệ SC CL chủ động (%), SM3, SM4, Tỷ lệ SC BH đúng QĐ (%), Điểm BSC

    2. C1.2 - Tỷ lệ thuê bao báo hỏng lặp lại
       File: KPI-DOWNLOAD/c12-nvktdb report.xlsx | Sheet: c12 kpi nvkt
       Cột: SM1, SM2, Tỷ lệ BH lặp lại (%), SM3, SM4, Tỷ lệ sự cố BRCĐ (%), Điểm BSC

    3. C1.4 - Độ hài lòng khách hàng
       File: downloads/baocao_hanoi/c1.4_chitiet_report.xlsx | Sheet: TH_HL_NVKT
       Cột: Phiếu KS thành công, Phiếu KHL, Tỉ lệ HL (%)

    4. C1.5 - Tỉ lệ thiết lập dịch vụ đạt thời gian
       File: downloads/baocao_hanoi/c1.5_chitiet_report.xlsx | Sheet: KQ_C15_chitiet
       Cột: Phiếu đạt, Phiếu không đạt, Tổng Hoàn công, Tỉ lệ đạt (%)

    5. KQ Tiếp thị - Kết quả tiếp thị dịch vụ BRCĐ và MyTV
       File: KQ-TIEP-THI/kq_tiep_thi_ddmmyyyy.xlsx (file mới nhất) | Sheet: kq_tiep_thi
       Cột: Dịch vụ BRCĐ, Dịch vụ MyTV, Tổng

    6. GHTT - Giao hoàn toàn trình (tỷ lệ hoàn thành phiếu giao việc)
       File: GHTT/tong_hop_ghtt_nvktdb.xlsx | Sheet: kq_nvktdb
       Cột: HT T, Giao T, Tỷ lệ T, Điểm T, Tỷ lệ T+1, Điểm T+1, Điểm Tổng

    7. Fiber thực tăng - Phát triển thuê bao Fiber
       File: PTTB-PSC/fiber_thuc_tang_ddmmyyyy.xlsx (file mới nhất) | Sheet: thuc_tang_theo_NVKT
       Cột: Hoàn công, Ngưng PSC, Thực tăng, Tỷ lệ ngưng/psc

    8. MyTV thực tăng - Phát triển thuê bao MyTV
       File: PTTB-PSC/mytv_thuc_tang_ddmmyyyy.xlsx (file mới nhất) | Sheet: thuc_tang_theo_NVKT
       Cột: Hoàn công, Ngưng PSC, Thực tăng, Tỷ lệ ngưng/psc

    9. I1.5 - Tỷ lệ suy hao cao trên đường dây
       File: downloads/baocao_hanoi/I1.5 report.xlsx | Sheet: TH_SHC_I15
       Cột: Số TB SHC K1, Số TB quản lý, Tỉ lệ SHC (%)

Quy trình xử lý:
    1. Đọc danh sách NVKT master từ dsnv.xlsx, lọc chỉ các Tổ kỹ thuật
       (Tổ Phúc Thọ, Tổ Quảng Oai, Tổ Suối Hai, Tổ Sơn Tây)
    2. Đọc từng nguồn dữ liệu, chuẩn hóa tên NVKT (Title Case, bỏ ngoặc đơn)
    3. Left merge tuần tự từ master list để đảm bảo tất cả NVKT đều có mặt
       (nguồn nào thiếu NVKT thì cột tương ứng sẽ là NaN)
    4. Ghi ra file Excel với multi-level header, freeze panes, auto-width

Output:
    File: KPI_TongHop_NVKT.xlsx (thư mục gốc dự án)
    Sheet: TongHop_KPI
    Cấu trúc: 45 cột (3 cột cố định + 42 cột KPI), ~40 NVKT

Cách chạy:
    python3 kpi_tonghop_nvkt.py
"""

import os
import glob
import re
import pandas as pd
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def tim_file_moi_nhat(thu_muc: str, pattern: str) -> str | None:
    """Tìm file mới nhất theo pattern trong thư mục.

    Sử dụng glob để tìm các file khớp pattern, sau đó chọn file có thời gian
    sửa đổi (mtime) gần nhất. Dùng cho các file có tên chứa ngày tháng
    (vd: kq_tiep_thi_04032026.xlsx, fiber_thuc_tang_04032026.xlsx).

    Args:
        thu_muc: Thư mục con tương đối từ BASE_DIR (vd: "KQ-TIEP-THI", "PTTB-PSC").
        pattern: Glob pattern để tìm file (vd: "kq_tiep_thi_*.xlsx").

    Returns:
        Đường dẫn tuyệt đối đến file mới nhất, hoặc None nếu không tìm thấy.
    """
    full_pattern = os.path.join(BASE_DIR, thu_muc, pattern)
    files = glob.glob(full_pattern)
    if not files:
        print(f"   ❌ Không tìm thấy file: {full_pattern}")
        return None
    latest = max(files, key=os.path.getmtime)
    print(f"   ✓ Tìm thấy: {os.path.basename(latest)}")
    return latest


def chuan_hoa_ten(name) -> str:
    """Chuẩn hóa tên NVKT để dùng làm join key thống nhất giữa các nguồn.

    Xử lý: strip khoảng trắng → bỏ nội dung trong ngoặc đơn → chuyển Title Case.
    Ví dụ: "nguyễn văn A(PTO)" → "Nguyễn Văn A"
           "  Chu Minh Tám  " → "Chu Minh Tám"

    Args:
        name: Tên NVKT thô từ file nguồn (có thể chứa ngoặc đơn, khoảng trắng thừa).

    Returns:
        Tên đã chuẩn hóa dạng Title Case, hoặc chuỗi rỗng nếu input không hợp lệ.
    """
    if not isinstance(name, str) or not name.strip():
        return ""
    name = name.strip()
    name = re.sub(r'\([^)]*\)', '', name).strip()
    name = name.title()
    return name


# ============================================================
# 9 HÀM ĐỌC DỮ LIỆU
# ============================================================

def doc_c11() -> pd.DataFrame | None:
    """Đọc C1.1 - Chất lượng sửa chữa dịch vụ BRCĐ.

    File: KPI-DOWNLOAD/c11-nvktdb report.xlsx | Sheet: c11 kpi nvkt
    Bỏ dòng "Tổng" (dòng tổng hợp toàn bộ). Truy cập cột theo index vì tên cột
    gốc rất dài (vd: "Tỷ lệ sửa chữa phiếu chất lượng chủ động dịch vụ FiberVNN, MyTV đạt yêu cầu").

    Returns:
        DataFrame với cột: nvkt, c11_sm1, c11_sm2, c11_ty_le_cl, c11_sm3,
        c11_sm4, c11_ty_le_bh, c11_bsc. Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "KPI-DOWNLOAD", "c11-nvktdb report.xlsx")
        df = pd.read_excel(file_path, sheet_name="c11 kpi nvkt")
        df = df[df['NVKT'] != 'Tổng'].copy()
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        cols = df.columns.tolist()
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'c11_sm1': df[cols[3]],
            'c11_sm2': df[cols[4]],
            'c11_ty_le_cl': df[cols[5]],
            'c11_sm3': df[cols[6]],
            'c11_sm4': df[cols[7]],
            'c11_ty_le_bh': df[cols[8]],
            'c11_bsc': df[cols[9]],
        })
        print(f"   ✓ C1.1: {len(result)} NVKT")
        return result
    except FileNotFoundError:
        print(f"   ❌ Không tìm thấy file C1.1")
        return None
    except Exception as e:
        print(f"   ❌ Lỗi đọc C1.1: {e}")
        return None


def doc_c12() -> pd.DataFrame | None:
    """Đọc C1.2 - Tỷ lệ thuê bao báo hỏng dịch vụ BRCĐ lặp lại.

    File: KPI-DOWNLOAD/c12-nvktdb report.xlsx | Sheet: c12 kpi nvkt
    Cấu trúc giống C1.1 (10 cột, dòng đầu là "Tổng").

    Returns:
        DataFrame với cột: nvkt, c12_sm1, c12_sm2, c12_ty_le_ll, c12_sm3,
        c12_sm4, c12_ty_le_sc, c12_bsc. Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "KPI-DOWNLOAD", "c12-nvktdb report.xlsx")
        df = pd.read_excel(file_path, sheet_name="c12 kpi nvkt")
        df = df[df['NVKT'] != 'Tổng'].copy()
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        cols = df.columns.tolist()
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'c12_sm1': df[cols[3]],
            'c12_sm2': df[cols[4]],
            'c12_ty_le_ll': df[cols[5]],
            'c12_sm3': df[cols[6]],
            'c12_sm4': df[cols[7]],
            'c12_ty_le_sc': df[cols[8]],
            'c12_bsc': df[cols[9]],
        })
        print(f"   ✓ C1.2: {len(result)} NVKT")
        return result
    except FileNotFoundError:
        print(f"   ❌ Không tìm thấy file C1.2")
        return None
    except Exception as e:
        print(f"   ❌ Lỗi đọc C1.2: {e}")
        return None


def doc_c14() -> pd.DataFrame | None:
    """Đọc C1.4 - Độ hài lòng khách hàng.

    File: downloads/baocao_hanoi/c1.4_chitiet_report.xlsx | Sheet: TH_HL_NVKT
    Gộp các dòng trùng tên NVKT (sum phiếu KS và KHL), sau đó tính lại tỉ lệ
    hài lòng = (KS - KHL) / KS * 100.

    Returns:
        DataFrame với cột: nvkt, c14_phieu_ks, c14_phieu_khl, c14_ty_le_hl.
        Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "downloads", "baocao_hanoi", "c1.4_chitiet_report.xlsx")
        df = pd.read_excel(file_path, sheet_name="TH_HL_NVKT")
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        # Gộp các dòng trùng tên
        df_agg = df.groupby('nvkt', as_index=False).agg({
            'Tổng phiếu KS thành công': 'sum',
            'Tổng phiếu KHL': 'sum',
        })
        df_agg['c14_ty_le_hl'] = df_agg.apply(
            lambda r: (r['Tổng phiếu KS thành công'] - r['Tổng phiếu KHL']) / r['Tổng phiếu KS thành công'] * 100
            if r['Tổng phiếu KS thành công'] > 0 else 0, axis=1
        ).round(2)
        result = pd.DataFrame({
            'nvkt': df_agg['nvkt'],
            'c14_phieu_ks': df_agg['Tổng phiếu KS thành công'],
            'c14_phieu_khl': df_agg['Tổng phiếu KHL'],
            'c14_ty_le_hl': df_agg['c14_ty_le_hl'],
        })
        print(f"   ✓ C1.4: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc C1.4: {e}")
        return None


def doc_c15() -> pd.DataFrame | None:
    """Đọc C1.5 - Tỉ lệ thiết lập dịch vụ đạt thời gian.

    File: downloads/baocao_hanoi/c1.5_chitiet_report.xlsx | Sheet: KQ_C15_chitiet
    Đo lường tỷ lệ phiếu hoàn công đạt thời gian quy định.

    Returns:
        DataFrame với cột: nvkt, c15_phieu_dat, c15_phieu_kdat, c15_tong_hc,
        c15_ty_le_dat. Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "downloads", "baocao_hanoi", "c1.5_chitiet_report.xlsx")
        df = pd.read_excel(file_path, sheet_name="KQ_C15_chitiet")
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'c15_phieu_dat': df['Phiếu đạt'],
            'c15_phieu_kdat': df['Phiếu không đạt'],
            'c15_tong_hc': df['Tổng Hoàn công'],
            'c15_ty_le_dat': df['Tỉ lệ đạt (%)'],
        })
        print(f"   ✓ C1.5: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc C1.5: {e}")
        return None


def doc_kq_tiep_thi() -> pd.DataFrame | None:
    """Đọc KQ Tiếp thị - Kết quả tiếp thị dịch vụ BRCĐ và MyTV.

    File: KQ-TIEP-THI/kq_tiep_thi_ddmmyyyy.xlsx (mới nhất) | Sheet: kq_tiep_thi
    Tự động tìm file mới nhất theo mtime. Gộp các dòng trùng tên NVKT (sum)
    vì cùng 1 NVKT có thể xuất hiện nhiều lần trong sheet gốc.

    Returns:
        DataFrame với cột: nvkt, kqtt_brcd, kqtt_mytv, kqtt_tong.
        Hoặc None nếu lỗi hoặc không tìm thấy file.
    """
    try:
        file_path = tim_file_moi_nhat("KQ-TIEP-THI", "kq_tiep_thi_*.xlsx")
        if not file_path:
            return None
        df = pd.read_excel(file_path, sheet_name="kq_tiep_thi")
        df['nvkt'] = df['Tên NV'].apply(chuan_hoa_ten)
        # Gộp các dòng trùng tên (cùng NVKT có thể xuất hiện nhiều lần)
        df_agg = df.groupby('nvkt', as_index=False).agg({
            'Dịch vụ BRCĐ': 'sum',
            'Dịch vụ MyTV': 'sum',
            'Tổng': 'sum',
        })
        result = pd.DataFrame({
            'nvkt': df_agg['nvkt'],
            'kqtt_brcd': df_agg['Dịch vụ BRCĐ'],
            'kqtt_mytv': df_agg['Dịch vụ MyTV'],
            'kqtt_tong': df_agg['Tổng'],
        })
        print(f"   ✓ KQ Tiếp thị: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc KQ Tiếp thị: {e}")
        return None


def doc_ghtt() -> pd.DataFrame | None:
    """Đọc GHTT - Giao hoàn toàn trình (tỷ lệ hoàn thành phiếu giao việc).

    File: GHTT/tong_hop_ghtt_nvktdb.xlsx | Sheet: kq_nvktdb
    Chứa dữ liệu hoàn thành phiếu công việc theo 2 kỳ T và T+1, bao gồm
    số hoàn thành, số giao, tỷ lệ, điểm và điểm tổng hợp.

    Returns:
        DataFrame với cột: nvkt, ghtt_ht_t, ghtt_giao_t, ghtt_ty_le_t,
        ghtt_diem_t, ghtt_ty_le_t1, ghtt_diem_t1, ghtt_diem_tong.
        Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "GHTT", "tong_hop_ghtt_nvktdb.xlsx")
        df = pd.read_excel(file_path, sheet_name="kq_nvktdb")
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'ghtt_ht_t': df['Hoàn thành T'],
            'ghtt_giao_t': df['Giao NVKT T'],
            'ghtt_ty_le_t': df['Tỷ lệ T'],
            'ghtt_diem_t': df['Điểm T'],
            'ghtt_ty_le_t1': df['Tỷ lệ T+1'],
            'ghtt_diem_t1': df['Điểm T+1'],
            'ghtt_diem_tong': df['ĐIểm Tổng'],
        })
        print(f"   ✓ GHTT: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc GHTT: {e}")
        return None


def doc_fiber() -> pd.DataFrame | None:
    """Đọc Fiber thực tăng - Phát triển thuê bao FiberVNN.

    File: PTTB-PSC/fiber_thuc_tang_ddmmyyyy.xlsx (mới nhất) | Sheet: thuc_tang_theo_NVKT
    Đo lường số hoàn công, số ngưng phát sinh cước, thực tăng ròng và tỷ lệ ngưng/psc.

    Returns:
        DataFrame với cột: nvkt, fiber_hc, fiber_ngung, fiber_thuc_tang, fiber_ty_le.
        Hoặc None nếu lỗi hoặc không tìm thấy file.
    """
    try:
        file_path = tim_file_moi_nhat("PTTB-PSC", "fiber_thuc_tang_*.xlsx")
        if not file_path:
            return None
        df = pd.read_excel(file_path, sheet_name="thuc_tang_theo_NVKT")
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'fiber_hc': df['Hoàn công'],
            'fiber_ngung': df['Ngưng phát sinh cước'],
            'fiber_thuc_tang': df['Thực tăng'],
            'fiber_ty_le': df['Tỷ lệ ngưng/psc'],
        })
        print(f"   ✓ Fiber thực tăng: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc Fiber thực tăng: {e}")
        return None


def doc_mytv() -> pd.DataFrame | None:
    """Đọc MyTV thực tăng - Phát triển thuê bao MyTV.

    File: PTTB-PSC/mytv_thuc_tang_ddmmyyyy.xlsx (mới nhất) | Sheet: thuc_tang_theo_NVKT
    Cấu trúc giống Fiber thực tăng.

    Returns:
        DataFrame với cột: nvkt, mytv_hc, mytv_ngung, mytv_thuc_tang, mytv_ty_le.
        Hoặc None nếu lỗi hoặc không tìm thấy file.
    """
    try:
        file_path = tim_file_moi_nhat("PTTB-PSC", "mytv_thuc_tang_*.xlsx")
        if not file_path:
            return None
        df = pd.read_excel(file_path, sheet_name="thuc_tang_theo_NVKT")
        df['nvkt'] = df['NVKT'].apply(chuan_hoa_ten)
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'mytv_hc': df['Hoàn công'],
            'mytv_ngung': df['Ngưng phát sinh cước'],
            'mytv_thuc_tang': df['Thực tăng'],
            'mytv_ty_le': df['Tỷ lệ ngưng/psc'],
        })
        print(f"   ✓ MyTV thực tăng: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc MyTV thực tăng: {e}")
        return None


def doc_i15() -> pd.DataFrame | None:
    """Đọc I1.5 - Tỷ lệ suy hao cao trên đường dây quang.

    File: downloads/baocao_hanoi/I1.5 report.xlsx | Sheet: TH_SHC_I15
    Cột join key là 'NVKT_DB' (khác với các nguồn khác dùng 'NVKT').
    Đo lường số thuê bao suy hao cao K1 trên tổng số thuê bao quản lý.

    Returns:
        DataFrame với cột: nvkt, i15_tb_shc, i15_tb_ql, i15_ty_le.
        Hoặc None nếu lỗi.
    """
    try:
        file_path = os.path.join(BASE_DIR, "downloads", "baocao_hanoi", "I1.5 report.xlsx")
        df = pd.read_excel(file_path, sheet_name="TH_SHC_I15")
        df['nvkt'] = df['NVKT_DB'].apply(chuan_hoa_ten)
        result = pd.DataFrame({
            'nvkt': df['nvkt'],
            'i15_tb_shc': df['Số TB Suy hao cao K1'],
            'i15_tb_ql': df['Số TB quản lý'],
            'i15_ty_le': df['Tỉ lệ SHC (%)'],
        })
        print(f"   ✓ I1.5: {len(result)} NVKT")
        return result
    except Exception as e:
        print(f"   ❌ Lỗi đọc I1.5: {e}")
        return None


# ============================================================
# HÀM TỔNG HỢP
# ============================================================

# Thứ tự cột trong output (sau STT, Đơn vị, NVKT)
DATA_COLUMNS = [
    'c11_sm1', 'c11_sm2', 'c11_ty_le_cl', 'c11_sm3', 'c11_sm4', 'c11_ty_le_bh', 'c11_bsc',
    'c12_sm1', 'c12_sm2', 'c12_ty_le_ll', 'c12_sm3', 'c12_sm4', 'c12_ty_le_sc', 'c12_bsc',
    'c14_phieu_ks', 'c14_phieu_khl', 'c14_ty_le_hl',
    'c15_phieu_dat', 'c15_phieu_kdat', 'c15_tong_hc', 'c15_ty_le_dat',
    'kqtt_brcd', 'kqtt_mytv', 'kqtt_tong',
    'ghtt_ht_t', 'ghtt_giao_t', 'ghtt_ty_le_t', 'ghtt_diem_t', 'ghtt_ty_le_t1', 'ghtt_diem_t1', 'ghtt_diem_tong',
    'fiber_hc', 'fiber_ngung', 'fiber_thuc_tang', 'fiber_ty_le',
    'mytv_hc', 'mytv_ngung', 'mytv_thuc_tang', 'mytv_ty_le',
    'i15_tb_shc', 'i15_tb_ql', 'i15_ty_le',
]

# Định nghĩa header level 1 (nhóm) và level 2 (cột con)
HEADER_GROUPS = [
    # (tên nhóm, số cột, danh sách tên cột con)
    ("C1.1 - CL Sửa chữa", 7, ["SM1", "SM2", "Tỷ lệ CL\nchủ động (%)", "SM3", "SM4", "Tỷ lệ SC BH\nđúng QĐ (%)", "Điểm BSC"]),
    ("C1.2 - BH lặp lại", 7, ["SM1", "SM2", "Tỷ lệ BH\nlặp lại (%)", "SM3", "SM4", "Tỷ lệ sự cố\nBRCĐ (%)", "Điểm BSC"]),
    ("C1.4 - Độ hài lòng KH", 3, ["Phiếu KS\nthành công", "Phiếu KHL", "Tỉ lệ HL (%)"]),
    ("C1.5 - TL thiết lập DV đạt t/g", 4, ["Phiếu đạt", "Phiếu\nkhông đạt", "Tổng HC", "Tỉ lệ đạt (%)"]),
    ("KQ Tiếp thị", 3, ["DV BRCĐ", "DV MyTV", "Tổng"]),
    ("GHTT", 7, ["HT T", "Giao T", "Tỷ lệ T", "Điểm T", "Tỷ lệ T+1", "Điểm T+1", "Điểm Tổng"]),
    ("Fiber thực tăng", 4, ["Hoàn công", "Ngưng PSC", "Thực tăng", "TL ngưng/psc"]),
    ("MyTV thực tăng", 4, ["Hoàn công", "Ngưng PSC", "Thực tăng", "TL ngưng/psc"]),
    ("I1.5 - Suy hao cao", 3, ["Số TB SHC", "Số TB QL", "Tỉ lệ SHC (%)"]),
]

# Đơn vị thuộc Tổ kỹ thuật
TO_KY_THUAT = ['Tổ Phúc Thọ', 'Tổ Quảng oai', 'Tổ Suối hai', 'Tổ Sơn Tây']


def tong_hop_kpi_nvkt() -> pd.DataFrame | None:
    """Tổng hợp KPI từ 9 nguồn dữ liệu thành 1 DataFrame duy nhất.

    Quy trình:
        1. Đọc master list NVKT từ dsnv.xlsx, lọc chỉ Tổ kỹ thuật
        2. Gọi 9 hàm doc_*(), mỗi hàm trả về DataFrame hoặc None (nếu lỗi)
        3. Left merge tuần tự từ master list → NVKT nào thiếu dữ liệu sẽ có NaN
        4. Sắp xếp theo đơn vị → tên NVKT, đánh STT

    Returns:
        DataFrame 45 cột (stt, don_vi, nvkt + 42 cột KPI), ~40 dòng.
        Hoặc None nếu không đọc được master list.
    """
    print("=" * 60)
    print("TỔNG HỢP KPI NVKT")
    print("=" * 60)

    # 1. Đọc master list
    print("\n[1/3] Đọc danh sách NVKT từ dsnv.xlsx...")
    dsnv_path = os.path.join(BASE_DIR, "dsnv.xlsx")
    df_dsnv = pd.read_excel(dsnv_path)
    df_master = df_dsnv[['Họ tên', 'đơn vị']].copy()
    df_master.columns = ['nvkt', 'don_vi']
    df_master['nvkt'] = df_master['nvkt'].apply(chuan_hoa_ten)
    # Lọc chỉ Tổ kỹ thuật
    df_master = df_master[df_master['don_vi'].isin(TO_KY_THUAT)].copy()
    df_master = df_master[df_master['nvkt'] != ''].copy()
    print(f"   ✓ {len(df_master)} NVKT thuộc Tổ kỹ thuật")

    # 2. Đọc 7 nguồn
    print("\n[2/3] Đọc dữ liệu từ 7 nguồn...")
    sources = {
        'C1.1': doc_c11(),
        'C1.2': doc_c12(),
        'C1.4': doc_c14(),
        'C1.5': doc_c15(),
        'KQ Tiếp thị': doc_kq_tiep_thi(),
        'GHTT': doc_ghtt(),
        'Fiber': doc_fiber(),
        'MyTV': doc_mytv(),
        'I1.5': doc_i15(),
    }

    # 3. Merge tuần tự
    print("\n[3/3] Ghép dữ liệu...")
    df_all = df_master.copy()
    for name, df_src in sources.items():
        if df_src is not None:
            df_all = df_all.merge(df_src, on='nvkt', how='left')
        else:
            print(f"   ⚠ Bỏ qua {name} (không có dữ liệu)")

    # Sắp xếp
    df_all = df_all.sort_values(['don_vi', 'nvkt']).reset_index(drop=True)
    df_all.insert(0, 'stt', range(1, len(df_all) + 1))

    # Đảm bảo thứ tự cột
    final_cols = ['stt', 'don_vi', 'nvkt']
    for col in DATA_COLUMNS:
        if col not in df_all.columns:
            df_all[col] = None
    final_cols.extend(DATA_COLUMNS)
    df_all = df_all[final_cols]

    print(f"\n   ✅ Tổng hợp xong: {len(df_all)} NVKT × {len(df_all.columns)} cột")
    return df_all


# ============================================================
# GHI EXCEL VỚI MULTI-LEVEL HEADER
# ============================================================

def ghi_file_tonghop(df: pd.DataFrame, output_path: str):
    """Ghi DataFrame tổng hợp ra file Excel với multi-level header (2 dòng).

    Cấu trúc header:
        - Row 1: Tên nhóm chỉ tiêu (merge ngang các cột con). Cột STT/Đơn vị/NVKT
          merge dọc 2 dòng.
        - Row 2: Tên chỉ tiêu con trong từng nhóm.
        - Row 3+: Dữ liệu.

    Định dạng:
        - Header: bold, nền xanh nhạt (#DCE6F1), căn giữa, wrap text, border.
        - Dữ liệu: border, căn giữa (trừ cột NVKT căn trái).
        - Freeze panes tại D3 (cố định 3 cột đầu + 2 dòng header).
        - Auto-width theo nội dung (tối đa 25 ký tự).

    Args:
        df: DataFrame từ tong_hop_kpi_nvkt(), 45 cột theo thứ tự DATA_COLUMNS.
        output_path: Đường dẫn file Excel output.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TongHop_KPI"

    # Style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 1-2: Headers
    # Cột 1-3: STT, Đơn vị, NVKT (merge dọc 2 rows)
    fixed_headers = ["STT", "Đơn vị", "NVKT"]
    for col_idx, label in enumerate(fixed_headers, 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        # Border cho cell row 2 (merged)
        ws.cell(row=2, column=col_idx).border = thin_border

    # Cột 4+: Các nhóm KPI
    current_col = 4
    for group_name, num_cols, sub_headers in HEADER_GROUPS:
        # Row 1: merge nhóm
        end_col = current_col + num_cols - 1
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=current_col, value=group_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        # Border cho các cell merged trong row 1
        for c in range(current_col + 1, end_col + 1):
            ws.cell(row=1, column=c).border = thin_border

        # Row 2: sub-headers
        for i, sub in enumerate(sub_headers):
            cell = ws.cell(row=2, column=current_col + i, value=sub)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        current_col = end_col + 1

    # Row 3+: Data
    data_align = Alignment(horizontal='center', vertical='center')
    for row_idx, row_data in enumerate(df.values, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx <= 1 or col_idx >= 4:  # STT và cột số liệu
                cell.alignment = data_align
            elif col_idx == 3:  # NVKT - căn trái
                cell.alignment = Alignment(vertical='center')

    # Auto-width
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                # Tính theo dòng dài nhất (hỗ trợ wrap text)
                lines = str(val).split('\n')
                line_max = max(len(line) for line in lines)
                max_len = max(max_len, line_max)
        width = min(max_len + 3, 25)
        if col_idx == 3:  # NVKT cần rộng hơn
            width = max(width, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 8)

    # Freeze panes: cố định 3 cột đầu + 2 dòng header
    ws.freeze_panes = "D3"

    wb.save(output_path)
    print(f"   ✅ Đã lưu file: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    """Chạy tổng hợp KPI và ghi file output, in thống kê ra console."""
    df = tong_hop_kpi_nvkt()
    if df is not None:
        output_path = os.path.join(BASE_DIR, "KPI_TongHop_NVKT.xlsx")
        ghi_file_tonghop(df, output_path)

        # Thống kê
        print("\n" + "=" * 60)
        print("THỐNG KÊ")
        print("=" * 60)
        print(f"Tổng NVKT: {len(df)}")
        print(f"Tổng cột: {len(df.columns)}")
        print(f"\nTheo đơn vị:")
        for dv, count in df.groupby('don_vi').size().items():
            print(f"   {dv}: {count} NVKT")

        # Kiểm tra dữ liệu thiếu
        print(f"\nDữ liệu thiếu (NaN):")
        for col in DATA_COLUMNS:
            na_count = df[col].isna().sum()
            if na_count > 0:
                print(f"   {col}: {na_count}/{len(df)} thiếu")
    else:
        print("❌ Không thể tổng hợp dữ liệu")


if __name__ == "__main__":
    main()
