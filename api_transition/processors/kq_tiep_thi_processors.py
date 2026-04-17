# -*- coding: utf-8 -*-
"""Processors cho nhom Ket qua tiep thi trong api_transition."""

from pathlib import Path
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side

from api_transition.processors.common import ensure_processed_workbook


DEFAULT_KQ_TIEP_THI_INPUT = (
    Path(__file__).resolve().parent.parent / "downloads" / "kq_tiep_thi" / "kq_tiep_thi report.xlsx"
)
DEFAULT_DSNV_FILE = Path.cwd() / "dsnv.xlsx"


def _resolve_path(input_path):
    path = Path(input_path).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    else:
        path = path.resolve()

    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay file input: {path}")
    return path


def _normalize_person_name(name):
    if pd.isna(name):
        return None

    name = str(name).strip()
    if not name:
        return None

    name = re.sub(r"\s+", " ", name)
    return name.lower().title()


def _normalize_multiindex_columns(columns):
    normalized = []
    for level_1, level_2 in columns:
        a_clean = str(level_1).strip()
        b_clean = str(level_2).strip()
        if "Unnamed" in b_clean or b_clean == a_clean:
            normalized.append((a_clean, ""))
        else:
            normalized.append((a_clean, b_clean))
    return pd.MultiIndex.from_tuples(normalized)


def _build_donvi_lookup(dsnv_file):
    dsnv_path = _resolve_path(dsnv_file)
    df_dsnv = pd.read_excel(dsnv_path)
    df_dsnv.columns = [str(col).strip() for col in df_dsnv.columns]

    if "Họ tên" not in df_dsnv.columns or "đơn vị" not in df_dsnv.columns:
        raise ValueError("File dsnv.xlsx phai co cot 'Họ tên' va 'đơn vị'")

    names = df_dsnv["Họ tên"].apply(_normalize_person_name)
    units = df_dsnv["đơn vị"].astype(str).str.strip()
    return dict(zip(names, units))


def _get_required_column(df, level_1, level_2_contains):
    for col in df.columns:
        if str(col[0]).strip() == level_1 and level_2_contains in str(col[1]).strip():
            return col
    raise ValueError(
        f"Khong tim thay cot yeu cau: ({level_1}, {level_2_contains})"
    )


def _write_sheet_with_manual_header(writer, sheet_name, df):
    df_for_save = df.copy()
    header_labels = [col[0] if isinstance(col, tuple) else str(col) for col in df_for_save.columns]
    df_for_save.columns = [f"Col_{i}" for i in range(len(df_for_save.columns))]
    df_for_save.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1)

    ws = writer.sheets[sheet_name]
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    last_row_idx = ws.max_row
    for col_idx in range(1, len(header_labels) + 1):
        cell = ws.cell(row=last_row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 5, 50)


def process_kq_tiep_thi_api_output(
    input_path=DEFAULT_KQ_TIEP_THI_INPUT,
    dsnv_file=DEFAULT_DSNV_FILE,
    overwrite_processed=False,
):
    """Xu ly bao cao ket qua tiep thi va ghi 2 sheet vao file processed."""
    raw_path = _resolve_path(input_path)
    df = pd.read_excel(raw_path, header=[0, 1])
    df.columns = _normalize_multiindex_columns(df.columns)

    cols_to_drop = [
        ("Dịch vụ BRCĐ", "Kết quả thực hiện tuần"),
        ("Dịch vụ BRCĐ", "Số liệu tiếp thị thuê bao lũy kế"),
        ("Dịch vụ MyTV", "Kết quả thực hiện tuần"),
        ("Dịch vụ MyTV", "Số liệu tiếp thị thuê bao lũy kế"),
    ]
    cols_present = [col for col in cols_to_drop if col in df.columns]
    df_processed = df.drop(columns=cols_present).copy()

    ten_nv_col = _get_required_column(df_processed, "Tên NV", "")
    brcd_month_col = _get_required_column(df_processed, "Dịch vụ BRCĐ", "Kết quả thực hiện trong tháng")
    mytv_month_col = _get_required_column(df_processed, "Dịch vụ MyTV", "Kết quả thực hiện trong tháng")

    lookup = _build_donvi_lookup(dsnv_file)
    units = df_processed[ten_nv_col].apply(_normalize_person_name).map(lookup).fillna("")
    df_processed.insert(1, ("Đơn vị", ""), units)

    df_processed[brcd_month_col] = pd.to_numeric(df_processed[brcd_month_col], errors="coerce").fillna(0)
    df_processed[mytv_month_col] = pd.to_numeric(df_processed[mytv_month_col], errors="coerce").fillna(0)
    df_processed[("Tổng", "")] = df_processed[brcd_month_col] + df_processed[mytv_month_col]

    if ("Đơn vị", "") in df_processed.columns:
        summary_cols = [brcd_month_col, mytv_month_col, ("Tổng", "")]
        df_summary = df_processed.groupby(("Đơn vị", ""))[summary_cols].sum().reset_index()
        df_summary.insert(0, ("STT", ""), range(1, len(df_summary) + 1))
        summary_total_row = {col: 0 for col in df_summary.columns}
        summary_total_row[("STT", "")] = ""
        summary_total_row[("Đơn vị", "")] = "TỔNG CỘNG"
        summary_total_row[brcd_month_col] = df_summary[brcd_month_col].sum()
        summary_total_row[mytv_month_col] = df_summary[mytv_month_col].sum()
        summary_total_row[("Tổng", "")] = df_summary[("Tổng", "")].sum()
        df_summary = pd.concat([df_summary, pd.DataFrame([summary_total_row])], ignore_index=True)
    else:
        df_summary = pd.DataFrame(
            columns=[("STT", ""), ("Đơn vị", ""), brcd_month_col, mytv_month_col, ("Tổng", "")]
        )

    total_row = {col: 0 for col in df_processed.columns}
    total_row[ten_nv_col] = "TỔNG CỘNG"
    total_row[("Đơn vị", "")] = ""
    total_row[brcd_month_col] = df_processed[brcd_month_col].sum()
    total_row[mytv_month_col] = df_processed[mytv_month_col].sum()
    total_row[("Tổng", "")] = df_processed[("Tổng", "")].sum()
    df_processed = pd.concat([df_processed, pd.DataFrame([total_row])], ignore_index=True)

    processed_path = ensure_processed_workbook(raw_path, overwrite=overwrite_processed)
    with pd.ExcelWriter(
        processed_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        _write_sheet_with_manual_header(writer, "kq_tiep_thi", df_processed)
        _write_sheet_with_manual_header(writer, "kq_th", df_summary)

    return processed_path
