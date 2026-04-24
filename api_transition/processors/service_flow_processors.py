# -*- coding: utf-8 -*-
"""Processors cho nhom luong dich vu trong api_transition."""

from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook, load_workbook

from api_transition.processors.common import (
    PROCESSED_DIR,
    append_or_replace_sheet,
    copy_raw_to_processed,
    ensure_processed_workbook,
    processed_group_dir,
)


DEFAULT_PHIEU_HOAN_CONG_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "phieu_hoan_cong_dich_vu"
    / "phieu_hoan_cong_dich_vu_chi_tiet.xlsx"
)
DEFAULT_TAM_DUNG_CHI_TIET_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "tam_dung_khoi_phuc_dich_vu_chi_tiet.xlsx"
)
DEFAULT_TAM_DUNG_CHI_TIET_KHOI_PHUC_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "tam_dung_khoi_phuc_dich_vu_chi_tiet_khoi_phuc.xlsx"
)
DEFAULT_TAM_DUNG_TONG_HOP_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "tam_dung_khoi_phuc_dich_vu_tong_hop.xlsx"
)
DEFAULT_TAM_DUNG_KHOI_PHUC_CHI_TIET_COMBINED_OUTPUT = (
    PROCESSED_DIR
    / "tam_dung_khoi_phuc_dich_vu"
    / "tam_dung_khoi_phuc_dich_vu_chi_tiet_combined_processed.xlsx"
)
DEFAULT_FIBER_T_MINUS_1_CAP_TO_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_fiber_thang_t-1_cap_to.xlsx"
)
DEFAULT_FIBER_T_MINUS_1_CAP_TTVT_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_fiber_thang_t-1_cap_ttvt.xlsx"
)
DEFAULT_MYTV_NGUNG_PSC_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_mytv_thang_t-1_cap_to.xlsx"
)
DEFAULT_MYTV_NGUNG_PSC_TTVT_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_mytv_thang_t-1_cap_ttvt.xlsx"
)
DEFAULT_MYTV_HOAN_CONG_INPUT = DEFAULT_PHIEU_HOAN_CONG_INPUT
DEFAULT_MYTV_NGUNG_PSC_OUTPUT = (
    PROCESSED_DIR
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_mytv_thang_t-1_cap_to_processed.xlsx"
)
DEFAULT_MYTV_HOAN_CONG_OUTPUT = (
    PROCESSED_DIR
    / "phieu_hoan_cong_dich_vu"
    / "phieu_hoan_cong_dich_vu_chi_tiet_processed.xlsx"
)
DEFAULT_MYTV_THUC_TANG_OUTPUT = (
    PROCESSED_DIR
    / "tam_dung_khoi_phuc_dich_vu"
    / "mytv_thuc_tang_processed.xlsx"
)
DEFAULT_MYTV_NGUNG_PSC_TTVT_OUTPUT = (
    PROCESSED_DIR
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_mytv_thang_t-1_cap_ttvt_processed.xlsx"
)
DEFAULT_FIBER_THUC_TANG_OUTPUT = (
    PROCESSED_DIR
    / "tam_dung_khoi_phuc_dich_vu"
    / "fiber_thuc_tang_processed.xlsx"
)
DEFAULT_SON_TAY_MYTV_NGUNG_T_MINUS_1_INPUT = (
    Path(__file__).resolve().parent.parent
    / "downloads"
    / "tam_dung_khoi_phuc_dich_vu"
    / "ngung_psc_mytv_thang_t-1_sontay.xlsx"
)


def _resolve_path(input_path):
    path = Path(input_path).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    else:
        path = path.resolve()

    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay file input: {path}")
    return path


def _normalize_person_name(value):
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    if "-" in text:
        text = text.split("-", 1)[1].strip()

    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.lower().title()


def _normalize_team_code(value):
    if pd.isna(value):
        return "(Chưa xác định)"

    text = str(value).strip().casefold()
    if not text:
        return "(Chưa xác định)"
    if "phúc thọ" in text or "phuc tho" in text or "tokt_phuctho" in text:
        return "ToKT_PhucTho"
    if "quảng oai" in text or "quang oai" in text or "tokt_quangoai" in text:
        return "ToKT_QuangOai"
    if "sơn tây" in text or "son tay" in text or "tokt_sontay" in text:
        return "ToKT_SonTay"
    if "suối hai" in text or "suoi hai" in text or "tokt_suoihai" in text:
        return "ToKT_SuoiHai"
    return str(value).strip()


def _default_mytv_ttvt(value=""):
    text = str(value).strip() if value is not None else ""
    return text or "TTVT Sơn Tây"


def _extract_nvkt_from_ten_kv(value):
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    parts = [part.strip() for part in text.split("-") if part.strip()]
    if len(parts) >= 2:
        return _normalize_person_name(parts[-1])
    return _normalize_person_name(text)


def _extract_nvkt_from_nhanvien_congviec(value):
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    first_line = text.splitlines()[0].strip()
    if ":" in first_line:
        first_line = first_line.split(":", 1)[0].strip()
    return _normalize_person_name(first_line)


def _coalesce(*values):
    for value in values:
        if value is not None and str(value).strip():
            return value
    return None


def _classify_service_type(value):
    if pd.isna(value):
        return "Khác"

    text = str(value).strip().casefold()
    if not text:
        return "Khác"
    if "mytv" in text:
        return "MyTV"
    if "cáp quang" in text or "cap quang" in text or "internet" in text or "fiber" in text:
        return "Fiber"
    return "Khác"


def _append_total_row(df, total_col, fixed_values):
    total_row = {col: "" for col in df.columns}
    total_row[total_col] = int(pd.to_numeric(df[total_col], errors="coerce").fillna(0).sum())
    for key, value in fixed_values.items():
        total_row[key] = value
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def _processed_target_for_group(raw_path, group_name):
    raw_path = _resolve_path(raw_path)
    return processed_group_dir(group_name) / f"{raw_path.stem}_processed{raw_path.suffix}"


def _ensure_processed_workbook_for_group(raw_path, group_name, overwrite=False):
    target_path = _processed_target_for_group(raw_path, group_name)
    return copy_raw_to_processed(raw_path, processed_path=target_path, overwrite=overwrite)


def _ensure_generated_workbook(output_path, overwrite=False):
    output_path = Path(output_path).expanduser()
    if not output_path.is_absolute():
        output_path = (Path.cwd() / output_path).resolve()
    else:
        output_path = output_path.resolve()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if overwrite or not output_path.exists():
        workbook = Workbook()
        workbook.save(output_path)
    return output_path


def _remove_empty_default_sheet(workbook_path):
    workbook = load_workbook(workbook_path)
    if len(workbook.sheetnames) > 1:
        for sheet_name in ["Sheet", "Sheet1"]:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
                    workbook.remove(sheet)
                    workbook.save(workbook_path)
                    break
    return workbook_path


def _prepare_phieu_hoan_cong_df(input_path):
    raw_path = _resolve_path(input_path)
    df = pd.read_excel(raw_path).copy()
    df.columns = [str(col).strip() for col in df.columns]

    required_cols = ["DOIVT", "TTVT"]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"Bao cao phieu hoan cong thieu cot bat buoc: {missing}")

    nvkt_from_db = df["TEN_NVKT_DB"].apply(_normalize_person_name) if "TEN_NVKT_DB" in df.columns else pd.Series([None] * len(df))
    nvkt_from_ten_kv = df["TEN_KV"].apply(_extract_nvkt_from_ten_kv) if "TEN_KV" in df.columns else pd.Series([None] * len(df))
    nvkt_from_congviec = (
        df["NHANVIEN_CONGVIEC"].apply(_extract_nvkt_from_nhanvien_congviec)
        if "NHANVIEN_CONGVIEC" in df.columns
        else pd.Series([None] * len(df))
    )

    df["NVKT"] = [
        _coalesce(a, b, c)
        for a, b, c in zip(nvkt_from_db, nvkt_from_ten_kv, nvkt_from_congviec)
    ]

    df["DOIVT"] = df["DOIVT"].fillna("(Chưa xác định)").astype(str).str.strip()
    df["TTVT"] = df["TTVT"].fillna("(Chưa xác định)").astype(str).str.strip()
    df["NVKT"] = df["NVKT"].fillna("(Chưa xác định)")

    return raw_path, df


def _build_nvkt_summary(df):
    grouped = (
        df.groupby(["TTVT", "DOIVT", "NVKT"], dropna=False)
        .size()
        .reset_index(name="SỐ LƯỢNG HOÀN CÔNG THÁNG")
        .sort_values(["TTVT", "DOIVT", "SỐ LƯỢNG HOÀN CÔNG THÁNG", "NVKT"], ascending=[True, True, False, True])
        .reset_index(drop=True)
    )
    return _append_total_row(
        grouped,
        "SỐ LƯỢNG HOÀN CÔNG THÁNG",
        {"TTVT": "TỔNG CỘNG", "DOIVT": "", "NVKT": ""},
    )


def _build_team_summary(df):
    grouped = (
        df.groupby(["TTVT", "DOIVT"], dropna=False)
        .size()
        .reset_index(name="SỐ LƯỢNG HOÀN CÔNG THÁNG")
        .sort_values(["TTVT", "SỐ LƯỢNG HOÀN CÔNG THÁNG", "DOIVT"], ascending=[True, False, True])
        .reset_index(drop=True)
    )
    return _append_total_row(
        grouped,
        "SỐ LƯỢNG HOÀN CÔNG THÁNG",
        {"TTVT": "TỔNG CỘNG", "DOIVT": ""},
    )


def _build_ttvt_summary(df):
    grouped = (
        df.groupby(["TTVT"], dropna=False)
        .size()
        .reset_index(name="SỐ LƯỢNG HOÀN CÔNG THÁNG")
        .sort_values(["SỐ LƯỢNG HOÀN CÔNG THÁNG", "TTVT"], ascending=[False, True])
        .reset_index(drop=True)
    )
    return _append_total_row(
        grouped,
        "SỐ LƯỢNG HOÀN CÔNG THÁNG",
        {"TTVT": "TỔNG CỘNG"},
    )


def _build_reason_summary(df, reason_col):
    if reason_col not in df.columns:
        return pd.DataFrame(columns=[reason_col, "Số lượng"])

    reason_df = df.copy()
    reason_df[reason_col] = reason_df[reason_col].fillna("(Không có lý do)").astype(str).str.strip()
    reason_df.loc[reason_df[reason_col] == "", reason_col] = "(Không có lý do)"

    return (
        reason_df.groupby(reason_col)
        .size()
        .reset_index(name="Số lượng")
        .sort_values(["Số lượng", reason_col], ascending=[False, True])
        .reset_index(drop=True)
    )


def _build_group_count_summary(df, group_cols, count_col):
    grouped = (
        df.groupby(group_cols, dropna=False)
        .size()
        .reset_index(name=count_col)
        .sort_values(group_cols[:-1] + [count_col, group_cols[-1]], ascending=[True] * (len(group_cols) - 1) + [False, True])
        .reset_index(drop=True)
    )
    return grouped


def _build_multirow_header_map(df_raw, header_row_count=3):
    header_map = {}
    for col_idx in range(df_raw.shape[1]):
        header_values = [
            str(df_raw.iat[row_idx, col_idx]).strip()
            for row_idx in range(min(header_row_count, len(df_raw)))
            if pd.notna(df_raw.iat[row_idx, col_idx]) and str(df_raw.iat[row_idx, col_idx]).strip()
        ]
        for value in header_values:
            header_map.setdefault(value, col_idx)
    return header_map


def _extract_mytv_t_minus_1_summary_df(input_path, unit_level):
    raw_path = _resolve_path(input_path)
    df_raw = pd.read_excel(raw_path, header=None)
    header_map = _build_multirow_header_map(df_raw)
    required_columns = {
        "Đơn vị/Nhân viên KT": "DON_VI",
        "Hoàn công(*) (1.5)": "HOAN_CONG",
        "Lũy kế tháng(1.6)": "LUY_KE_THANG",
        "Lũy kế năm(1.7)": "LUY_KE_NAM",
        "TB ngưng PSC tháng(4.2)(cột 1.6 - 4.1)": "NGUNG_PSC_THANG",
        "Ngưng PSC tạm tính tháng T(5.1)": "NGUNG_PSC_TAM_TINH",
    }
    missing_headers = [header for header in required_columns if header not in header_map]
    if missing_headers:
        raise ValueError(
            "Bao cao MyTV T-1 thieu cot bat buoc: " + ", ".join(missing_headers)
        )

    selected_cols = [header_map[header] for header in required_columns]
    df = df_raw.iloc[3:, selected_cols].copy()
    df.columns = list(required_columns.values())
    df = df[df["DON_VI"].notna()].copy()
    df["DON_VI"] = df["DON_VI"].astype(str).str.strip()
    df = df[df["DON_VI"] != ""].reset_index(drop=True)

    numeric_cols = ["HOAN_CONG", "LUY_KE_THANG", "LUY_KE_NAM", "NGUNG_PSC_THANG", "NGUNG_PSC_TAM_TINH"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    if unit_level == "team":
        df["TEN_DOI"] = df["DON_VI"].apply(_normalize_team_code)
        df["TEN_TTVT"] = "TTVT Sơn Tây"
    elif unit_level == "ttvt":
        df["TEN_TTVT"] = df["DON_VI"].astype(str).str.strip()
        df["TEN_DOI"] = ""
    else:
        raise ValueError(f"Unsupported unit_level: {unit_level}")

    return raw_path, df


def _build_mytv_ngung_psc_detail_rows(summary_df):
    records = []
    for row in summary_df.itertuples(index=False):
        count = int(getattr(row, "NGUNG_PSC_TAM_TINH", 0))
        for index in range(count):
            records.append(
                {
                    "STT": len(records) + 1,
                    "MA_TB": f"SUMMARY-{row.TEN_DOI or row.TEN_TTVT}-{index + 1:04d}",
                    "NGAY_TD": None,
                    "NGAY_HUY": None,
                    "MA_KV": None,
                    "TEN_KV": "(Tổng hợp cấp tổ)",
                    "TEN_DOI": row.TEN_DOI,
                    "TEN_TTVT": row.TEN_TTVT,
                    "TRANGTHAI_TB": "Tạm dừng",
                    "LOAIHINH_TB": "MyTV",
                }
            )
    return pd.DataFrame(
        records,
        columns=[
            "STT",
            "MA_TB",
            "NGAY_TD",
            "NGAY_HUY",
            "MA_KV",
            "TEN_KV",
            "TEN_DOI",
            "TEN_TTVT",
            "TRANGTHAI_TB",
            "LOAIHINH_TB",
        ],
    )


def _prepare_mytv_hoan_cong_df(input_path):
    raw_path, df = _prepare_phieu_hoan_cong_df(input_path)
    if "TEN_DVVT_HNI" in df.columns:
        service_series = df["TEN_DVVT_HNI"].apply(_classify_service_type)
    elif "LOAIHINH_TB" in df.columns:
        service_series = df["LOAIHINH_TB"].apply(_classify_service_type)
    else:
        service_series = pd.Series(["Khác"] * len(df), index=df.index)

    df = df[service_series == "MyTV"].copy()
    if df.empty:
        raise ValueError("Khong tim thay dong MyTV nao trong bao cao phieu hoan cong dich vu chi tiet.")

    df["DOIVT"] = df["DOIVT"].apply(_normalize_team_code)
    df["TEN_TTVT"] = df["TTVT"].apply(_default_mytv_ttvt)
    df["NHANVIEN_KT"] = df["NVKT"].fillna("(Chưa xác định)")
    data_df = pd.DataFrame(
        {
            "STT": range(1, len(df) + 1),
            "MA_TB": df.get("MA_TB"),
            "HDTB_ID": df.get("HDTB_ID"),
            "NGAY_INS": df.get("NGAY_HC"),
            "NGAY_YC": df.get("NGAYLAP_HD"),
            "NHOM_DIABAN": df.get("TEN_KV"),
            "DOIVT": df.get("DOIVT"),
            "TEN_TTVT": df.get("TEN_TTVT"),
            "TRANGTHAI_HD": df.get("TRANGTHAI_HD"),
            "NHANVIEN_KT": df.get("NHANVIEN_KT"),
            "MA_GD": df.get("MA_GD"),
        }
    )
    return raw_path, data_df


def _append_mytv_total_row(df, team_col, person_col, value_col):
    total_row = pd.DataFrame(
        [{team_col: "TỔNG CỘNG", person_col: "", value_col: int(pd.to_numeric(df[value_col], errors="coerce").fillna(0).sum())}]
    )
    return pd.concat([df, total_row], ignore_index=True)


def _append_mytv_team_total_row(df, team_col, value_col):
    total_row = pd.DataFrame(
        [{team_col: "TỔNG CỘNG", value_col: int(pd.to_numeric(df[value_col], errors="coerce").fillna(0).sum())}]
    )
    return pd.concat([df, total_row], ignore_index=True)


def _finalize_thuc_tang_df(df, sort_cols, total_values):
    df["Hoàn công"] = pd.to_numeric(df["Hoàn công"], errors="coerce").fillna(0).astype(int)
    df["Ngưng phát sinh cước"] = pd.to_numeric(df["Ngưng phát sinh cước"], errors="coerce").fillna(0).astype(int)
    df["Thực tăng"] = df["Hoàn công"] - df["Ngưng phát sinh cước"]
    df["Tỷ lệ ngưng/psc"] = df.apply(
        lambda row: round((row["Ngưng phát sinh cước"] / row["Hoàn công"]) * 100, 2) if row["Hoàn công"] else 0.0,
        axis=1,
    )
    df = df.sort_values(sort_cols, ascending=[False] + [True] * (len(sort_cols) - 1)).reset_index(drop=True)

    total_hoan_cong = int(df["Hoàn công"].sum())
    total_ngung = int(df["Ngưng phát sinh cước"].sum())
    total_row = {
        "Hoàn công": total_hoan_cong,
        "Ngưng phát sinh cước": total_ngung,
        "Thực tăng": int(df["Thực tăng"].sum()),
        "Tỷ lệ ngưng/psc": round((total_ngung / total_hoan_cong) * 100, 2) if total_hoan_cong else 0.0,
    }
    total_row.update(total_values)
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def _extract_son_tay_ngung_psc_subset(df_raw):
    data_row_indices = [3, 4, 5, 6]
    if len(df_raw) <= max(data_row_indices):
        raise ValueError(
            f"File Son Tay khong du so dong du lieu. Can it nhat {max(data_row_indices) + 1} dong, nhung chi co {len(df_raw)} dong."
        )

    header_row_count = min(3, len(df_raw))
    header_map = {}
    for col_idx in range(df_raw.shape[1]):
        header_values = [
            str(df_raw.iat[row_idx, col_idx]).strip()
            for row_idx in range(header_row_count)
            if pd.notna(df_raw.iat[row_idx, col_idx]) and str(df_raw.iat[row_idx, col_idx]).strip()
        ]
        for value in header_values:
            header_map.setdefault(value, col_idx)

    required_columns = {
        "Đơn vị/Nhân viên KT": "Đơn vị/Nhân viên KT",
        "Hoàn công(*) (1.5)": "Hoàn công(*) (1.5)",
        "Lũy kế tháng(1.6)": "Lũy kế tháng(1.6)",
        "Lũy kế năm(1.7)": "Lũy kế năm(1.7)",
        "Ngưng PSC tạm tính tháng T(5.1)": "Ngưng PSC tạm tính tháng T(5.1)",
        "TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)": "TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)",
    }
    missing_headers = [header for header in required_columns if header not in header_map]
    if missing_headers:
        raise ValueError(
            "Khong tim thay cac cot bat buoc trong bao cao Son Tay: " + ", ".join(missing_headers)
        )

    selected_col_indices = [header_map[header] for header in required_columns]
    df_subset = df_raw.iloc[data_row_indices, selected_col_indices].copy()
    df_subset.columns = list(required_columns.values())
    return df_subset


def _flatten_header_rows(df_raw, header_rows=3):
    flat_headers = []
    for col_idx in range(df_raw.shape[1]):
        parts = []
        for row_idx in range(min(header_rows, len(df_raw))):
            value = df_raw.iat[row_idx, col_idx]
            if pd.notna(value):
                text = str(value).strip()
                if text and text not in parts:
                    parts.append(text)
        if parts:
            flat_headers.append(" | ".join(parts))
        else:
            flat_headers.append(f"COL_{col_idx + 1}")
    return flat_headers


def process_phieu_hoan_cong_dich_vu_chi_tiet_api_output(
    input_path=DEFAULT_PHIEU_HOAN_CONG_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao phieu hoan cong dich vu chi tiet."""
    raw_path, df = _prepare_phieu_hoan_cong_df(input_path)
    df_nvkt = _build_nvkt_summary(df)
    df_team = _build_team_summary(df)
    df_ttvt = _build_ttvt_summary(df)

    processed_path = ensure_processed_workbook(raw_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "Data", df)
    append_or_replace_sheet(processed_path, "fiber_hoan_cong_thang", df_nvkt)
    append_or_replace_sheet(processed_path, "fiber_hoan_cong_thang_theo_to", df_team)
    append_or_replace_sheet(processed_path, "fiber_hoan_cong_thang_theo_ttvt", df_ttvt)
    return processed_path


def _prepare_tam_dung_chi_tiet_df(input_path):
    raw_path = _resolve_path(input_path)
    df = pd.read_excel(raw_path).copy()
    df.columns = [str(col).strip() for col in df.columns]

    required_cols = ["DOIVT", "TTVT"]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"Bao cao tam dung/khoi phuc thieu cot bat buoc: {missing}")

    nvkt_from_db = df["TEN_NVKT_DB"].apply(_normalize_person_name) if "TEN_NVKT_DB" in df.columns else pd.Series([None] * len(df))
    nvkt_from_ten_kv = df["TEN_KV"].apply(_extract_nvkt_from_ten_kv) if "TEN_KV" in df.columns else pd.Series([None] * len(df))
    nvkt_from_xm = df["NVKT_XM"].apply(_normalize_person_name) if "NVKT_XM" in df.columns else pd.Series([None] * len(df))

    df["NVKT"] = [
        _coalesce(a, b, c)
        for a, b, c in zip(nvkt_from_db, nvkt_from_ten_kv, nvkt_from_xm)
    ]

    df["DOIVT"] = df["DOIVT"].fillna("(Chưa xác định)").astype(str).str.strip()
    df["TTVT"] = df["TTVT"].fillna("(Chưa xác định)").astype(str).str.strip()
    df["NVKT"] = df["NVKT"].fillna("(Chưa xác định)")
    df["LOAI_DICH_VU"] = (
        df["TEN_DVVT_HNI"].apply(_classify_service_type)
        if "TEN_DVVT_HNI" in df.columns
        else "Khác"
    )
    return raw_path, df


def _build_pause_restore_summary(df_pause, df_restore, group_cols):
    pause_df = df_pause[df_pause["LOAI_DICH_VU"].isin(["Fiber", "MyTV"])].copy()
    if pause_df.empty:
        pause_pivot = pd.DataFrame(columns=group_cols + ["Tạm dừng Fiber", "Tạm dừng MyTV"])
    else:
        pause_counts = (
            pause_df.groupby(group_cols + ["LOAI_DICH_VU"], dropna=False)
            .size()
            .reset_index(name="Số lượng")
        )
        pause_pivot = (
            pause_counts.pivot_table(
                index=group_cols,
                columns="LOAI_DICH_VU",
                values="Số lượng",
                fill_value=0,
                aggfunc="sum",
            )
            .reset_index()
        )
        pause_pivot.columns.name = None
        pause_pivot = pause_pivot.rename(
            columns={
                "Fiber": "Tạm dừng Fiber",
                "MyTV": "Tạm dừng MyTV",
            }
        )

    restore_df = df_restore[df_restore["LOAI_DICH_VU"].isin(["Fiber", "MyTV"])].copy()
    if restore_df.empty:
        restore_pivot = pd.DataFrame(columns=group_cols + ["Khôi phục Fiber", "Khôi phục MyTV"])
    else:
        restore_counts = (
            restore_df.groupby(group_cols + ["LOAI_DICH_VU"], dropna=False)
            .size()
            .reset_index(name="Số lượng")
        )
        restore_pivot = (
            restore_counts.pivot_table(
                index=group_cols,
                columns="LOAI_DICH_VU",
                values="Số lượng",
                fill_value=0,
                aggfunc="sum",
            )
            .reset_index()
        )
        restore_pivot.columns.name = None
        restore_pivot = restore_pivot.rename(
            columns={
                "Fiber": "Khôi phục Fiber",
                "MyTV": "Khôi phục MyTV",
            }
        )

    summary_df = pd.merge(pause_pivot, restore_pivot, on=group_cols, how="outer").fillna(0)
    for col in ["Tạm dừng Fiber", "Tạm dừng MyTV", "Khôi phục Fiber", "Khôi phục MyTV"]:
        if col not in summary_df.columns:
            summary_df[col] = 0
        summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce").fillna(0).astype(int)

    summary_df["Chưa khôi phục Fiber"] = summary_df["Tạm dừng Fiber"] - summary_df["Khôi phục Fiber"]
    summary_df["Chưa khôi phục MyTV"] = summary_df["Tạm dừng MyTV"] - summary_df["Khôi phục MyTV"]

    sort_cols = group_cols[:-1] + [
        "Chưa khôi phục Fiber",
        "Chưa khôi phục MyTV",
        "Tạm dừng Fiber",
        "Tạm dừng MyTV",
        "Khôi phục Fiber",
        "Khôi phục MyTV",
        group_cols[-1],
    ]
    ascending = [True] * (len(group_cols) - 1) + [False, False, False, False, False, False, True]
    summary_df = summary_df.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    total_row = {col: "" for col in summary_df.columns}
    for col in [
        "Tạm dừng Fiber",
        "Tạm dừng MyTV",
        "Khôi phục Fiber",
        "Khôi phục MyTV",
        "Chưa khôi phục Fiber",
        "Chưa khôi phục MyTV",
    ]:
        total_row[col] = int(summary_df[col].sum())
    if group_cols:
        total_row[group_cols[0]] = "TỔNG CỘNG"
        for col in group_cols[1:]:
            total_row[col] = ""

    return pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)


def _build_unrestored_detail(df_pause, df_restore):
    pause_detail = df_pause.copy()
    restore_detail = df_restore.copy()

    if "MA_TB" not in pause_detail.columns:
        raise ValueError("Bao cao tam dung thieu cot MA_TB de doi chieu voi du lieu khoi phuc.")

    pause_detail["MA_TB"] = pause_detail["MA_TB"].fillna("").astype(str).str.strip()
    if "MA_TB" in restore_detail.columns:
        restore_ma_tb = set(
            restore_detail["MA_TB"].fillna("").astype(str).str.strip().loc[lambda s: s != ""]
        )
    else:
        restore_ma_tb = set()

    unrestored_df = pause_detail.loc[
        (pause_detail["MA_TB"] != "") & (~pause_detail["MA_TB"].isin(restore_ma_tb))
    ].copy()

    sort_cols = [col for col in ["LOAI_DICH_VU", "TTVT", "DOIVT", "NVKT", "MA_TB"] if col in unrestored_df.columns]
    if sort_cols:
        unrestored_df = unrestored_df.sort_values(sort_cols, ascending=[True] * len(sort_cols)).reset_index(drop=True)
    return unrestored_df


def process_tam_dung_khoi_phuc_dich_vu_chi_tiet_api_output(
    input_path=DEFAULT_TAM_DUNG_CHI_TIET_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao tam dung/khoi phuc dich vu chi tiet."""
    return _process_tam_dung_khoi_phuc_dich_vu_chi_tiet_output(
        input_path=input_path,
        overwrite_processed=overwrite_processed,
        summary_label="SỐ LƯỢNG NGƯNG PSC THÁNG",
        nvkt_sheet_name="fiber_ngung_psc_thang",
        team_sheet_name="fiber_ngung_psc_thang_theo_to",
        ttvt_sheet_name="fiber_ngung_psc_thang_theo_ttvt",
        reason_sheet_name="tong_hop_ly_do_huy",
    )


def _process_tam_dung_khoi_phuc_dich_vu_chi_tiet_output(
    input_path,
    overwrite_processed,
    summary_label,
    nvkt_sheet_name,
    team_sheet_name,
    ttvt_sheet_name,
    reason_sheet_name,
):
    raw_path, df = _prepare_tam_dung_chi_tiet_df(input_path)
    df_nvkt = _build_nvkt_summary(df).rename(
        columns={"SỐ LƯỢNG HOÀN CÔNG THÁNG": summary_label}
    )
    df_team = _build_team_summary(df).rename(
        columns={"SỐ LƯỢNG HOÀN CÔNG THÁNG": summary_label}
    )
    df_ttvt = _build_ttvt_summary(df).rename(
        columns={"SỐ LƯỢNG HOÀN CÔNG THÁNG": summary_label}
    )
    df_reasons = _build_reason_summary(df, "LYDOHUY")

    processed_path = ensure_processed_workbook(raw_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "Data", df)
    append_or_replace_sheet(processed_path, nvkt_sheet_name, df_nvkt)
    append_or_replace_sheet(processed_path, team_sheet_name, df_team)
    append_or_replace_sheet(processed_path, ttvt_sheet_name, df_ttvt)
    append_or_replace_sheet(processed_path, reason_sheet_name, df_reasons)
    return processed_path


def process_tam_dung_khoi_phuc_dich_vu_chi_tiet_khoi_phuc_api_output(
    input_path=DEFAULT_TAM_DUNG_CHI_TIET_KHOI_PHUC_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao tam dung/khoi phuc dich vu chi tiet - khoi phuc."""
    return _process_tam_dung_khoi_phuc_dich_vu_chi_tiet_output(
        input_path=input_path,
        overwrite_processed=overwrite_processed,
        summary_label="SỐ LƯỢNG KHÔI PHỤC THÁNG",
        nvkt_sheet_name="fiber_khoi_phuc_thang",
        team_sheet_name="fiber_khoi_phuc_thang_theo_to",
        ttvt_sheet_name="fiber_khoi_phuc_thang_theo_ttvt",
        reason_sheet_name="tong_hop_ly_do_huy_khoi_phuc",
    )


def process_tam_dung_khoi_phuc_dich_vu_chi_tiet_combined_api_output(
    tam_dung_input_path=DEFAULT_TAM_DUNG_CHI_TIET_INPUT,
    khoi_phuc_input_path=DEFAULT_TAM_DUNG_CHI_TIET_KHOI_PHUC_INPUT,
    combined_output_path=DEFAULT_TAM_DUNG_KHOI_PHUC_CHI_TIET_COMBINED_OUTPUT,
    overwrite_processed=False,
):
    """Xu ly dong thoi 2 file chi tiet tam dung va khoi phuc, kem workbook tong hop."""
    tam_dung_processed_path = process_tam_dung_khoi_phuc_dich_vu_chi_tiet_api_output(
        input_path=tam_dung_input_path,
        overwrite_processed=overwrite_processed,
    )
    khoi_phuc_processed_path = process_tam_dung_khoi_phuc_dich_vu_chi_tiet_khoi_phuc_api_output(
        input_path=khoi_phuc_input_path,
        overwrite_processed=overwrite_processed,
    )

    _, df_tam_dung = _prepare_tam_dung_chi_tiet_df(tam_dung_input_path)
    _, df_khoi_phuc = _prepare_tam_dung_chi_tiet_df(khoi_phuc_input_path)

    df_tam_dung_combined = df_tam_dung.copy()
    df_tam_dung_combined["LUONG_BAO_CAO"] = "Tạm dừng"
    df_khoi_phuc_combined = df_khoi_phuc.copy()
    df_khoi_phuc_combined["LUONG_BAO_CAO"] = "Khôi phục"
    df_combined = pd.concat([df_tam_dung_combined, df_khoi_phuc_combined], ignore_index=True, sort=False)

    df_nvkt = _build_pause_restore_summary(df_tam_dung, df_khoi_phuc, ["TTVT", "DOIVT", "NVKT"])
    df_team = _build_pause_restore_summary(df_tam_dung, df_khoi_phuc, ["TTVT", "DOIVT"])
    df_ttvt = _build_pause_restore_summary(df_tam_dung, df_khoi_phuc, ["TTVT"])
    df_unrestored = _build_unrestored_detail(df_tam_dung, df_khoi_phuc)

    combined_path = _ensure_generated_workbook(combined_output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(combined_path, "Data_tam_dung", df_tam_dung)
    append_or_replace_sheet(combined_path, "Data_khoi_phuc", df_khoi_phuc)
    append_or_replace_sheet(combined_path, "Data_combined", df_combined)
    append_or_replace_sheet(combined_path, "tong_hop_theo_NVKT", df_nvkt)
    append_or_replace_sheet(combined_path, "tong_hop_theo_to", df_team)
    append_or_replace_sheet(combined_path, "tong_hop_theo_TTVT", df_ttvt)
    append_or_replace_sheet(combined_path, "chi_tiet_chua_khoi_phuc", df_unrestored)
    _remove_empty_default_sheet(combined_path)

    return {
        "tam_dung_processed_path": tam_dung_processed_path,
        "khoi_phuc_processed_path": khoi_phuc_processed_path,
        "combined_processed_path": combined_path,
    }


def process_tam_dung_khoi_phuc_dich_vu_tong_hop_api_output(
    input_path=DEFAULT_TAM_DUNG_TONG_HOP_INPUT,
    overwrite_processed=False,
):
    """Xu ly workbook tong hop tam dung/khoi phuc dich vu.

    Hien tai file raw co the chua du lieu. Truong hop do, van tao file processed
    va ghi sheet thong bao de luong sau khong bi gay.
    """
    raw_path = _resolve_path(input_path)
    try:
        df = pd.read_excel(raw_path)
    except ValueError:
        df = pd.DataFrame()

    processed_path = ensure_processed_workbook(raw_path, overwrite=overwrite_processed)
    if df.empty and len(df.columns) == 0:
        note_df = pd.DataFrame(
            [{"Trạng thái": "Không có dữ liệu", "Ghi chú": "Workbook nguồn đang rỗng, chưa có dữ liệu để chuẩn hóa."}]
        )
        append_or_replace_sheet(processed_path, "thong_bao", note_df)
        return processed_path

    df.columns = [str(col).strip() for col in df.columns]
    append_or_replace_sheet(processed_path, "du_lieu_sach", df)
    return processed_path


def process_fiber_thuc_tang_api_output(
    ngung_psc_input=DEFAULT_TAM_DUNG_CHI_TIET_INPUT,
    hoan_cong_input=DEFAULT_PHIEU_HOAN_CONG_INPUT,
    output_path=DEFAULT_FIBER_THUC_TANG_OUTPUT,
    overwrite_processed=False,
):
    """Tao bao cao Fiber thuc tang tu 2 nguon raw: hoan cong va tam dung/khoi phuc."""
    _, df_ngung = _prepare_tam_dung_chi_tiet_df(ngung_psc_input)
    _, df_hc = _prepare_phieu_hoan_cong_df(hoan_cong_input)

    df_ngung_team = (
        df_ngung.groupby(["TTVT", "DOIVT"], dropna=False)
        .size()
        .reset_index(name="Ngưng phát sinh cước")
    )
    df_hc_team = (
        df_hc.groupby(["TTVT", "DOIVT"], dropna=False)
        .size()
        .reset_index(name="Hoàn công")
    )
    df_thuc_tang_to = pd.merge(df_hc_team, df_ngung_team, on=["TTVT", "DOIVT"], how="outer").fillna(0)
    df_thuc_tang_to = df_thuc_tang_to.rename(columns={"DOIVT": "Đội VT"})
    df_thuc_tang_to = _finalize_thuc_tang_df(
        df_thuc_tang_to,
        ["Thực tăng", "TTVT", "Đội VT"],
        {"TTVT": "TỔNG CỘNG", "Đội VT": ""},
    )

    df_ngung_nvkt = (
        df_ngung.groupby(["TTVT", "DOIVT", "NVKT"], dropna=False)
        .size()
        .reset_index(name="Ngưng phát sinh cước")
    )
    df_hc_nvkt = (
        df_hc.groupby(["TTVT", "DOIVT", "NVKT"], dropna=False)
        .size()
        .reset_index(name="Hoàn công")
    )
    df_thuc_tang_nvkt = pd.merge(
        df_hc_nvkt,
        df_ngung_nvkt,
        on=["TTVT", "DOIVT", "NVKT"],
        how="outer",
    ).fillna(0)
    df_thuc_tang_nvkt = df_thuc_tang_nvkt.rename(columns={"DOIVT": "Đội VT"})
    df_thuc_tang_nvkt = _finalize_thuc_tang_df(
        df_thuc_tang_nvkt,
        ["Thực tăng", "TTVT", "Đội VT", "NVKT"],
        {"TTVT": "TỔNG CỘNG", "Đội VT": "", "NVKT": ""},
    )

    processed_path = _ensure_generated_workbook(output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "thuc_tang_theo_to", df_thuc_tang_to)
    append_or_replace_sheet(processed_path, "thuc_tang_theo_NVKT", df_thuc_tang_nvkt)
    _remove_empty_default_sheet(processed_path)
    return processed_path


def process_mytv_ngung_psc_api_output(
    input_path=DEFAULT_MYTV_NGUNG_PSC_INPUT,
    ttvt_input_path=DEFAULT_MYTV_NGUNG_PSC_TTVT_INPUT,
    output_path=DEFAULT_MYTV_NGUNG_PSC_OUTPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao MyTV ngung PSC."""
    _, df_team_source = _extract_mytv_t_minus_1_summary_df(input_path, unit_level="team")
    _, df_ttvt_source = _extract_mytv_t_minus_1_summary_df(ttvt_input_path, unit_level="ttvt")

    data_df = _build_mytv_ngung_psc_detail_rows(df_team_source)
    df_nvkt = (
        df_team_source[["TEN_DOI", "NGUNG_PSC_TAM_TINH"]]
        .rename(columns={"NGUNG_PSC_TAM_TINH": "SỐ LƯỢNG NGƯNG PSC THÁNG"})
        .copy()
    )
    df_nvkt["TEN_KV"] = "(Tổng hợp cấp tổ)"
    df_nvkt = df_nvkt[["TEN_DOI", "TEN_KV", "SỐ LƯỢNG NGƯNG PSC THÁNG"]]
    df_nvkt = df_nvkt.sort_values(["SỐ LƯỢNG NGƯNG PSC THÁNG", "TEN_DOI"], ascending=[False, True]).reset_index(drop=True)
    df_nvkt = _append_mytv_total_row(df_nvkt, "TEN_DOI", "TEN_KV", "SỐ LƯỢNG NGƯNG PSC THÁNG")

    df_team = (
        df_team_source[["TEN_DOI", "NGUNG_PSC_TAM_TINH"]]
        .rename(columns={"NGUNG_PSC_TAM_TINH": "SỐ LƯỢNG NGƯNG PSC THÁNG"})
        .sort_values(["SỐ LƯỢNG NGƯNG PSC THÁNG", "TEN_DOI"], ascending=[False, True])
        .reset_index(drop=True)
    )
    df_team = _append_mytv_team_total_row(df_team, "TEN_DOI", "SỐ LƯỢNG NGƯNG PSC THÁNG")

    df_ttvt = (
        df_ttvt_source[["TEN_TTVT", "NGUNG_PSC_TAM_TINH"]]
        .rename(columns={"TEN_TTVT": "TTVT", "NGUNG_PSC_TAM_TINH": "SỐ LƯỢNG NGƯNG PSC THÁNG"})
        .sort_values(["SỐ LƯỢNG NGƯNG PSC THÁNG", "TTVT"], ascending=[False, True])
        .reset_index(drop=True)
    )
    df_ttvt = _append_mytv_team_total_row(df_ttvt, "TTVT", "SỐ LƯỢNG NGƯNG PSC THÁNG")

    processed_path = _ensure_generated_workbook(output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "Data", data_df)
    append_or_replace_sheet(processed_path, "mytv_ngung_psc_thang", df_nvkt)
    append_or_replace_sheet(processed_path, "mytv_ngung_psc_thang_theo_to", df_team)
    append_or_replace_sheet(processed_path, "mytv_ngung_psc_thang_theo_ttvt", df_ttvt)
    append_or_replace_sheet(processed_path, "source_cap_to", df_team_source)
    append_or_replace_sheet(processed_path, "source_cap_ttvt", df_ttvt_source)
    _remove_empty_default_sheet(processed_path)
    return processed_path


def process_mytv_ngung_psc_ttvt_api_output(
    input_path=DEFAULT_MYTV_NGUNG_PSC_TTVT_INPUT,
    output_path=DEFAULT_MYTV_NGUNG_PSC_TTVT_OUTPUT,
    overwrite_processed=False,
):
    """Xu ly rieng bao cao MyTV ngung PSC cap TTVT."""
    raw_path = _resolve_path(input_path)
    df_raw = pd.read_excel(raw_path, header=None).copy()
    if len(df_raw) <= 3:
        raise ValueError(
            f"File MyTV TTVT khong du dong du lieu. Can it nhat 4 dong, nhung chi co {len(df_raw)} dong."
        )

    df_data = df_raw.iloc[3:].reset_index(drop=True).copy()
    df_data.columns = _flatten_header_rows(df_raw, header_rows=3)
    first_col = df_data.columns[0]
    df_data = df_data[df_data[first_col].notna() & (df_data[first_col].astype(str).str.strip() != "")].reset_index(drop=True)

    processed_path = _ensure_generated_workbook(output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "TH_ngung_PSC-Thang T-1", df_data)
    _remove_empty_default_sheet(processed_path)
    return processed_path


def process_mytv_hoan_cong_api_output(
    input_path=DEFAULT_MYTV_HOAN_CONG_INPUT,
    output_path=DEFAULT_MYTV_HOAN_CONG_OUTPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao MyTV hoan cong."""
    _, df = _prepare_mytv_hoan_cong_df(input_path)
    df_nvkt = _build_group_count_summary(df, ["DOIVT", "NHANVIEN_KT"], "SỐ LƯỢNG HOÀN CÔNG THÁNG")
    df_nvkt = _append_mytv_total_row(df_nvkt, "DOIVT", "NHANVIEN_KT", "SỐ LƯỢNG HOÀN CÔNG THÁNG")
    df_team = (
        df.groupby("DOIVT", dropna=False)
        .size()
        .reset_index(name="SỐ LƯỢNG HOÀN CÔNG THÁNG")
        .sort_values(["SỐ LƯỢNG HOÀN CÔNG THÁNG", "DOIVT"], ascending=[False, True])
        .reset_index(drop=True)
    )
    df_team = _append_mytv_team_total_row(df_team, "DOIVT", "SỐ LƯỢNG HOÀN CÔNG THÁNG")

    df_ttvt = (
        df.groupby("TEN_TTVT", dropna=False)
        .size()
        .reset_index(name="SỐ LƯỢNG HOÀN CÔNG THÁNG")
        .sort_values(["SỐ LƯỢNG HOÀN CÔNG THÁNG", "TEN_TTVT"], ascending=[False, True])
        .reset_index(drop=True)
    )
    df_ttvt = _append_mytv_team_total_row(df_ttvt, "TEN_TTVT", "SỐ LƯỢNG HOÀN CÔNG THÁNG")

    processed_path = _ensure_generated_workbook(output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "Data", df)
    append_or_replace_sheet(processed_path, "mytv_hoan_cong_thang", df_nvkt)
    append_or_replace_sheet(processed_path, "mytv_hoan_cong_thang_theo_to", df_team)
    append_or_replace_sheet(processed_path, "mytv_hoan_cong_thang_theo_ttvt", df_ttvt)
    _remove_empty_default_sheet(processed_path)
    return processed_path


def process_mytv_thuc_tang_api_output(
    ngung_psc_input=DEFAULT_MYTV_NGUNG_PSC_INPUT,
    hoan_cong_input=DEFAULT_MYTV_HOAN_CONG_INPUT,
    output_path=DEFAULT_MYTV_THUC_TANG_OUTPUT,
    overwrite_processed=False,
):
    """Tao bao cao MyTV thuc tang tu 2 nguon raw: hoan cong va ngung PSC."""
    _, df_ngung_source = _extract_mytv_t_minus_1_summary_df(ngung_psc_input, unit_level="team")
    _, df_hc = _prepare_mytv_hoan_cong_df(hoan_cong_input)

    df_ngung_team = (
        df_ngung_source[["TEN_TTVT", "TEN_DOI", "NGUNG_PSC_TAM_TINH"]]
        .rename(
            columns={
                "TEN_TTVT": "TTVT",
                "TEN_DOI": "Đội VT",
                "NGUNG_PSC_TAM_TINH": "Ngưng phát sinh cước",
            }
        )
        .copy()
    )
    df_hc_team = (
        df_hc.groupby(["TEN_TTVT", "DOIVT"], dropna=False)
        .size()
        .reset_index(name="Hoàn công")
        .rename(columns={"TEN_TTVT": "TTVT", "DOIVT": "Đội VT"})
    )
    df_thuc_tang_to = pd.merge(df_hc_team, df_ngung_team, on=["TTVT", "Đội VT"], how="outer").fillna(0)
    df_thuc_tang_to = _finalize_thuc_tang_df(
        df_thuc_tang_to,
        ["Thực tăng", "TTVT", "Đội VT"],
        {"TTVT": "TỔNG CỘNG", "Đội VT": ""},
    )

    processed_path = _ensure_generated_workbook(output_path, overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "thuc_tang_theo_to", df_thuc_tang_to)
    append_or_replace_sheet(
        processed_path,
        "thong_bao",
        pd.DataFrame(
            [
                {
                    "Trạng thái": "Không có dữ liệu NVKT",
                    "Ghi chú": "Nguồn MyTV ngưng PSC API hiện chỉ có cấp tổ/TTVT, chưa có chi tiết theo NVKT nên không sinh sheet thuc_tang_theo_NVKT.",
                }
            ]
        ),
    )
    _remove_empty_default_sheet(processed_path)
    return processed_path


def process_son_tay_mytv_ngung_psc_t_minus_1_api_output(
    input_path=DEFAULT_SON_TAY_MYTV_NGUNG_T_MINUS_1_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao MyTV Son Tay ngung PSC thang T-1."""
    raw_path = _resolve_path(input_path)
    df_raw = pd.read_excel(raw_path, header=None)
    df_subset = _extract_son_tay_ngung_psc_subset(df_raw)

    numeric_cols = list(df_subset.columns[1:])
    for col in numeric_cols:
        df_subset[col] = pd.to_numeric(df_subset[col], errors="coerce").fillna(0).astype(int)

    total_row = pd.DataFrame(
        [{
            "Đơn vị/Nhân viên KT": "Tổng",
            "Hoàn công(*) (1.5)": int(df_subset["Hoàn công(*) (1.5)"].sum()),
            "Lũy kế tháng(1.6)": int(df_subset["Lũy kế tháng(1.6)"].sum()),
            "Lũy kế năm(1.7)": int(df_subset["Lũy kế năm(1.7)"].sum()),
            "Ngưng PSC tạm tính tháng T(5.1)": int(df_subset["Ngưng PSC tạm tính tháng T(5.1)"].sum()),
            "TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)": int(df_subset["TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)"].sum()),
        }]
    )
    df_final = pd.concat([df_subset, total_row], ignore_index=True)

    processed_path = _ensure_processed_workbook_for_group(raw_path, "tam_dung_khoi_phuc_dich_vu", overwrite=overwrite_processed)
    append_or_replace_sheet(processed_path, "TH_ngung_PSC-Thang T-1", df_final)
    return processed_path


def process_son_tay_fiber_ngung_psc_t_minus_1_api_output(
    input_path=DEFAULT_FIBER_T_MINUS_1_CAP_TO_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao Fiber Son Tay ngung PSC thang T-1."""
    raw_path = _resolve_path(input_path)
    df_raw = pd.read_excel(raw_path, header=None)
    df_subset = _extract_son_tay_ngung_psc_subset(df_raw)

    numeric_cols = list(df_subset.columns[1:])
    for col in numeric_cols:
        df_subset[col] = pd.to_numeric(df_subset[col], errors="coerce").fillna(0).astype(int)

    total_row = pd.DataFrame(
        [{
            "Đơn vị/Nhân viên KT": "Tổng",
            "Hoàn công(*) (1.5)": int(df_subset["Hoàn công(*) (1.5)"].sum()),
            "Lũy kế tháng(1.6)": int(df_subset["Lũy kế tháng(1.6)"].sum()),
            "Lũy kế năm(1.7)": int(df_subset["Lũy kế năm(1.7)"].sum()),
            "Ngưng PSC tạm tính tháng T(5.1)": int(df_subset["Ngưng PSC tạm tính tháng T(5.1)"].sum()),
            "TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)": int(df_subset["TB Ngưng PSC lũy kế năm(4.6) (4.7-4.4)"].sum()),
        }]
    )
    df_final = pd.concat([df_subset, total_row], ignore_index=True)

    processed_path = _ensure_processed_workbook_for_group(
        raw_path,
        "tam_dung_khoi_phuc_dich_vu",
        overwrite=overwrite_processed,
    )
    append_or_replace_sheet(processed_path, "TH_ngung_PSC-Thang T-1", df_final)
    return processed_path


def process_son_tay_fiber_ngung_psc_t_minus_1_ttvt_api_output(
    input_path=DEFAULT_FIBER_T_MINUS_1_CAP_TTVT_INPUT,
    overwrite_processed=False,
):
    """Xu ly bao cao Fiber Son Tay ngung PSC thang T-1 cap TTVT."""
    raw_path = _resolve_path(input_path)
    df_raw = pd.read_excel(raw_path, header=None).copy()
    if len(df_raw) <= 3:
        raise ValueError(
            f"File Fiber TTVT khong du dong du lieu. Can it nhat 4 dong, nhung chi co {len(df_raw)} dong."
        )

    df_data = df_raw.iloc[3:].reset_index(drop=True).copy()
    df_data.columns = _flatten_header_rows(df_raw, header_rows=3)
    first_col = df_data.columns[0]
    df_data = df_data[df_data[first_col].notna() & (df_data[first_col].astype(str).str.strip() != "")].reset_index(drop=True)

    processed_path = _ensure_processed_workbook_for_group(
        raw_path,
        "tam_dung_khoi_phuc_dich_vu",
        overwrite=overwrite_processed,
    )
    append_or_replace_sheet(processed_path, "TH_ngung_PSC-Thang T-1", df_data)
    return processed_path
