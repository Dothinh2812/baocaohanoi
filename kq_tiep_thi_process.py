# -*- coding: utf-8 -*-
"""
Module chứa các hàm xử lý báo cáo Kết quả Tiếp thị
"""
import os
import pandas as pd
from datetime import datetime

def process_kq_tiep_thi_report():
    """
    Xử lý báo cáo Kết quả Tiếp thị:
    1. Đọc file kq_tiep_thi_DDMMYYYY.xlsx trong thư mục KQ-TIEP-THI
    2. Loại bỏ 4 cột theo yêu cầu
    3. Lưu vào sheet mới 'kq_tiep_thi' trong cùng file
    """
    print("\n=== Bắt đầu xử lý báo cáo Kết quả Tiếp thị ===")

    try:
        # 1. Tìm file mới nhất
        download_dir = "KQ-TIEP-THI"
        date_str = datetime.now().strftime("%d%m%Y")
        file_path = os.path.join(download_dir, f"kq_tiep_thi_{date_str}.xlsx")

        if not os.path.exists(file_path):
            # Thử tìm file không có ngày nếu file có ngày không tồn tại
            file_path = os.path.join(download_dir, "kq_tiep_thi.xlsx")
            if not os.path.exists(file_path):
                print(f"❌ Không tìm thấy file báo cáo trong thư mục {download_dir}")
                return

        print(f"📂 Đang xử lý file: {file_path}")

        # 2. Đọc file với header 2 dòng (MultiIndex)
        df = pd.read_excel(file_path, header=[0, 1])
        print(f"✅ Đọc thành công {len(df)} dòng dữ liệu")

        # Chuẩn hóa tên cột (xóa khoảng trắng thừa)
        # Nếu dòng 2 giống dòng 1 hoặc là 'Unnamed', thì xóa dòng 2
        new_cols = []
        for a, b in df.columns:
            a_clean = str(a).strip()
            b_clean = str(b).strip()
            if 'Unnamed' in b_clean or b_clean == a_clean:
                new_cols.append((a_clean, ''))
            else:
                new_cols.append((a_clean, b_clean))
        df.columns = pd.MultiIndex.from_tuples(new_cols)

        # 3. Xác định các cột cần bỏ
        cols_to_drop = [
            ('Dịch vụ BRCĐ', 'Kết quả thực hiện tuần'),
            ('Dịch vụ BRCĐ', 'Số liệu tiếp thị thuê bao lũy kế'),
            ('Dịch vụ MyTV', 'Kết quả thực hiện tuần'),
            ('Dịch vụ MyTV', 'Số liệu tiếp thị thuê bao lũy kế')
        ]

        # Kiểm tra sự tồn tại của các cột trước khi bỏ
        cols_present = [col for col in cols_to_drop if col in df.columns]
        
        if len(cols_present) < len(cols_to_drop):
            print("⚠️ Cảnh báo: Một số cột yêu cầu không tìm thấy trong file.")
            print(f"   Các cột tìm thấy: {cols_present}")
            print(f"   Các cột có sẵn: {df.columns.tolist()}")

        # Thực hiện loại bỏ cột
        df_processed = df.drop(columns=cols_present)
        print(f"✅ Đã loại bỏ {len(cols_present)} cột. Còn lại {len(df_processed.columns)} cột.")

        # 3.0 Bổ sung cột Đơn vị từ dsnv.xlsx
        print("🔍 Đang bổ sung cột Đơn vị từ dsnv.xlsx...")
        dsnv_path = "dsnv.xlsx"
        if os.path.exists(dsnv_path):
            df_dsnv = pd.read_excel(dsnv_path)
            # Tạo mapping từ Họ tên -> đơn vị
            # Chuẩn hóa tên để tránh lỗi khoảng trắng
            df_dsnv['Họ tên'] = df_dsnv['Họ tên'].astype(str).str.strip()
            unit_mapping = dict(zip(df_dsnv['Họ tên'], df_dsnv['đơn vị']))
            
            # Hàm lấy đơn vị cho từng nhân viên
            def get_unit(name):
                if pd.isna(name): return ""
                return unit_mapping.get(str(name).strip(), "")
            
            # Lấy cột Tên NV
            col_name = ('Tên NV', '')
            if col_name in df_processed.columns:
                units = df_processed[col_name].apply(get_unit)
                # Chèn cột Đơn vị sau cột STT (vị trí 1)
                df_processed.insert(1, ('Đơn vị', ''), units)
                print("✅ Đã bổ sung cột Đơn vị.")
            else:
                print("⚠️ Không tìm thấy cột 'Tên NV' để mapping đơn vị.")
        else:
            print(f"⚠️ Không tìm thấy file {dsnv_path}, bỏ qua bước bổ sung đơn vị.")

        # 3.1 Thêm cột Tổng ở cuối
        print("📊 Đang tính tổng...")
        col_brcd = ('Dịch vụ BRCĐ', 'Kết quả thực hiện trong tháng')
        col_mytv = ('Dịch vụ MyTV', 'Kết quả thực hiện trong tháng')
        
        # Đảm bảo các cột số liệu là kiểu số
        df_processed[col_brcd] = pd.to_numeric(df_processed[col_brcd], errors='coerce').fillna(0)
        df_processed[col_mytv] = pd.to_numeric(df_processed[col_mytv], errors='coerce').fillna(0)
        
        # Tính tổng
        df_processed[('Tổng', '')] = df_processed[col_brcd] + df_processed[col_mytv]
        print("✅ Đã thêm cột Tổng.")

        # 3.1.1 Tạo bảng tổng hợp theo Đơn vị (Sheet kq_th)
        print("📊 Đang tạo bảng tổng hợp theo Đơn vị...")
        
        # Kiểm tra xem có cột Đơn vị không
        if ('Đơn vị', '') in df_processed.columns:
            summary_cols = [col_brcd, col_mytv, ('Tổng', '')]
            df_summary = df_processed.groupby(('Đơn vị', ''))[summary_cols].sum().reset_index()
            
            # Thêm STT cho bảng tổng hợp
            df_summary.insert(0, ('STT', ''), range(1, len(df_summary) + 1))
            
            # Thêm hàng Tổng Cộng cho bảng tổng hợp
            summary_total_row = {col: 0 for col in df_summary.columns}
            summary_total_row[('Đơn vị', '')] = 'TỔNG CỘNG'
            summary_total_row[('STT', '')] = ''
            summary_total_row[col_brcd] = df_summary[col_brcd].sum()
            summary_total_row[col_mytv] = df_summary[col_mytv].sum()
            summary_total_row[('Tổng', '')] = df_summary[('Tổng', '')].sum()
            df_summary = pd.concat([df_summary, pd.DataFrame([summary_total_row])], ignore_index=True)
            print("✅ Đã tạo bảng tổng hợp.")
        else:
            print("⚠️ Không có dữ liệu 'Đơn vị', tạo bảng tổng hợp rỗng.")
            df_summary = pd.DataFrame(columns=[('STT', ''), ('Đơn vị', ''), col_brcd, col_mytv, ('Tổng', '')])

        # 3.2 Thêm hàng Tổng ở dưới cùng bản chi tiết
        print("📊 Đang tạo hàng Tổng Cộng...")
        total_row = {col: 0 for col in df_processed.columns}
        total_row[('Tên NV', '')] = 'TỔNG CỘNG'
        if ('Đơn vị', '') in total_row:
            total_row[('Đơn vị', '')] = ''
        total_row[col_brcd] = df_processed[col_brcd].sum()
        total_row[col_mytv] = df_processed[col_mytv].sum()
        total_row[('Tổng', '')] = df_processed[('Tổng', '')].sum()
        
        # Append hàng tổng
        df_processed = pd.concat([df_processed, pd.DataFrame([total_row])], ignore_index=True)
        print("✅ Đã thêm hàng Tổng Cộng.")

        # 4. Lưu lại vào file (thêm sheet mới)
        print(f"💾 Đang lưu vào sheet 'kq_tiep_thi'...")
        
        # Để kiểm soát hoàn toàn header, chúng ta sẽ tự tạo header ở dòng 1
        # và ghi dữ liệu từ dòng 2 (startrow=1)
        
        df_for_save = df_processed.copy()
        # Lấy nhãn header dòng 1
        header_labels = [col[0] for col in df_for_save.columns]
        # Làm phẳng tên cột để tránh pandas tự viết MultiIndex
        df_for_save.columns = [f"Col_{i}" for i in range(len(df_for_save.columns))]

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # --- Ghi sheet CHI TIẾT ---
            df_for_save.to_excel(writer, sheet_name='kq_tiep_thi', index=False, header=False, startrow=1)
            ws = writer.sheets['kq_tiep_thi']
            
            from openpyxl.styles import Alignment, Font, Border, Side
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

            # Ghi Header dòng 1
            for col_idx, label in enumerate(header_labels, 1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Định dạng hàng TỔNG CỘNG (hàng cuối cùng)
            last_row_idx = ws.max_row
            for col_idx in range(1, len(header_labels) + 1):
                cell = ws.cell(row=last_row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border

            # Tự động chỉnh độ rộng cột
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 5, 50)

            # --- Ghi sheet TỔNG HỢP ---
            print(f"💾 Đang lưu vào sheet 'kq_th'...")
            # Chuẩn bị dữ liệu để lưu
            summary_header_labels = [col[0] for col in df_summary.columns]
            df_summary_save = df_summary.copy()
            df_summary_save.columns = [f"Col_{i}" for i in range(len(df_summary_save.columns))]
            
            df_summary_save.to_excel(writer, sheet_name='kq_th', index=False, header=False, startrow=1)
            ws_th = writer.sheets['kq_th']
            
            # Ghi Header dòng 1
            for col_idx, label in enumerate(summary_header_labels, 1):
                cell = ws_th.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Định dạng hàng TỔNG CỘNG (hàng cuối cùng)
            last_row_idx_th = ws_th.max_row
            for col_idx in range(1, len(summary_header_labels) + 1):
                cell = ws_th.cell(row=last_row_idx_th, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border
                
            # Tự động chỉnh độ rộng cột
            for column in ws_th.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_th.column_dimensions[column_letter].width = min(max_length + 5, 50)

        print(f"✅ Hoàn thành! Báo cáo đã được lưu vào sheet 'kq_tiep_thi' và 'kq_th' của file: {file_path}")

    except Exception as e:
        print(f"❌ Lỗi khi xử lý báo cáo: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    process_kq_tiep_thi_report()
