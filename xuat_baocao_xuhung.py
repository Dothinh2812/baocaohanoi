#!/usr/bin/env python3
"""
Script tạo báo cáo xu hướng theo tháng từ database baocao_hanoi_thang.db.
Xuất file Excel với các sheet thể hiện xu hướng qua các tháng cho:
- Từng đơn vị (C1.1, C1.2, C1.3, C1.4)
- Từng NVKT (C1.4 chi tiết, SM1-C12, SM4-C11)

Usage:
    python xuat_baocao_xuhung.py
    python xuat_baocao_xuhung.py --output "bao_cao_xu_huong_2025.xlsx"
"""

import argparse
import sqlite3
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

# Cấu hình
DB_PATH = Path(__file__).parent / "baocao_hanoi_thang.db"
OUTPUT_DIR = Path(__file__).parent / "downloads" / "baocao_hanoi"


def get_available_months(conn: sqlite3.Connection) -> list:
    """Lấy danh sách các tháng có dữ liệu."""
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT thang_bao_cao FROM bao_cao_c11 ORDER BY thang_bao_cao")
    return [row[0] for row in cursor.fetchall()]


def create_trend_c11(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng C1.1 theo đơn vị."""
    query = """
    SELECT 
        d.ten_don_vi,
        c.thang_bao_cao,
        c.sm1_cl_chu_dong,
        c.sm2_cl_chu_dong,
        c.ty_le_cl_chu_dong,
        c.sm3_brcd,
        c.sm4_brcd,
        c.ty_le_brcd,
        c.chi_tieu_bsc
    FROM bao_cao_c11 c
    JOIN don_vi d ON c.don_vi_id = d.id
    ORDER BY d.ten_don_vi, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    # Pivot để mỗi tháng là một nhóm cột
    pivot_data = []
    for don_vi in df['ten_don_vi'].unique():
        row = {'Đơn vị': don_vi}
        dv_data = df[df['ten_don_vi'] == don_vi]
        
        for _, record in dv_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_SM1'] = record['sm1_cl_chu_dong']
            row[f'{month}_SM2'] = record['sm2_cl_chu_dong']
            row[f'{month}_TL_CLCD'] = record['ty_le_cl_chu_dong']
            row[f'{month}_SM3'] = record['sm3_brcd']
            row[f'{month}_SM4'] = record['sm4_brcd']
            row[f'{month}_TL_BRCD'] = record['ty_le_brcd']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_clcd_first = row.get(f'{first_month}_TL_CLCD')
            tl_clcd_last = row.get(f'{last_month}_TL_CLCD')
            tl_brcd_first = row.get(f'{first_month}_TL_BRCD')
            tl_brcd_last = row.get(f'{last_month}_TL_BRCD')
            
            if tl_clcd_first and tl_clcd_last:
                row['XH_TL_CLCD'] = tl_clcd_last - tl_clcd_first
            if tl_brcd_first and tl_brcd_last:
                row['XH_TL_BRCD'] = tl_brcd_last - tl_brcd_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_c12(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng C1.2 theo đơn vị."""
    query = """
    SELECT 
        d.ten_don_vi,
        c.thang_bao_cao,
        c.sm1_lap_lai,
        c.sm2_lap_lai,
        c.ty_le_lap_lai,
        c.sm3_su_co,
        c.sm4_su_co,
        c.ty_le_su_co
    FROM bao_cao_c12 c
    JOIN don_vi d ON c.don_vi_id = d.id
    ORDER BY d.ten_don_vi, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for don_vi in df['ten_don_vi'].unique():
        row = {'Đơn vị': don_vi}
        dv_data = df[df['ten_don_vi'] == don_vi]
        
        for _, record in dv_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_SM1'] = record['sm1_lap_lai']
            row[f'{month}_SM2'] = record['sm2_lap_lai']
            row[f'{month}_TL_LL'] = record['ty_le_lap_lai']
            row[f'{month}_SM3'] = record['sm3_su_co']
            row[f'{month}_SM4'] = record['sm4_su_co']
            row[f'{month}_TL_SC'] = record['ty_le_su_co']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_ll_first = row.get(f'{first_month}_TL_LL')
            tl_ll_last = row.get(f'{last_month}_TL_LL')
            tl_sc_first = row.get(f'{first_month}_TL_SC')
            tl_sc_last = row.get(f'{last_month}_TL_SC')
            
            if tl_ll_first and tl_ll_last:
                row['XH_TL_LapLai'] = tl_ll_last - tl_ll_first
            if tl_sc_first and tl_sc_last:
                row['XH_TL_SuCo'] = tl_sc_last - tl_sc_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_c13(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng C1.3 theo đơn vị."""
    query = """
    SELECT 
        d.ten_don_vi,
        c.thang_bao_cao,
        c.ty_le_sua_chua,
        c.ty_le_lap_lai,
        c.ty_le_su_co
    FROM bao_cao_c13 c
    JOIN don_vi d ON c.don_vi_id = d.id
    ORDER BY d.ten_don_vi, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for don_vi in df['ten_don_vi'].unique():
        row = {'Đơn vị': don_vi}
        dv_data = df[df['ten_don_vi'] == don_vi]
        
        for _, record in dv_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_TL_SC'] = record['ty_le_sua_chua']
            row[f'{month}_TL_LL'] = record['ty_le_lap_lai']
            row[f'{month}_TL_SCo'] = record['ty_le_su_co']
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_c14(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng C1.4 theo đơn vị."""
    query = """
    SELECT 
        d.ten_don_vi,
        c.thang_bao_cao,
        c.tong_phieu,
        c.sl_kh_hai_long,
        c.ty_le_hl_kt_phuc_vu,
        c.ty_le_hl_kt_dich_vu,
        c.ty_le_kh_hai_long,
        c.diem_bsc
    FROM bao_cao_c14 c
    JOIN don_vi d ON c.don_vi_id = d.id
    ORDER BY d.ten_don_vi, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for don_vi in df['ten_don_vi'].unique():
        row = {'Đơn vị': don_vi}
        dv_data = df[df['ten_don_vi'] == don_vi]
        
        for _, record in dv_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_TongPhieu'] = record['tong_phieu']
            row[f'{month}_SL_HL'] = record['sl_kh_hai_long']
            row[f'{month}_TL_HL_PV'] = record['ty_le_hl_kt_phuc_vu']
            row[f'{month}_TL_HL_DV'] = record['ty_le_hl_kt_dich_vu']
            row[f'{month}_TL_HL'] = record['ty_le_kh_hai_long']
            row[f'{month}_DiemBSC'] = record['diem_bsc']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_hl_first = row.get(f'{first_month}_TL_HL')
            tl_hl_last = row.get(f'{last_month}_TL_HL')
            
            if tl_hl_first and tl_hl_last:
                row['XH_TL_HaiLong'] = tl_hl_last - tl_hl_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_c14_nvkt(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng C1.4 theo NVKT."""
    query = """
    SELECT 
        d.ten_don_vi,
        n.ten_nvkt,
        c.thang_bao_cao,
        c.tong_phieu_ks_thanh_cong,
        c.tong_phieu_khl,
        c.ty_le_hai_long
    FROM bao_cao_c14_nvkt c
    JOIN nhan_vien_kt n ON c.nvkt_id = n.id
    JOIN don_vi d ON n.don_vi_id = d.id
    ORDER BY d.ten_don_vi, n.ten_nvkt, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for (don_vi, nvkt) in df[['ten_don_vi', 'ten_nvkt']].drop_duplicates().values:
        row = {'Đơn vị': don_vi, 'NVKT': nvkt}
        nvkt_data = df[(df['ten_don_vi'] == don_vi) & (df['ten_nvkt'] == nvkt)]
        
        for _, record in nvkt_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_TongPhieu'] = record['tong_phieu_ks_thanh_cong']
            row[f'{month}_SL_HL'] = record['tong_phieu_khl']
            row[f'{month}_TL_HL'] = record['ty_le_hai_long']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_first = row.get(f'{first_month}_TL_HL')
            tl_last = row.get(f'{last_month}_TL_HL')
            
            if tl_first is not None and tl_last is not None:
                row['XH_TL_HL'] = tl_last - tl_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_sm1c12(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng SM1-C12 (Hỏng lại) theo NVKT."""
    query = """
    SELECT 
        d.ten_don_vi,
        n.ten_nvkt,
        c.thang_bao_cao,
        c.so_phieu_hll,
        c.so_phieu_bao_hong,
        c.ty_le_hll
    FROM bao_cao_sm1c12_hll c
    JOIN nhan_vien_kt n ON c.nvkt_id = n.id
    JOIN don_vi d ON n.don_vi_id = d.id
    ORDER BY d.ten_don_vi, n.ten_nvkt, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for (don_vi, nvkt) in df[['ten_don_vi', 'ten_nvkt']].drop_duplicates().values:
        row = {'Đơn vị': don_vi, 'NVKT': nvkt}
        nvkt_data = df[(df['ten_don_vi'] == don_vi) & (df['ten_nvkt'] == nvkt)]
        
        for _, record in nvkt_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_HLL'] = record['so_phieu_hll']
            row[f'{month}_BaoHong'] = record['so_phieu_bao_hong']
            row[f'{month}_TL_HLL'] = record['ty_le_hll']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_first = row.get(f'{first_month}_TL_HLL')
            tl_last = row.get(f'{last_month}_TL_HLL')
            
            if tl_first is not None and tl_last is not None:
                row['XH_TL_HLL'] = tl_last - tl_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_sm4c11(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng SM4-C11 (BRCD chi tiết) theo NVKT."""
    query = """
    SELECT 
        d.ten_don_vi,
        n.ten_nvkt,
        c.thang_bao_cao,
        c.tong_phieu,
        c.so_phieu_dat,
        c.ty_le_dat
    FROM bao_cao_sm4c11_chitiet c
    JOIN nhan_vien_kt n ON c.nvkt_id = n.id
    JOIN don_vi d ON n.don_vi_id = d.id
    ORDER BY d.ten_don_vi, n.ten_nvkt, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for (don_vi, nvkt) in df[['ten_don_vi', 'ten_nvkt']].drop_duplicates().values:
        row = {'Đơn vị': don_vi, 'NVKT': nvkt}
        nvkt_data = df[(df['ten_don_vi'] == don_vi) & (df['ten_nvkt'] == nvkt)]
        
        for _, record in nvkt_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_TongPhieu'] = record['tong_phieu']
            row[f'{month}_Dat'] = record['so_phieu_dat']
            row[f'{month}_TL_Dat'] = record['ty_le_dat']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_first = row.get(f'{first_month}_TL_Dat')
            tl_last = row.get(f'{last_month}_TL_Dat')
            
            if tl_first is not None and tl_last is not None:
                row['XH_TL_Dat'] = tl_last - tl_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_trend_sm4c11_18h(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo báo cáo xu hướng SM4-C11 18h theo NVKT."""
    query = """
    SELECT 
        d.ten_don_vi,
        n.ten_nvkt,
        c.thang_bao_cao,
        c.tong_phieu,
        c.so_phieu_dat,
        c.ty_le_dat
    FROM bao_cao_sm4c11_18h c
    JOIN nhan_vien_kt n ON c.nvkt_id = n.id
    JOIN don_vi d ON n.don_vi_id = d.id
    ORDER BY d.ten_don_vi, n.ten_nvkt, c.thang_bao_cao
    """
    df = pd.read_sql_query(query, conn)
    
    pivot_data = []
    for (don_vi, nvkt) in df[['ten_don_vi', 'ten_nvkt']].drop_duplicates().values:
        row = {'Đơn vị': don_vi, 'NVKT': nvkt}
        nvkt_data = df[(df['ten_don_vi'] == don_vi) & (df['ten_nvkt'] == nvkt)]
        
        for _, record in nvkt_data.iterrows():
            month = record['thang_bao_cao']
            row[f'{month}_TongPhieu'] = record['tong_phieu']
            row[f'{month}_Dat'] = record['so_phieu_dat']
            row[f'{month}_TL_Dat'] = record['ty_le_dat']
        
        # Tính xu hướng
        if len(months) >= 2:
            first_month = months[0]
            last_month = months[-1]
            tl_first = row.get(f'{first_month}_TL_Dat')
            tl_last = row.get(f'{last_month}_TL_Dat')
            
            if tl_first is not None and tl_last is not None:
                row['XH_TL_Dat'] = tl_last - tl_first
        
        pivot_data.append(row)
    
    return pd.DataFrame(pivot_data)


def create_summary_sheet(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo sheet tổng hợp."""
    summary = []
    
    # Tổng hợp C1.1 - TTVT Sơn Tây
    cursor = conn.cursor()
    cursor.execute("""
        SELECT c.thang_bao_cao, c.ty_le_cl_chu_dong, c.ty_le_brcd
        FROM bao_cao_c11 c
        JOIN don_vi d ON c.don_vi_id = d.id
        WHERE d.ten_don_vi = 'TTVT Sơn Tây'
        ORDER BY c.thang_bao_cao
    """)
    for row in cursor.fetchall():
        summary.append({
            'Chỉ tiêu': 'C1.1 - Tỷ lệ CLCD',
            'Tháng': row[0],
            'Giá trị': row[1]
        })
        summary.append({
            'Chỉ tiêu': 'C1.1 - Tỷ lệ BRCD',
            'Tháng': row[0],
            'Giá trị': row[2]
        })
    
    # Tổng hợp C1.2 - TTVT Sơn Tây
    cursor.execute("""
        SELECT c.thang_bao_cao, c.ty_le_lap_lai, c.ty_le_su_co
        FROM bao_cao_c12 c
        JOIN don_vi d ON c.don_vi_id = d.id
        WHERE d.ten_don_vi = 'TTVT Sơn Tây'
        ORDER BY c.thang_bao_cao
    """)
    for row in cursor.fetchall():
        summary.append({
            'Chỉ tiêu': 'C1.2 - Tỷ lệ lặp lại',
            'Tháng': row[0],
            'Giá trị': row[1]
        })
        summary.append({
            'Chỉ tiêu': 'C1.2 - Tỷ lệ sự cố',
            'Tháng': row[0],
            'Giá trị': row[2]
        })
    
    # Tổng hợp C1.4 - TTVT Sơn Tây
    cursor.execute("""
        SELECT c.thang_bao_cao, c.ty_le_kh_hai_long
        FROM bao_cao_c14 c
        JOIN don_vi d ON c.don_vi_id = d.id
        WHERE d.ten_don_vi = 'TTVT Sơn Tây'
        ORDER BY c.thang_bao_cao
    """)
    for row in cursor.fetchall():
        summary.append({
            'Chỉ tiêu': 'C1.4 - Tỷ lệ hài lòng',
            'Tháng': row[0],
            'Giá trị': row[1]
        })
    
    df = pd.DataFrame(summary)
    
    # Pivot để mỗi tháng là một cột
    if not df.empty:
        pivot_df = df.pivot(index='Chỉ tiêu', columns='Tháng', values='Giá trị').reset_index()
        return pivot_df
    
    return df


def create_summary_by_donvi(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo sheet tổng hợp theo từng đơn vị - tất cả chỉ số."""
    cursor = conn.cursor()
    
    # Lấy danh sách đơn vị
    cursor.execute("SELECT id, ten_don_vi FROM don_vi ORDER BY ten_don_vi")
    don_vis = cursor.fetchall()
    
    all_data = []
    
    for don_vi_id, ten_don_vi in don_vis:
        for month in months:
            row = {'Đơn vị': ten_don_vi, 'Tháng': month}
            
            # C1.1
            cursor.execute("""
                SELECT ty_le_cl_chu_dong, ty_le_brcd 
                FROM bao_cao_c11 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c11 = cursor.fetchone()
            if c11:
                row['C1.1_TL_CLCD'] = c11[0]
                row['C1.1_TL_BRCD'] = c11[1]
            
            # C1.2
            cursor.execute("""
                SELECT ty_le_lap_lai, ty_le_su_co 
                FROM bao_cao_c12 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c12 = cursor.fetchone()
            if c12:
                row['C1.2_TL_LapLai'] = c12[0]
                row['C1.2_TL_SuCo'] = c12[1]
            
            # C1.3
            cursor.execute("""
                SELECT ty_le_sua_chua, ty_le_lap_lai, ty_le_su_co 
                FROM bao_cao_c13 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c13 = cursor.fetchone()
            if c13:
                row['C1.3_TL_SuaChua'] = c13[0]
                row['C1.3_TL_LapLai'] = c13[1]
                row['C1.3_TL_SuCo'] = c13[2]
            
            # C1.4
            cursor.execute("""
                SELECT ty_le_kh_hai_long, diem_bsc 
                FROM bao_cao_c14 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c14 = cursor.fetchone()
            if c14:
                row['C1.4_TL_HaiLong'] = c14[0]
                row['C1.4_DiemBSC'] = c14[1]
            
            all_data.append(row)
    
    return pd.DataFrame(all_data)


def create_summary_by_nvkt(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo sheet tổng hợp theo từng NVKT - tất cả chỉ số."""
    cursor = conn.cursor()
    
    # Lấy danh sách NVKT với đơn vị
    cursor.execute("""
        SELECT n.id, d.ten_don_vi, n.ten_nvkt 
        FROM nhan_vien_kt n
        JOIN don_vi d ON n.don_vi_id = d.id
        ORDER BY d.ten_don_vi, n.ten_nvkt
    """)
    nvkts = cursor.fetchall()
    
    all_data = []
    
    for nvkt_id, ten_don_vi, ten_nvkt in nvkts:
        for month in months:
            row = {'Đơn vị': ten_don_vi, 'NVKT': ten_nvkt, 'Tháng': month}
            
            # C1.4 chi tiết NVKT
            cursor.execute("""
                SELECT tong_phieu_ks_thanh_cong, tong_phieu_khl, ty_le_hai_long 
                FROM bao_cao_c14_nvkt WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            c14 = cursor.fetchone()
            if c14:
                row['C1.4_TongPhieu'] = c14[0]
                row['C1.4_SL_HL'] = c14[1]
                row['C1.4_TL_HL'] = c14[2]
            
            # SM1-C12 Hỏng lại
            cursor.execute("""
                SELECT so_phieu_hll, so_phieu_bao_hong, ty_le_hll 
                FROM bao_cao_sm1c12_hll WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm1 = cursor.fetchone()
            if sm1:
                row['SM1C12_HLL'] = sm1[0]
                row['SM1C12_BaoHong'] = sm1[1]
                row['SM1C12_TL_HLL'] = sm1[2]
            
            # SM4-C11 chi tiết
            cursor.execute("""
                SELECT tong_phieu, so_phieu_dat, ty_le_dat 
                FROM bao_cao_sm4c11_chitiet WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm4 = cursor.fetchone()
            if sm4:
                row['SM4C11_TongPhieu'] = sm4[0]
                row['SM4C11_Dat'] = sm4[1]
                row['SM4C11_TL_Dat'] = sm4[2]
            
            # SM4-C11 18h
            cursor.execute("""
                SELECT tong_phieu, so_phieu_dat, ty_le_dat 
                FROM bao_cao_sm4c11_18h WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm4_18h = cursor.fetchone()
            if sm4_18h:
                row['SM4C11_18h_TongPhieu'] = sm4_18h[0]
                row['SM4C11_18h_Dat'] = sm4_18h[1]
                row['SM4C11_18h_TL_Dat'] = sm4_18h[2]
            
            all_data.append(row)
    
    return pd.DataFrame(all_data)


def create_pivot_by_donvi(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo bảng pivot theo đơn vị - mỗi hàng là 1 đơn vị, cột là chỉ số theo tháng."""
    cursor = conn.cursor()
    
    # Lấy danh sách đơn vị
    cursor.execute("SELECT id, ten_don_vi FROM don_vi ORDER BY ten_don_vi")
    don_vis = cursor.fetchall()
    
    all_data = []
    
    for don_vi_id, ten_don_vi in don_vis:
        row = {'Đơn vị': ten_don_vi}
        
        for month in months:
            # C1.1
            cursor.execute("""
                SELECT ty_le_cl_chu_dong, ty_le_brcd 
                FROM bao_cao_c11 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c11 = cursor.fetchone()
            if c11:
                row[f'{month}_C11_CLCD'] = c11[0]
                row[f'{month}_C11_BRCD'] = c11[1]
            
            # C1.2
            cursor.execute("""
                SELECT ty_le_lap_lai, ty_le_su_co 
                FROM bao_cao_c12 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c12 = cursor.fetchone()
            if c12:
                row[f'{month}_C12_LL'] = c12[0]
                row[f'{month}_C12_SC'] = c12[1]
            
            # C1.4
            cursor.execute("""
                SELECT ty_le_kh_hai_long 
                FROM bao_cao_c14 WHERE don_vi_id = ? AND thang_bao_cao = ?
            """, (don_vi_id, month))
            c14 = cursor.fetchone()
            if c14:
                row[f'{month}_C14_HL'] = c14[0]
        
        all_data.append(row)
    
    return pd.DataFrame(all_data)


def create_pivot_by_nvkt(conn: sqlite3.Connection, months: list) -> pd.DataFrame:
    """Tạo bảng pivot theo NVKT - mỗi hàng là 1 NVKT, cột là chỉ số theo tháng."""
    cursor = conn.cursor()
    
    # Lấy danh sách NVKT
    cursor.execute("""
        SELECT n.id, d.ten_don_vi, n.ten_nvkt 
        FROM nhan_vien_kt n
        JOIN don_vi d ON n.don_vi_id = d.id
        ORDER BY d.ten_don_vi, n.ten_nvkt
    """)
    nvkts = cursor.fetchall()
    
    all_data = []
    
    for nvkt_id, ten_don_vi, ten_nvkt in nvkts:
        row = {'Đơn vị': ten_don_vi, 'NVKT': ten_nvkt}
        
        for month in months:
            # C1.4 chi tiết NVKT
            cursor.execute("""
                SELECT ty_le_hai_long 
                FROM bao_cao_c14_nvkt WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            c14 = cursor.fetchone()
            if c14:
                row[f'{month}_C14_HL'] = c14[0]
            
            # SM1-C12 Hỏng lại
            cursor.execute("""
                SELECT ty_le_hll 
                FROM bao_cao_sm1c12_hll WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm1 = cursor.fetchone()
            if sm1:
                row[f'{month}_SM1_HLL'] = sm1[0]
            
            # SM4-C11 chi tiết
            cursor.execute("""
                SELECT ty_le_dat 
                FROM bao_cao_sm4c11_chitiet WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm4 = cursor.fetchone()
            if sm4:
                row[f'{month}_SM4_Dat'] = sm4[0]
            
            # SM4-C11 18h
            cursor.execute("""
                SELECT ty_le_dat 
                FROM bao_cao_sm4c11_18h WHERE nvkt_id = ? AND thang_bao_cao = ?
            """, (nvkt_id, month))
            sm4_18h = cursor.fetchone()
            if sm4_18h:
                row[f'{month}_SM4_18h'] = sm4_18h[0]
        
        all_data.append(row)
    
    return pd.DataFrame(all_data)


def add_bar_chart_to_sheet(ws, title: str, data_cols: list, months: list, 
                           start_row: int = 2, category_col: int = 1,
                           chart_position: str = "A15", width: int = 18, height: int = 10):
    """
    Thêm biểu đồ bar chart vào worksheet.
    
    Args:
        ws: Worksheet object
        title: Tiêu đề biểu đồ
        data_cols: List các cột dữ liệu (1-indexed)
        months: Danh sách tháng
        start_row: Dòng bắt đầu dữ liệu
        category_col: Cột chứa tên category (đơn vị/NVKT)
        chart_position: Vị trí đặt biểu đồ
        width: Chiều rộng biểu đồ
        height: Chiều cao biểu đồ
    """
    max_row = ws.max_row
    
    if max_row < start_row:
        return
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = "Tỷ lệ (%)"
    chart.x_axis.title = "Đơn vị"
    chart.style = 10
    chart.width = width
    chart.height = height
    
    # Thêm dữ liệu
    for i, col in enumerate(data_cols):
        data = Reference(ws, min_col=col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    
    # Categories (tên đơn vị/NVKT)
    categories = Reference(ws, min_col=category_col, min_row=start_row, max_row=max_row)
    chart.set_categories(categories)
    
    # Hiển thị data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, chart_position)


def add_summary_chart(ws, months: list):
    """Thêm biểu đồ tổng hợp cho sheet Tong_hop."""
    max_row = ws.max_row
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Xu hướng các chỉ tiêu qua các tháng"
    chart.y_axis.title = "Tỷ lệ (%)"
    chart.style = 10
    chart.width = 18
    chart.height = 10
    
    # Thêm dữ liệu cho từng tháng
    for i, month in enumerate(months):
        col = i + 2  # Cột 2, 3, 4... cho các tháng
        data = Reference(ws, min_col=col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    
    # Categories (tên chỉ tiêu)
    categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.set_categories(categories)
    
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "A10")


def add_donvi_chart(ws, months: list, indicator_name: str, indicator_cols: list):
    """Thêm biểu đồ cho sheet đơn vị."""
    max_row = ws.max_row
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"Xu hướng {indicator_name} theo đơn vị"
    chart.y_axis.title = "Tỷ lệ (%)"
    chart.x_axis.title = "Đơn vị"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    
    # Thêm dữ liệu cho từng cột chỉ tiêu
    for col in indicator_cols:
        data = Reference(ws, min_col=col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    
    # Categories (tên đơn vị)
    categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.set_categories(categories)
    
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "A10")


def add_nvkt_chart(ws, months: list, indicator_name: str, indicator_cols: list):
    """Thêm biểu đồ cho sheet NVKT."""
    max_row = ws.max_row
    
    # Giới hạn số dòng để biểu đồ không quá phức tạp
    display_rows = min(max_row, 20)  # Chỉ hiển thị 18 NVKT đầu tiên
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"Xu hướng {indicator_name} theo NVKT (top 18)"
    chart.y_axis.title = "Tỷ lệ (%)"
    chart.x_axis.title = "NVKT"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    
    # Thêm dữ liệu cho từng cột chỉ tiêu
    for col in indicator_cols:
        data = Reference(ws, min_col=col, min_row=1, max_row=display_rows)
        chart.add_data(data, titles_from_data=True)
    
    # Categories (tên NVKT - cột 2)
    categories = Reference(ws, min_col=2, min_row=2, max_row=display_rows)
    chart.set_categories(categories)
    
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    # Đặt vị trí biểu đồ phía dưới dữ liệu
    chart_row = max_row + 3
    ws.add_chart(chart, f"A{chart_row}")


def export_trend_report(output_file: str = None):
    """Xuất báo cáo xu hướng ra file Excel."""
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"bao_cao_xu_huong_{timestamp}.xlsx"
    else:
        output_file = Path(output_file)
    
    print(f"Database: {DB_PATH}")
    print(f"Output: {output_file}")
    
    conn = sqlite3.connect(DB_PATH)
    
    try:
        months = get_available_months(conn)
        print(f"Các tháng có dữ liệu: {months}")
        
        if not months:
            print("❌ Không có dữ liệu trong database")
            return 1
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet tổng hợp
            print("  ✓ Tạo sheet Tổng hợp...")
            summary_df = create_summary_sheet(conn, months)
            summary_df.to_excel(writer, sheet_name='Tong_hop', index=False)
            
            # Xu hướng theo đơn vị
            print("  ✓ Tạo sheet C1.1 theo đơn vị...")
            c11_df = create_trend_c11(conn, months)
            c11_df.to_excel(writer, sheet_name='C1.1_DonVi', index=False)
            
            print("  ✓ Tạo sheet C1.2 theo đơn vị...")
            c12_df = create_trend_c12(conn, months)
            c12_df.to_excel(writer, sheet_name='C1.2_DonVi', index=False)
            
            print("  ✓ Tạo sheet C1.3 theo đơn vị...")
            c13_df = create_trend_c13(conn, months)
            c13_df.to_excel(writer, sheet_name='C1.3_DonVi', index=False)
            
            print("  ✓ Tạo sheet C1.4 theo đơn vị...")
            c14_df = create_trend_c14(conn, months)
            c14_df.to_excel(writer, sheet_name='C1.4_DonVi', index=False)
            
            # Xu hướng theo NVKT
            print("  ✓ Tạo sheet C1.4 theo NVKT...")
            c14_nvkt_df = create_trend_c14_nvkt(conn, months)
            c14_nvkt_df.to_excel(writer, sheet_name='C1.4_NVKT', index=False)
            
            print("  ✓ Tạo sheet SM1-C12 (Hỏng lại) theo NVKT...")
            sm1c12_df = create_trend_sm1c12(conn, months)
            sm1c12_df.to_excel(writer, sheet_name='SM1C12_HLL_NVKT', index=False)
            
            print("  ✓ Tạo sheet SM4-C11 (BRCD chi tiết) theo NVKT...")
            sm4c11_df = create_trend_sm4c11(conn, months)
            sm4c11_df.to_excel(writer, sheet_name='SM4C11_ChiTiet_NVKT', index=False)
            
            print("  ✓ Tạo sheet SM4-C11 18h theo NVKT...")
            sm4c11_18h_df = create_trend_sm4c11_18h(conn, months)
            sm4c11_18h_df.to_excel(writer, sheet_name='SM4C11_18h_NVKT', index=False)
            
            # === SHEET MỚI: Thống kê theo từng đơn vị (tất cả chỉ số) ===
            print("\n  📊 Tạo sheet thống kê theo đơn vị/NVKT...")
            
            print("  ✓ Tạo sheet Tổng hợp theo đơn vị (chi tiết)...")
            donvi_detail_df = create_summary_by_donvi(conn, months)
            donvi_detail_df.to_excel(writer, sheet_name='TH_DonVi_ChiTiet', index=False)
            
            print("  ✓ Tạo sheet Pivot theo đơn vị...")
            donvi_pivot_df = create_pivot_by_donvi(conn, months)
            donvi_pivot_df.to_excel(writer, sheet_name='Pivot_DonVi', index=False)
            
            print("  ✓ Tạo sheet Tổng hợp theo NVKT (chi tiết)...")
            nvkt_detail_df = create_summary_by_nvkt(conn, months)
            nvkt_detail_df.to_excel(writer, sheet_name='TH_NVKT_ChiTiet', index=False)
            
            print("  ✓ Tạo sheet Pivot theo NVKT...")
            nvkt_pivot_df = create_pivot_by_nvkt(conn, months)
            nvkt_pivot_df.to_excel(writer, sheet_name='Pivot_NVKT', index=False)
            
            # Thêm biểu đồ cho từng sheet
            print("\n  📊 Tạo biểu đồ...")
            workbook = writer.book
            
            # Biểu đồ sheet Tổng hợp
            print("    ✓ Biểu đồ Tổng hợp...")
            add_summary_chart(workbook['Tong_hop'], months)
            
            # Biểu đồ C1.1 - cột TL_BRCD cho mỗi tháng (cột 4, 7, 10 cho 3 tháng)
            print("    ✓ Biểu đồ C1.1 theo đơn vị...")
            tl_brcd_cols = [i * 6 + 4 for i in range(len(months))]  # Cột TL_BRCD cho mỗi tháng
            add_donvi_chart(workbook['C1.1_DonVi'], months, "C1.1 - Tỷ lệ BRCD", tl_brcd_cols)
            
            # Biểu đồ C1.2 - cột TL_LL cho mỗi tháng
            print("    ✓ Biểu đồ C1.2 theo đơn vị...")
            tl_ll_cols = [i * 6 + 4 for i in range(len(months))]  # Cột TL_LL
            add_donvi_chart(workbook['C1.2_DonVi'], months, "C1.2 - Tỷ lệ lặp lại", tl_ll_cols)
            
            # Biểu đồ C1.3 - cột TL_SC (sửa chữa)
            print("    ✓ Biểu đồ C1.3 theo đơn vị...")
            tl_sc_cols = [i * 3 + 2 for i in range(len(months))]
            add_donvi_chart(workbook['C1.3_DonVi'], months, "C1.3 - Tỷ lệ sửa chữa", tl_sc_cols)
            
            # Biểu đồ C1.4 - cột TL_HL (hài lòng)
            print("    ✓ Biểu đồ C1.4 theo đơn vị...")
            tl_hl_cols = [i * 6 + 6 for i in range(len(months))]
            add_donvi_chart(workbook['C1.4_DonVi'], months, "C1.4 - Tỷ lệ hài lòng", tl_hl_cols)
            
            # Biểu đồ C1.4 NVKT - cột TL_HL
            print("    ✓ Biểu đồ C1.4 theo NVKT...")
            tl_hl_nvkt_cols = [i * 3 + 5 for i in range(len(months))]
            add_nvkt_chart(workbook['C1.4_NVKT'], months, "C1.4 - Tỷ lệ hài lòng", tl_hl_nvkt_cols)
            
            # Biểu đồ SM1-C12 HLL
            print("    ✓ Biểu đồ SM1-C12 HLL theo NVKT...")
            tl_hll_cols = [i * 3 + 5 for i in range(len(months))]
            add_nvkt_chart(workbook['SM1C12_HLL_NVKT'], months, "SM1-C12 - Tỷ lệ hỏng lại", tl_hll_cols)
            
            # Biểu đồ SM4-C11 chi tiết
            print("    ✓ Biểu đồ SM4-C11 chi tiết theo NVKT...")
            tl_dat_cols = [i * 3 + 5 for i in range(len(months))]
            add_nvkt_chart(workbook['SM4C11_ChiTiet_NVKT'], months, "SM4-C11 - Tỷ lệ đạt", tl_dat_cols)
            
            # Biểu đồ SM4-C11 18h
            print("    ✓ Biểu đồ SM4-C11 18h theo NVKT...")
            add_nvkt_chart(workbook['SM4C11_18h_NVKT'], months, "SM4-C11 18h - Tỷ lệ đạt", tl_dat_cols)
        
        print(f"\n✓ Đã xuất báo cáo xu hướng: {output_file}")
        
        # Thống kê
        print(f"\n=== THỐNG KÊ ===")
        print(f"  Số tháng: {len(months)} ({', '.join(months)})")
        print(f"\n  [Theo chỉ tiêu]")
        print(f"  C1.1 theo đơn vị: {len(c11_df)} bản ghi")
        print(f"  C1.2 theo đơn vị: {len(c12_df)} bản ghi")
        print(f"  C1.3 theo đơn vị: {len(c13_df)} bản ghi")
        print(f"  C1.4 theo đơn vị: {len(c14_df)} bản ghi")
        print(f"  C1.4 theo NVKT: {len(c14_nvkt_df)} bản ghi")
        print(f"  SM1-C12 HLL theo NVKT: {len(sm1c12_df)} bản ghi")
        print(f"  SM4-C11 chi tiết theo NVKT: {len(sm4c11_df)} bản ghi")
        print(f"  SM4-C11 18h theo NVKT: {len(sm4c11_18h_df)} bản ghi")
        print(f"\n  [Theo đơn vị/NVKT - Tất cả chỉ số]")
        print(f"  Tổng hợp theo đơn vị (chi tiết): {len(donvi_detail_df)} bản ghi")
        print(f"  Pivot theo đơn vị: {len(donvi_pivot_df)} bản ghi")
        print(f"  Tổng hợp theo NVKT (chi tiết): {len(nvkt_detail_df)} bản ghi")
        print(f"  Pivot theo NVKT: {len(nvkt_pivot_df)} bản ghi")
        
    finally:
        conn.close()
    
    return 0


def main():
    parser = argparse.ArgumentParser(description="Xuất báo cáo xu hướng theo tháng")
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Đường dẫn file Excel đầu ra. Mặc định: downloads/baocao_hanoi/bao_cao_xu_huong_<timestamp>.xlsx"
    )
    args = parser.parse_args()
    
    return export_trend_report(args.output)


if __name__ == "__main__":
    exit(main())
