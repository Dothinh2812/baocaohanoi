# -*- coding: utf-8 -*-
import time
import os
from datetime import datetime
from config import Config


def xac_minh_tam_dung_download(page_baocao):
    """
    Tải báo cáo xác minh tạm dừng từ hệ thống báo cáo VNPT Hà Nội.

    Các bước:
    1. Truy cập URL báo cáo
    2. Chọn đơn vị TTVT Sơn Tây từ dropdown
    3. Mở dropdown dịch vụ (div:nth-child(4)) và tick 2 checkbox (item 5, item 6)
    4. Xóa trắng ô ngày và điền ngày đầu tiên của tháng hiện tại
    5. Click "Xem báo cáo"
    6. Click "Xuất Excel"
    7. Click icon download để tải file
    8. Lưu file với tên tam_dung_xac_minh trong thư mục XAC MINH TAM DUNG

    Args:
        page_baocao: Playwright page object đã đăng nhập
    """

    try:
        print("\n" + "=" * 80)
        print("BẮT ĐẦU TẢI BÁO CÁO XÁC MINH TẠM DỪNG")
        print("=" * 80)

        # Tính ngày đầu tiên của tháng hiện tại
        today = datetime.now()
        first_day = today.replace(day=1)
        first_day_str = first_day.strftime("%d/%m/%Y")
        print(f"Ngày đầu tháng: {first_day_str}")

        # ====================================================================
        # Bước 1: Truy cập URL báo cáo
        # ====================================================================
        url = "https://baocao.hanoi.vnpt.vn/report/report-info?id=267844&menu_id=276199"
        print(f"\n✓ Đang truy cập: {url}")
        page_baocao.goto(url, wait_until="networkidle", timeout=Config.PAGE_LOAD_TIMEOUT)
        time.sleep(3)
        print("✅ Đã tải trang thành công")

        # ====================================================================
        # Bước 2: Click vào dropdown chọn đơn vị
        # ====================================================================
        print("\n✓ Đang click vào dropdown chọn đơn vị...")
        dropdown_btn_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[1]/div/div/div/div/div/div/ngx-dropdown-treeview-select/ngx-dropdown-treeview/div/button"
        dropdown_btn = page_baocao.locator(f"xpath={dropdown_btn_xpath}")
        dropdown_btn.wait_for(state="visible", timeout=10000)
        dropdown_btn.click()
        time.sleep(1)
        print("✅ Đã mở dropdown đơn vị")

        # ====================================================================
        # Bước 3: Tìm input search và nhập "ttvt sơn tây"
        # ====================================================================
        print("\n✓ Đang tìm input search và nhập 'ttvt sơn tây'...")
        search_input_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[1]/div/div/div/div/div/div/ngx-dropdown-treeview-select/ngx-dropdown-treeview/div/div/div/ngx-treeview/div[1]/div[1]/div/input"
        search_input = page_baocao.locator(f"xpath={search_input_xpath}")
        search_input.wait_for(state="visible", timeout=10000)
        search_input.fill("ttvt sơn tây")
        time.sleep(2)
        print("✅ Đã nhập 'ttvt sơn tây' vào ô tìm kiếm")

        # ====================================================================
        # Bước 4: Chọn TTVT Sơn Tây từ kết quả dropdown
        # ====================================================================
        print("\n✓ Đang chọn 'TTVT Sơn Tây'...")
        sontay_option_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[1]/div/div/div/div/div/div/ngx-dropdown-treeview-select/ngx-dropdown-treeview/div/div/div/ngx-treeview/div[2]/div/ngx-treeview-item/div/div[2]/ngx-treeview-item/div/div/span"
        sontay_option = page_baocao.locator(f"xpath={sontay_option_xpath}")
        sontay_option.wait_for(state="visible", timeout=10000)
        sontay_option.click()
        time.sleep(1)
        print("✅ Đã chọn 'TTVT Sơn Tây'")

        # Click vào body để đóng dropdown
        page_baocao.click('body')
        time.sleep(1)
        print("✅ Đã đóng dropdown đơn vị")

        # ====================================================================
        # Bước 5: Mở dropdown dịch vụ (div:nth-child(4))
        # ====================================================================
        print("\n✓ Đang mở dropdown dịch vụ...")
        dichvu_dropdown_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[4]/div/div/div/div/div/div/ngx-dropdown-treeview/div/button"
        dichvu_dropdown = page_baocao.locator(f"xpath={dichvu_dropdown_xpath}")
        dichvu_dropdown.wait_for(state="visible", timeout=10000)
        dichvu_dropdown.click()
        time.sleep(1)
        print("✅ Đã mở dropdown dịch vụ")

        # ====================================================================
        # Bước 6: Tick vào checkbox item 6 (ngx-treeview-item[6])
        # ====================================================================
        print("\n✓ Đang tick checkbox item 6...")
        checkbox6_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[4]/div/div/div/div/div/div/ngx-dropdown-treeview/div/div/div/ngx-treeview/div[2]/div/ngx-treeview-item[6]/div/div[1]/div/input"
        checkbox6 = page_baocao.locator(f"xpath={checkbox6_xpath}")
        checkbox6.wait_for(state="visible", timeout=10000)
        checkbox6.click(force=True)
        time.sleep(1)
        print("✅ Đã tick checkbox item 6")

        # ====================================================================
        # Bước 7: Tick vào checkbox item 5 (ngx-treeview-item[5])
        # ====================================================================
        print("\n✓ Đang tick checkbox item 5...")
        checkbox5_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[1]/div[4]/div/div/div/div/div/div/ngx-dropdown-treeview/div/div/div/ngx-treeview/div[2]/div/ngx-treeview-item[5]/div/div[1]/div/input"
        checkbox5 = page_baocao.locator(f"xpath={checkbox5_xpath}")
        checkbox5.wait_for(state="visible", timeout=10000)
        checkbox5.click(force=True)
        time.sleep(1)
        print("✅ Đã tick checkbox item 5")

        # Đóng dropdown dịch vụ bằng cách click ra ngoài
        page_baocao.keyboard.press("Escape")
        time.sleep(1)

        # ====================================================================
        # Bước 8: Xóa trắng ô input ngày và điền ngày đầu tháng
        # ====================================================================
        print(f"\n✓ Đang điền ngày đầu tháng: {first_day_str}...")
        date_field = page_baocao.locator('//*[@id="mat-input-0"]')
        date_field.wait_for(state="visible", timeout=10000)
        # Xóa trắng ô input
        date_field.click()
        time.sleep(0.5)
        date_field.press("Control+A")
        time.sleep(0.5)
        date_field.fill(first_day_str)
        time.sleep(1)
        print(f"✅ Đã điền ngày: {first_day_str}")

        # Đóng datepicker nếu có
        page_baocao.keyboard.press("Escape")
        time.sleep(1)

        # Đợi backdrop biến mất
        try:
            page_baocao.wait_for_selector(".cdk-overlay-backdrop", state="hidden", timeout=5000)
            print("✅ Datepicker đã đóng")
        except:
            print("✅ Không có backdrop hoặc đã đóng")
        time.sleep(1)

        # ====================================================================
        # Bước 9: Click "Xem báo cáo"
        # ====================================================================
        print("\n✓ Đang click 'Xem báo cáo'...")
        baocao_btn_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[2]/button"
        baocao_btn = page_baocao.locator(f"xpath={baocao_btn_xpath}")
        baocao_btn.wait_for(state="visible", timeout=10000)
        baocao_btn.click()
        time.sleep(2)
        print("✅ Đã click 'Xem báo cáo'")

        # Đợi dữ liệu load
        print("\n✓ Đang đợi dữ liệu load...")
        page_baocao.wait_for_load_state("networkidle", timeout=Config.NETWORK_IDLE_TIMEOUT)
        time.sleep(3)

        # Đợi loading overlay biến mất
        try:
            page_baocao.wait_for_selector("ngx-loading .backdrop", state="hidden", timeout=60000)
            print("✅ Loading overlay đã biến mất")
        except:
            try:
                page_baocao.wait_for_selector(".backdrop.full-screen", state="hidden", timeout=10000)
                print("✅ Loading overlay đã biến mất")
            except:
                print("✅ Không có loading overlay hoặc đã biến mất")

        time.sleep(1)
        print("✅ Dữ liệu đã load xong")

        # ====================================================================
        # Bước 10: Click "Xuất Excel"
        # ====================================================================
        print("\n✓ Đang click 'Xuất Excel'...")
        xuatexcel_btn_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[2]/div/button"
        xuatexcel_btn = page_baocao.locator(f"xpath={xuatexcel_btn_xpath}")
        xuatexcel_btn.wait_for(state="visible", timeout=10000)
        xuatexcel_btn.click()
        time.sleep(2)
        print("✅ Đã click 'Xuất Excel', dropdown đã mở")

        # ====================================================================
        # Bước 11: Click icon download (i[2]) để tải file
        # ====================================================================
        print("\n✓ Đang click icon download để tải file...")
        download_icon_xpath = "/html/body/app-root/app-layout/app-vertical/div[2]/div[2]/div/app-report-info-list/div/div[1]/div[2]/div/div/div[2]/div[2]/div/div/i[2]"
        download_icon = page_baocao.locator(f"xpath={download_icon_xpath}")
        download_icon.wait_for(state="visible", timeout=10000)

        # Đảm bảo thư mục lưu file tồn tại
        download_dir = "XAC MINH TAM DUNG"
        os.makedirs(download_dir, exist_ok=True)

        # Bắt đầu tải file
        with page_baocao.expect_download(timeout=300000) as download_info:
            download_icon.click()
            time.sleep(2)

        download = download_info.value

        # Lấy tên file gốc và extension
        original_filename = download.suggested_filename
        print(f"Tên file gốc: {original_filename}")

        file_extension = os.path.splitext(original_filename)[1]
        current_month = datetime.now().strftime("%m-%Y")
        new_filename = f"tam_dung_xac_minh_thang_{current_month}{file_extension}"

        # Lưu file
        save_path = os.path.join(download_dir, new_filename)
        download.save_as(save_path)

        print(f"✅ Đã tải file về: {save_path}")

        print("\n" + "=" * 80)
        print("✅ HOÀN THÀNH TẢI BÁO CÁO XÁC MINH TẠM DỪNG")
        print("=" * 80)

        return True

    except Exception as e:
        print(f"\n❌ Lỗi khi tải báo cáo xác minh tạm dừng: {e}")
        import traceback
        traceback.print_exc()
        return False


def chuan_hoa_ten_kv(ten_kv):
    """
    Chuẩn hóa cột TEN_KV: lấy tên người sau dấu '-'.
    Ví dụ:
        'PCT1-Nguyễn Mạnh Hùng(PTO)' -> 'Nguyễn Mạnh Hùng'
        'Tản Lĩnh 06 - Chu Văn Hùng' -> 'Chu Văn Hùng'
        'Vạn Thắng 09 - Bùi Văn Biên' -> 'Bùi Văn Biên'
        'Sơn Lộc 5 - Đỗ Huy Thông' -> 'Đỗ Huy Thông'
    """
    import re
    if not ten_kv or not isinstance(ten_kv, str):
        return ten_kv

    # Tách phần sau dấu '-' (lấy phần cuối cùng nếu có nhiều dấu '-')
    if '-' in ten_kv:
        name_part = ten_kv.split('-', 1)[1].strip()
    else:
        name_part = ten_kv.strip()

    # Loại bỏ phần trong ngoặc đơn, ví dụ: (PTO), (TGG)
    name_part = re.sub(r'\([^)]*\)', '', name_part).strip()

    return name_part


def xac_minh_tam_dung_process():
    """
    Xử lý báo cáo xác minh tạm dừng:
    - Đọc file tam_dung_xac_minh_thang_mm-yyyy.xlsx
    - Tạo sheet chi_tiet_tam_dung với các cột yêu cầu
    - Chuẩn hóa cột TEN_KV
    """
    import openpyxl

    current_month = datetime.now().strftime("%m-%Y")
    filename = f"tam_dung_xac_minh_thang_{current_month}.xlsx"
    file_path = os.path.join("XAC MINH TAM DUNG", filename)
    print(f"\n{'='*80}")
    print("BẮT ĐẦU XỬ LÝ BÁO CÁO XÁC MINH TẠM DỪNG")
    print(f"{'='*80}")
    print(f"Đang đọc file: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path)
        # Luôn đọc từ sheet gốc (Sheet1), không phải sheet đã xử lý
        ws = wb['Sheet1']
        print(f"✅ Đã đọc file, sheet: {ws.title}, {ws.max_row} dòng, {ws.max_column} cột")

        # Đọc header (dòng 1)
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")

        # Các cột cần lấy (theo thứ tự yêu cầu, không tính STT)
        target_columns = [
            'DOIVT', 'TEN_KV', 'ND_HUY_NVKT', 'MA_TB', 'TEN_DVVT_HNI',
            'TEN_KIEULD', 'NGAYLAP_HD', 'NGAY_THUCHIEN', 'TEN_TB', 'DIACHI_LAPDAT',
            'LYDOHUY', 'GHICHU', 'SO_DT', 'ND_HUY_GDV'
        ]

        # Tìm index của từng cột trong file gốc
        col_indices = {}
        for col_name in target_columns:
            if col_name in headers:
                col_indices[col_name] = headers.index(col_name)
            else:
                print(f"⚠️ Không tìm thấy cột: {col_name}")
                col_indices[col_name] = None

        # Thu thập tất cả dòng dữ liệu
        all_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_idx]]
            if not any(row_data):
                continue

            extracted = {}
            for col_name in target_columns:
                src_idx = col_indices.get(col_name)
                value = row_data[src_idx] if src_idx is not None else None
                if col_name == 'TEN_KV':
                    value = chuan_hoa_ten_kv(value)
                extracted[col_name] = value
            all_rows.append(extracted)

        # Sắp xếp theo NGAYLAP_HD mới nhất lên đầu
        all_rows.sort(key=lambda r: str(r.get('NGAYLAP_HD') or ''), reverse=True)
        print(f"✅ Đã sắp xếp {len(all_rows)} dòng theo NGAYLAP_HD giảm dần")

        # Tạo sheet mới hoặc ghi đè
        sheet_name = "chi_tiet_tam_dung"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"✅ Đã xóa sheet cũ: {sheet_name}")

        ws_new = wb.create_sheet(sheet_name)
        print(f"✅ Đã tạo sheet mới: {sheet_name}")

        # Ghi header
        new_headers = ['STT'] + target_columns
        for col_idx, header in enumerate(new_headers, 1):
            ws_new.cell(row=1, column=col_idx, value=header)

        # Ghi dữ liệu
        for stt, row_dict in enumerate(all_rows, 1):
            ws_new.cell(row=stt + 1, column=1, value=stt)
            for col_offset, col_name in enumerate(target_columns, 2):
                ws_new.cell(row=stt + 1, column=col_offset, value=row_dict.get(col_name))

        print(f"✅ Đã ghi {len(all_rows)} dòng dữ liệu")

        # Lưu file
        wb.save(file_path)
        print(f"✅ Đã lưu file: {file_path}")

        print(f"\n{'='*80}")
        print("✅ HOÀN THÀNH XỬ LÝ BÁO CÁO XÁC MINH TẠM DỪNG")
        print(f"{'='*80}")

        return True

    except Exception as e:
        print(f"\n❌ Lỗi khi xử lý báo cáo: {e}")
        import traceback
        traceback.print_exc()
        return False


def send_warning_tam_dung_xac_minh():
    """
    Gửi cảnh báo các bản ghi tạm dừng mới đến nhóm Zalo + Telegram tương ứng theo DOIVT.
    - Đọc sheet chi_tiet_tam_dung từ file tam_dung_xac_minh_thang_MM-YYYY.xlsx
    - So sánh với log để chỉ gửi bản ghi mới (MA_TB chưa gửi)
    - Gộp bản ghi theo DOIVT, gửi 1 tin nhắn/nhóm
    """
    import pandas as pd
    import requests

    print("\n" + "=" * 80)
    print("GỬI CẢNH BÁO TẠM DỪNG XÁC MINH QUA ZALO + TELEGRAM")
    print("=" * 80)

    # Import Telegram
    TELEGRAM_TOKEN = None
    try:
        import sys
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'config_for_send_zalo'))
        from send_tele import TELEGRAM_TOKEN, send_message as send_telegram_message
    except ImportError:
        print("⚠️ Không thể import module send_tele. Chỉ gửi Zalo.")

    current_month = datetime.now().strftime("%m-%Y")
    filename = f"tam_dung_xac_minh_thang_{current_month}.xlsx"
    file_path = os.path.join("XAC MINH TAM DUNG", filename)
    webhook_url_zalo = os.getenv("WEBHOOK_TEXT_URL", "")

    # Mapping DOIVT -> thread_id Zalo, telegram_chat_id
    doi_vt_mapping = {
        'Tổ Kỹ thuật Địa bàn Phúc Thọ': {
            'thread_id': '3142012656522650111',
            'telegram_chat_id': '-4616062001',
            'display_name': 'Phúc Thọ'
        },
        'Tổ Kỹ thuật Địa bàn Sơn Tây': {
            'thread_id': '4761925886931896176',
            'telegram_chat_id': '-4654883926',
            'display_name': 'Sơn Tây'
        },
        'Tổ Kỹ thuật Địa bàn Quảng Oai': {
            'thread_id': '7968537750365285360',
            'telegram_chat_id': '-4734554771',
            'display_name': 'Quảng Oai'
        },
        'Tổ Kỹ thuật Địa bàn Suối Hai': {
            'thread_id': '6052111621047664',
            'telegram_chat_id': '-4607586268',
            'display_name': 'Suối Hai'
        },
        'Tổ Kỹ thuật Địa bàn Suối hai': {
            'thread_id': '6052111621047664',
            'telegram_chat_id': '-4607586268',
            'display_name': 'Suối Hai'
        },
    }

    # Log config
    log_dir = "log_message"
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, "tam_dung_xac_minh_log.xlsx")

    try:
        if not os.path.exists(file_path):
            print(f"❌ Không tìm thấy file: {file_path}")
            return False

        df = pd.read_excel(file_path, sheet_name='chi_tiet_tam_dung')
        if df.empty:
            print("ℹ️ Sheet chi_tiet_tam_dung không có dữ liệu")
            return True

        print(f"📋 Tổng số bản ghi trong sheet: {len(df)}")

        # Load log để lọc bản ghi đã gửi
        sent_ma_tb = set()
        if os.path.exists(log_file):
            try:
                existing_log = pd.read_excel(log_file)
                sent_ma_tb = set(existing_log['MA_TB'].astype(str).tolist())
                print(f"📂 Đã load log: {len(sent_ma_tb)} MA_TB đã gửi trước đó")
            except Exception as e:
                print(f"⚠️ Lỗi khi đọc log: {e}")

        # Lọc bản ghi mới
        df['MA_TB_str'] = df['MA_TB'].astype(str)
        df_new = df[~df['MA_TB_str'].isin(sent_ma_tb)].copy()

        if df_new.empty:
            print("ℹ️ Không có bản ghi mới nào cần gửi")
            return True

        print(f"🔍 Tìm thấy {len(df_new)} bản ghi mới cần gửi")

        current_time = datetime.now()
        log_data = []
        total_sent_zalo = 0
        total_sent_telegram = 0
        total_failed_zalo = 0
        total_failed_telegram = 0

        # Nhóm theo DOIVT
        grouped = df_new.groupby('DOIVT')

        for doi_vt, group in grouped:
            mapping_info = doi_vt_mapping.get(doi_vt)
            if not mapping_info:
                print(f"⚠️ Không tìm thấy mapping cho '{doi_vt}', bỏ qua")
                continue

            thread_id = mapping_info['thread_id']
            telegram_chat_id = mapping_info['telegram_chat_id']
            display_name = mapping_info['display_name']

            print(f"\n📤 Gửi cảnh báo cho Tổ {display_name}: {len(group)} bản ghi")

            # Tạo tin nhắn gộp
            lines = []
            for stt, (_, row) in enumerate(group.iterrows(), 1):
                ma_tb = str(row.get('MA_TB', 'N/A'))
                ten_dvvt = str(row.get('TEN_DVVT_HNI', 'N/A'))
                ten_kieuld = str(row.get('TEN_KIEULD', '')) if pd.notna(row.get('TEN_KIEULD')) else ''
                ten_kv = str(row.get('TEN_KV', '')) if pd.notna(row.get('TEN_KV')) else ''
                ghichu = str(row.get('GHICHU', '')) if pd.notna(row.get('GHICHU')) else ''
                so_dt = str(row.get('SO_DT', '')) if pd.notna(row.get('SO_DT')) else ''
                ngaylap_hd = row.get('NGAYLAP_HD', '')
                if pd.notna(ngaylap_hd):
                    if isinstance(ngaylap_hd, datetime):
                        ngaylap_hd = ngaylap_hd.strftime('%d/%m/%Y')
                    else:
                        ngaylap_hd = str(ngaylap_hd)
                else:
                    ngaylap_hd = 'N/A'
                line = (
                    f"{stt}. {ma_tb} | {ten_dvvt} | {ngaylap_hd}\n"
                    f"   Kiểu LĐ: {ten_kieuld} | KV: {ten_kv}\n"
                    f"   SĐT: {so_dt}"
                )
                if ghichu:
                    line += f" | Ghi chú: {ghichu}"
                lines.append(line)

            message = (
                f"🔔 Cảnh báo TB tạm dừng mới - Tổ {display_name}\n"
                f"Thời gian: {current_time.strftime('%d/%m/%Y %H:%M')}\n\n"
                + "\n".join(lines)
                + f"\n\nTổng: {len(group)} thuê bao"
            )

            # Gửi Zalo
            zalo_success = False
            try:
                data = {
                    'threadID': thread_id,
                    'message': message
                }
                if not webhook_url_zalo:
                    raise ValueError("Thiếu WEBHOOK_TEXT_URL trong biến môi trường")
                response = requests.get(webhook_url_zalo, json=data, timeout=10)
                if response.status_code == 200:
                    total_sent_zalo += 1
                    zalo_success = True
                    print(f"  ✅ [Zalo] Gửi thành công tới {display_name}")
                else:
                    total_failed_zalo += 1
                    print(f"  ❌ [Zalo] Gửi thất bại (Status: {response.status_code})")
            except Exception as e:
                total_failed_zalo += 1
                print(f"  ❌ [Zalo] Lỗi: {e}")

            import time as time_module
            time_module.sleep(0.5)

            # Gửi Telegram
            telegram_success = False
            if telegram_chat_id and TELEGRAM_TOKEN:
                try:
                    tele_url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
                    tele_data = {
                        "chat_id": telegram_chat_id,
                        "text": message,
                        "parse_mode": "HTML"
                    }
                    response = requests.post(tele_url, data=tele_data, timeout=10)
                    if response.status_code == 200:
                        total_sent_telegram += 1
                        telegram_success = True
                        print(f"  ✅ [Telegram] Gửi thành công tới {display_name}")
                    else:
                        total_failed_telegram += 1
                        print(f"  ❌ [Telegram] Gửi thất bại (Status: {response.status_code})")
                except Exception as e:
                    total_failed_telegram += 1
                    print(f"  ❌ [Telegram] Lỗi: {e}")

            # Ghi log cho từng bản ghi trong nhóm nếu gửi thành công
            if zalo_success or telegram_success:
                for _, row in group.iterrows():
                    log_data.append({
                        'MA_TB': str(row.get('MA_TB', '')),
                        'DOIVT': str(doi_vt),
                        'TEN_DVVT_HNI': str(row.get('TEN_DVVT_HNI', '')),
                        'NGAYLAP_HD': row.get('NGAYLAP_HD', ''),
                        'Thời gian gửi': current_time,
                        'Zalo': 'Thành công' if zalo_success else 'Thất bại',
                        'Telegram': 'Thành công' if telegram_success else 'Thất bại'
                    })

            time_module.sleep(1)

        # Cập nhật log file
        if log_data:
            new_log_df = pd.DataFrame(log_data)
            if os.path.exists(log_file):
                try:
                    old_log_df = pd.read_excel(log_file)
                    combined_df = pd.concat([old_log_df, new_log_df], ignore_index=True)
                except Exception:
                    combined_df = new_log_df
            else:
                combined_df = new_log_df
            combined_df.to_excel(log_file, index=False)
            print(f"\n📝 Đã cập nhật log: {log_file} ({len(log_data)} bản ghi mới)")

        print(f"\n📊 Kết quả: Zalo {total_sent_zalo} OK / {total_failed_zalo} lỗi | "
              f"Telegram {total_sent_telegram} OK / {total_failed_telegram} lỗi")
        print("=" * 80)
        return True

    except Exception as e:
        print(f"\n❌ Lỗi khi gửi cảnh báo tạm dừng: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    from login import login_baocao_hanoi

    page_baocao = None
    browser_baocao = None
    playwright_baocao = None

    try:
        # Đăng nhập
        page_baocao, browser_baocao, playwright_baocao = login_baocao_hanoi()

        # Tải báo cáo xác minh tạm dừng
        xac_minh_tam_dung_download(page_baocao)

    except Exception as e:
        print(f"\n❌ Lỗi: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # Đóng browser và playwright
        if browser_baocao:
            browser_baocao.close()
        if playwright_baocao:
            playwright_baocao.stop()
        print("\n✅ Đã đóng browser.")

    # Xử lý báo cáo
    xac_minh_tam_dung_process()

    # Gửi cảnh báo tạm dừng qua Zalo + Telegram
    send_warning_tam_dung_xac_minh()
